"""
Stage 1 (v2) — build a state's SNAP QC rules workbook: the delivery rule
lists, measured against the state's QC cases, with no tuning of any kind.

No search or tuning runs here (decision 2026-08-16). The original v1 builder's
unconstrained in-sample search was the winner's curse with no brake on it
(retired to custom_one_off/legacy_dashboard); its guarded tiered replacement
in tuning.py never fires at public-QC sample sizes and cannot re-run inside
Excel on a state's pasted data, so the workbook no longer carries it. The
workbook measures the delivered lists; it does not modify them. tuning.py
stays in the package for pipeline-side use on a state's internal data
(methods/tuning_principles.md).

Sheets produced (names from workbook_layout.py):
  Step 3. Select Rules   the effective rule list (blended delivery list after
                         the rule_selection.py transform: count_divisible_by_100
                         rules dropped, trivially-true bbce_state_i conjuncts
                         stripped, buffer rules promoted to refill capacity),
                         sorted by error dollars caught on the state frame
  See cases flagged by a rule  pick a rule, list the cases it flags (errors-only
                               or all flagged, via a toggle)
  Step 2. Import Testing Data  input-contract headers (left, filled by
                         make_input_workbook) + reconstructed state QC case data
  Dashboard          one threshold block per rule + PR chart   (hidden engine)
  Grid Search        bracket-bounded threshold search          (hidden engine)
  RuleFlags          case x rule hit matrices                  (hidden engine)

Pick the state with the SNAP_STATE environment variable (default WA); add new
states in states.py.  Run through make_state.py so the calc chain gets dropped
and the result is verified in Excel. Rule selection is plain TRUE/FALSE text in
the Include? column (native checkboxes were dropped 2026-08-18 for
compatibility with pre-365 Excel).
"""
import os
import re
import math
import shutil
import subprocess
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
import numpy as np
import pandas as pd

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG — state settings live in states.py; nothing here needs editing
# ══════════════════════════════════════════════════════════════════════════════
import states as STATE_REGISTRY
import tuning
import rule_selection
from workbook_layout import (DATA_SHEET, BLENDED_SHEET, EXPORT_SHEET,
                             VIEWER_SHEET, RAW_COLS, RAW_OFF, qref)

DQ = qref(DATA_SHEET)                # the Data sheet as written in formulas

PKG          = os.path.dirname(os.path.abspath(__file__))
BASE         = os.path.dirname(PKG)                # project root
BUILD_DIR    = os.path.join(PKG, '.build')         # stage-to-stage handoff files
_cfg         = STATE_REGISTRY.get(os.environ.get('SNAP_STATE', 'WA'))

STATE_NAME   = _cfg['name']
STATE_ABBR   = _cfg['abbr']
STATE_FIPS   = _cfg['fips']          # informational; nothing reads the .sav
                                     # files any more (2026-08-16: the demo
                                     # carries reconstructed frame values)
FY_LABEL     = _cfg['fy_label']


def _find_repo(start):
    """Locate the snap_qc checkout by walking up and looking for the two files
    v2 actually needs: the munged frame and the miner's helpers."""
    cands = [os.environ['SNAP_REPO']] if os.environ.get('SNAP_REPO') else []
    d = start
    for _ in range(6):
        cands += [d, os.path.join(d, 'snap_qc')]
        d = os.path.dirname(d)
    for c in cands:
        if (os.path.isfile(os.path.join(c, 'reg_model_data.rds'))
                and os.path.isfile(os.path.join(c, 'rule_mining_helpers.R'))):
            return os.path.abspath(c)
    raise SystemExit('cannot find the snap_qc checkout (needs reg_model_data.rds, '
                     'written by the munging script, and rule_mining_helpers.R); '
                     'set SNAP_REPO')


REPO = _find_repo(PKG)


def _find_delivery(rel):
    """The delivery list, wherever the state keeps it. Falls back to the repo's
    tracked state_delivery_lists/ copy, which is the same public artifact."""
    base = os.path.basename(rel).replace('_core.csv', '.csv')
    for c in (os.path.join(BASE, rel), os.path.join(REPO, rel),
              os.path.join(PKG, rel), os.path.join(REPO, 'state_delivery_lists', base)):
        if os.path.isfile(c):
            return os.path.abspath(c)
    raise SystemExit(f'delivery list not found: {rel}')


DELIVERY_CSV = _find_delivery(_cfg['delivery_csv'])
STATE_DIR    = os.environ.get('SNAP_OUT_DIR') or os.path.join(
                   PKG, 'state_workbooks', STATE_ABBR)
os.makedirs(STATE_DIR, exist_ok=True)
CASES_CSV    = os.path.join(STATE_DIR, f'{STATE_ABBR.lower()}_cases.csv')
OUT          = os.path.join(STATE_DIR, f'snap_qc_dashboard_{STATE_ABBR}.xlsx')

MAX_ROW      = 3000            # Data-sheet formula range ceiling
MAXG         = 2500            # grid rows allocated per stratum
STRATA       = ['1', '2-3', '4+']
NSLOTS       = 4               # condition slots in the grid engine

# The tuning contract. Defaults live in tuning.TuningConfig (and are documented
# in methods/tuning_principles.md); states.py may override per deployment.
TCFG         = tuning.TuningConfig(**getattr(STATE_REGISTRY, 'TUNING', {}))
PREC_FLOOR   = TCFG.min_variant_precision   # applied to a BOUND, not raw precision
MIN_FLAGGED  = TCFG.min_support             # the non-negotiable support floor
RECALL_FLOOR = 0.0                          # v1 device; the objective is dollar-max
                                            # among qualifiers, so no floor is needed

# ══════════════════════════════════════════════════════════════════════════════
# 1. LOAD THE STATE DATASET from the munged modelling frame
#
# The delivery rules were mined on reg_model_data.rds, whose variables the
# munging script restores to their pre-QC-process values
# (1_data_munging_..._for_using_public_qc_data.R, correct_variables <- TRUE).
# Scoring them against anything else compares thresholds to a different scale:
# rebuilding these features from the raw .sav files in Python, as v1 does,
# reproduces the row universe exactly but leaves 21 of Washington's 114 core
# rules never firing at all. So this reads the munged frame instead, exported by
# export_state_frame.R with the miner's own prep_features() applied.
#
# The export is cached per state under .frames/. Set SNAP_REFRESH_FRAME=1 to
# rebuild it (needed after the munging script is re-run).
# ══════════════════════════════════════════════════════════════════════════════
FRAME_DIR = os.path.join(PKG, '.frames')
FRAME_CSV = os.path.join(FRAME_DIR, f'{STATE_ABBR.lower()}_frame.csv')
YEARS = [str(y) for y in _cfg.get('years', (2022, 2023, 2024))]


def _rscript():
    cands = ([os.environ['RSCRIPT']] if os.environ.get('RSCRIPT') else []) + [
        os.path.join('C:' + os.sep, 'Program Files', 'R', 'R-4.5.1', 'bin', 'Rscript.exe'),
        'Rscript']
    for c in cands:
        if os.path.isabs(c):
            if os.path.isfile(c):
                return c
        elif shutil.which(c):
            return shutil.which(c)
    raise SystemExit('Rscript not found; set RSCRIPT to its full path. The state '
                     'frame comes from reg_model_data.rds via export_state_frame.R.')


def _export_frame():
    cmd = [_rscript(), os.path.join(PKG, 'export_state_frame.R'),
           '--state', STATE_NAME, '--out', FRAME_CSV,
           '--years', ','.join(YEARS), '--repo', REPO]
    print('$ ' + ' '.join(cmd))
    if subprocess.run(cmd, cwd=REPO).returncode != 0:
        raise SystemExit('export_state_frame.R failed (see the output above)')


if os.environ.get('SNAP_REFRESH_FRAME') == '1' or not os.path.isfile(FRAME_CSV):
    _export_frame()
else:
    print(f'using cached frame {FRAME_CSV} (SNAP_REFRESH_FRAME=1 to rebuild)')

df = pd.read_csv(FRAME_CSV)
print(f'{STATE_ABBR}: {len(df)} rows from the munged frame, FY{"+".join(YEARS)}')

# The pipeline's error convention, carried over by the exporter: an error is
# over_threshold != 0, and the dashboard's sheets read a 0/1 `over_threshold`.
df['over_threshold'] = pd.to_numeric(df['is_error'], errors='coerce').fillna(0).astype(int)
df['total_error_amount'] = pd.to_numeric(df['total_error_amount'],
                                         errors='coerce').fillna(0).abs().round(0)
df['fiscal_year'] = pd.to_numeric(df['fiscal_year'], errors='coerce').astype(int)
df['hh_group'] = df['hh_group'].astype(str)
df['hh_size_raw'] = pd.to_numeric(df['hh_size_raw'], errors='coerce')
bad = df.hh_group.isin(['nan', 'None'])
if bad.any():
    print(f'dropping {int(bad.sum())} rows with no household-size stratum')
    df = df[~bad].reset_index(drop=True)
print(f'error rate: {df.over_threshold.mean():.3f} '
      f'({int(df.over_threshold.sum())} errors)')

df.to_csv(CASES_CSV, index=False)

# The Data-sheet formula ranges must cover every case: large states (CA, TX, FL)
# carry more public QC rows than the 3,000-row default.
MAX_ROW = max(MAX_ROW, len(df) + 101)

# ══════════════════════════════════════════════════════════════════════════════
# 2. PARSE THE DELIVERY RULES
#
# One rule list ships per state: the BLENDED delivery list (the state's own
# mined rules merged into the national pool), passed through the
# rule_selection.py transform (count_divisible_by_100 rules dropped, trivially
# true bbce_state_i conjuncts stripped, buffer rules promoted to refill the
# freed capacity) and sorted by error dollars caught on the state frame.
# ══════════════════════════════════════════════════════════════════════════════

# Per-rule characterization carried through from the delivery CSVs: what the
# rule catches as mined on national data. Fixed context on the rules tabs —
# deliberately NOT recomputed from pasted data.
# (csv column, tab header, number format, column width)
CHAR_COLS = [
    ('pool',                   'Rule source pool',                 '@',      12),
    ('precision_train',        'Train precision (natl.)',          '0.0%',   12),
    ('precision_train_lcb',    'Train precision lower bound',      '0.0%',   13),
    ('dollars_per_flag_train', 'Train error $ per flagged case',   '$#,##0', 13),
    ('n_error_cases_national', 'Natl. error cases behind rule',    '#,##0',  13),
    ('element_groups_to_75',   'Error elements caught (to 75%)',   '@',      46),
    ('nature_groups_to_75',    'Error natures caught (to 75%)',    '@',      46),
    ('share_overissuance',     'Share overissuance',               '0%',     12),
    ('cause_agency',           'Share agency-caused',              '0%',     12),
    ('found_in_case_record',   'Share discovered in case file',    '0%',     13),
    ('timing_at_certification','Share at certification',           '0%',     12),
]

# ── plain-English rendering of a rule's conditions ────────────────────────────
# The 19-variable vocabulary, phrased for program staff. Indicators render as
# the phrase alone (or its negation); paired bounds on one variable collapse
# to a range. The exact machine expression stays in the tabs' last column.
PLAIN_VARS = {
    'HH_size_n':                  ('household size', 'n'),
    'children_i':                 ('children present', 'i'),
    'elderly_disabled_i':         ('elderly or disabled member', 'i'),
    'total_deductions_by_hh_size': ('total deductions per person', '$'),
    'expedited_i':                ('expedited service', 'i'),
    'bbce_state_i':               ('rule applies to BBCE states', 'i'),
    'rawben_rel_max':             ('benefit relative to the maximum', 'n'),
    'medical_deductions':         ('medical deduction', '$'),
    'shelter_expenses_by_hh_size': ('shelter costs per person', '$'),
    'utilities':                  ('utility costs', '$'),
    'married':                    ('spouse present', 'i'),
    'homeless':                   ('homeless', 'i'),
    'earned_by_hh_size':          ('earned income per person', '$'),
    'unearned_by_hh_size':        ('unearned income per person', '$'),
    'gross_by_hh_size':           ('gross income per person', '$'),
    'percent_abawd':              ('share of members with ABAWD status', 'n'),
    'unc_rawben_rel_max':         ('uncapped benefit relative to the maximum', 'n'),
    'months_since_cert_n':        ('months since certification', 'n'),
    'count_divisible_by_100':     ('number of round-$100 amounts', 'n'),
}


# negation phrasings that read better than 'not <phrase>'
PLAIN_NEG = {'bbce_state_i': 'rule applies to non-BBCE states'}


def _plain_val(v, kind):
    s = f'{v:g}'
    return f'${s}' if kind == '$' else s


def render_plain(conds):
    """'unc_rawben_rel_max > 0.999 & unc_rawben_rel_max <= 1' ->
    'uncapped benefit relative to the maximum between 0.999 and 1'."""
    by_var, order = {}, []
    for c in conds:
        if c['var'] not in by_var:
            by_var[c['var']] = []
            order.append(c['var'])
        by_var[c['var']].append(c)
    parts = []
    for v in order:
        phrase, kind = PLAIN_VARS.get(v, (v, 'n'))
        cs = by_var[v]
        if kind == 'i' and len(cs) == 1:
            up = cs[0]['op'] in ('>=', '>')
            thr = cs[0]['thr']
            yes = (up and thr > 0) or (not up and thr >= 1)
            parts.append(phrase if yes else PLAIN_NEG.get(v, f'not {phrase}'))
            continue
        los = [c for c in cs if c['op'] in ('>', '>=')]
        his = [c for c in cs if c['op'] in ('<', '<=')]
        if len(cs) == 2 and los and his:
            parts.append(f'{phrase} between {_plain_val(los[0]["thr"], kind)} '
                         f'and {_plain_val(his[0]["thr"], kind)}')
            continue
        for c in cs:
            word = {'>=': 'at least', '>': 'over',
                    '<=': 'at most', '<': 'under'}[c['op']]
            parts.append(f'{phrase} {word} {_plain_val(c["thr"], kind)}')
    return '; '.join(parts)


os.makedirs(BUILD_DIR, exist_ok=True)
RULES = rule_selection.effective_rules(
    DELIVERY_CSV, df, char_keys=[k for k, _, _, _ in CHAR_COLS],
    out_csv=os.path.join(BUILD_DIR, f'effective_rules_{STATE_ABBR}.csv'))
print('effective rules implemented:', len(RULES))

RULE_VARS = sorted({c['var'] for r in RULES for c in r['conds']})
for v in RULE_VARS:
    assert v in df.columns, f'missing variable: {v}'

# snapped_grid steps by variable type (dollar 50, ratio 0.05, months/counts 1)
RATIO_VARS = {'rawben_rel_max', 'unc_rawben_rel_max', 'percent_abawd'}
UNIT_VARS  = {'months_since_cert_n', 'HH_size_n', 'count_divisible_by_100', 'cat_elig',
              'bbce_state_i', 'expedited_i', 'elderly_disabled_i', 'married',
              'children_i', 'homeless'}
def base_step(v):
    if v in RATIO_VARS: return 0.05
    if v in UNIT_VARS:  return 1.0
    return 50.0

# ══════════════════════════════════════════════════════════════════════════════
# 3. WORKBOOK SCAFFOLDING
# ══════════════════════════════════════════════════════════════════════════════
DCOLS = ['fiscal_year', 'hh_size_raw', 'hh_group'] + RULE_VARS + \
        ['over_threshold', 'total_error_amount']
# the raw input contract occupies columns 1..RAW_OFF (see workbook_layout);
# the feature block sits to its right, so every feature reference is offset
FEAT0   = get_column_letter(RAW_OFF + 1)
LASTCOL = get_column_letter(RAW_OFF + len(DCOLS))

def dc(varname): return get_column_letter(RAW_OFF + DCOLS.index(varname) + 1)
def dr(varname):
    c = dc(varname)
    return f'{DQ}!${c}$2:${c}${MAX_ROW}'

YELLOW     = PatternFill('solid', fgColor='FFFF99')
BLUE_LIGHT = PatternFill('solid', fgColor='BDD7EE')
BLUE_DARK  = PatternFill('solid', fgColor='2F5496')
GRAY       = PatternFill('solid', fgColor='F2F2F2')
GREEN      = PatternFill('solid', fgColor='E2EFDA')
WHITE      = PatternFill('solid', fgColor='FFFFFF')
FONT = 'Calibri'
def bold_font(size=11, color='000000'): return Font(name=FONT, bold=True, size=size, color=color)
center = Alignment(horizontal='center', vertical='center')
left   = Alignment(horizontal='left',   vertical='center')

def thin():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value=None, formula=None, fill=None, font=None,
             align=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col)
    cell.value = formula if formula else value
    if fill:   cell.fill = fill
    if font:   cell.font = font
    if align:  cell.alignment = align
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell

def merge(ws, r1, c1, r2, c2, value=None, fill=None, font=None, align=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    if value is not None: cell.value = value
    if fill:  cell.fill = fill
    if font:  cell.font = font
    if align: cell.alignment = align
    return cell

wb = openpyxl.Workbook()

# ── Data sheet ────────────────────────────────────────────────────────────────
ws_data = wb.create_sheet(DATA_SHEET)
dfx = df[DCOLS]
# the input contract's headers on the LEFT; make_input_workbook fills the
# values. Light yellow = the workbook-wide "yellow means interactive"
# convention (revision 2026-08-18).
INPUT_YELLOW = PatternFill('solid', fgColor='FFFF99')
for ci, col in enumerate(RAW_COLS, 1):
    c = ws_data.cell(row=1, column=ci, value=col)
    c.fill = INPUT_YELLOW
    c.font = Font(name=FONT, bold=True)
    c.alignment = center
for ci, col in enumerate(DCOLS, 1):
    c = ws_data.cell(row=1, column=RAW_OFF + ci, value=col)
    c.fill = BLUE_DARK
    c.font = Font(name=FONT, bold=True, color='FFFFFF')
    c.alignment = center
for ri, row in enumerate(dfx.itertuples(index=False), 2):
    for ci, val in enumerate(row, 1):
        ws_data.cell(row=ri, column=RAW_OFF + ci, value=val)
ws_data.freeze_panes = 'B2'

# ══════════════════════════════════════════════════════════════════════════════
# 4. DASHBOARD — one tuning block per delivery rule
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Dashboard'
ws.sheet_view.showGridLines = False
for col_letter, width in {'A':34,'B':8,'C':16,'D':14,'E':3,'F':18,'G':13,'H':13,'I':3,
                          'J':9,'K':7,'L':12,'M':12,'N':12,'O':12,'P':12,'Q':10}.items():
    ws.column_dimensions[col_letter].width = width

merge(ws,1,1,1,9,
      value=f'SNAP QC Inclusion Rule Tuner — {STATE_NAME} (FY2022–2024) — '
            f'{len(RULES)} core delivery rules',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=16,color='FFFFFF'), align=center)
ws.row_dimensions[1].height = 32

# stratum summary (referenced by every rule block)
merge(ws,3,1,3,9, value='Dataset Summary by Household Size',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
for col, txt in [(1,'HH size'),(2,''),(3,'Cases'),(4,'Errors'),(6,'Error $'),(7,'Base error rate')]:
    if txt:
        set_cell(ws,4,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())
S_ROWS = {}   # stratum -> summary row
for s, lbl in enumerate(STRATA):
    r = 5 + s
    S_ROWS[lbl] = r
    set_cell(ws,r,1,lbl, font=bold_font(), align=center, border=thin())
    set_cell(ws,r,3,formula=f'=COUNTIF({dr("hh_group")},"{lbl}")',
             fill=BLUE_LIGHT, align=center, border=thin(), number_format='#,##0')
    set_cell(ws,r,4,formula=f'=COUNTIFS({dr("hh_group")},"{lbl}",{dr("over_threshold")},1)',
             fill=BLUE_LIGHT, align=center, border=thin(), number_format='#,##0')
    set_cell(ws,r,6,formula=f'=SUMIFS({dr("total_error_amount")},{dr("hh_group")},"{lbl}",{dr("over_threshold")},1)',
             fill=BLUE_LIGHT, align=center, border=thin(), number_format='$#,##0')
    set_cell(ws,r,7,formula=f'=D{r}/C{r}', fill=BLUE_LIGHT, align=center,
             border=thin(), number_format='0.0%')

def rule_countifs(rule, thr_cells, add_error=False):
    parts = [f'{dr("hh_group")},"{rule["hh"]}"']
    for c, tc in zip(rule['conds'], thr_cells):
        parts.append(f'{dr(c["var"])},"{c["op"]}"&{tc}')
    if add_error:
        parts.append(f'{dr("over_threshold")},1')
    return f'COUNTIFS({",".join(parts)})'

def rule_sumifs(rule, thr_cells):
    parts = [f'{dr("total_error_amount")},{dr("hh_group")},"{rule["hh"]}",{dr("over_threshold")},1']
    for c, tc in zip(rule['conds'], thr_cells):
        parts.append(f'{dr(c["var"])},"{c["op"]}"&{tc}')
    return f'SUMIFS({",".join(parts)})'

METRIC_NAMES = ['n_flagged','errors caught','precision','recall (count)','dollar recall','lift']
METRIC_FMTS  = ['#,##0','#,##0','0.0%','0.0%','0.0%','0.00']
BLOCK_HEIGHT = 10
BASE_ROW = 9
summary_cells = []   # per rule: dict of metric cell refs

for bi, rule in enumerate(RULES):
    r0 = BASE_ROW + bi * BLOCK_HEIGHT
    merge(ws,r0,1,r0,9,
          value=(f'Rule {rule["num"]}  │  HH size {rule["hh"]}  │  {rule["engine"]}, '
                 f'{rule["frame"]}  │  train precision {rule["prec_train"]:.1%} '
                 f'(LCB {rule["prec_lcb"]:.1%})'),
          fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=11,color='FFFFFF'), align=left)
    r1 = r0 + 1
    for col, txt in [(1,'Variable'),(2,'Op'),(3,'Threshold ← EDIT'),(4,'Original'),
                     (6,'Metric'),(7,'This state'),(8,'Train (natl.)')]:
        set_cell(ws,r1,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())

    thr_cells = []
    for ci, cond in enumerate(rule['conds']):
        rc = r0 + 2 + ci
        set_cell(ws,rc,1,cond['var'], align=left, border=thin())
        set_cell(ws,rc,2,cond['op'], align=center, border=thin())
        set_cell(ws,rc,3,cond['thr'], fill=YELLOW, align=center, border=thin(), font=bold_font())
        set_cell(ws,rc,4,cond['thr'], align=center, border=thin(),
                 font=Font(name=FONT,color='808080'))
        thr_cells.append(f'C{rc}')

    srow = S_ROWS[rule['hh']]
    flag_f, err_f = rule_countifs(rule, thr_cells), rule_countifs(rule, thr_cells, True)
    dol_f = rule_sumifs(rule, thr_cells)
    formulas = [
        f'={flag_f}',
        f'={err_f}',
        f'=IFERROR({err_f}/{flag_f},0)',
        f'=IFERROR({err_f}/$D${srow},0)',
        f'=IFERROR({dol_f}/$F${srow},0)',
        f'=IFERROR(({err_f}/{flag_f})/($D${srow}/$C${srow}),0)',
    ]
    nat = [None, None, rule['prec_train'], None, None, None]
    cells = {}
    for mi, (mname, mfmt, mf) in enumerate(zip(METRIC_NAMES, METRIC_FMTS, formulas)):
        rm = r0 + 2 + mi
        set_cell(ws,rm,6,mname, font=bold_font(10), fill=GRAY, align=left, border=thin())
        set_cell(ws,rm,7,formula=mf, fill=BLUE_LIGHT, align=center, border=thin(),
                 number_format=mfmt)
        if nat[mi] is not None:
            set_cell(ws,rm,8,nat[mi], align=center, border=thin(), number_format=mfmt,
                     font=Font(name=FONT,color='808080'))
        cells[mname] = f'G{rm}'
    summary_cells.append(cells)

# rule comparison table (right side, below the chart)
SUMM_ROW = 26
merge(ws,SUMM_ROW,10,SUMM_ROW,17, value='Rule Comparison Summary',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
for off, hdr in enumerate(['Rule','HH','Train prec.','Precision','Recall','$ Recall',
                           'n_flagged','Errors']):
    set_cell(ws,SUMM_ROW+1,10+off,hdr, font=bold_font(10), fill=GRAY, align=center, border=thin())
for bi, rule in enumerate(RULES):
    sr = SUMM_ROW + 2 + bi
    cc = summary_cells[bi]
    fill_r = BLUE_LIGHT if bi % 2 == 0 else WHITE
    vals = [(f'Rule {rule["num"]}', None), (rule['hh'], None),
            (rule['prec_train'], '0.0%'), (f'={cc["precision"]}', '0.0%'),
            (f'={cc["recall (count)"]}', '0.0%'), (f'={cc["dollar recall"]}', '0.0%'),
            (f'={cc["n_flagged"]}', '#,##0'), (f'={cc["errors caught"]}', '#,##0')]
    for off, (v, fmt) in enumerate(vals):
        c = set_cell(ws,sr,10+off,formula=v if isinstance(v,str) and v.startswith('=') else None,
                     value=None if isinstance(v,str) and v.startswith('=') else v,
                     fill=fill_r, align=center, border=thin())
        if fmt: c.number_format = fmt

# precision-recall chart (79 points; labels off — identify via the table)
chart = ScatterChart()
chart.title = 'Precision vs Recall (current Dashboard thresholds)'
chart.x_axis.title = 'Recall'; chart.y_axis.title = 'Precision'
chart.x_axis.numFmt = '0%';   chart.y_axis.numFmt = '0%'
chart.width, chart.height = 15, 12
for bi, rule in enumerate(RULES):
    rec_c, prec_c = summary_cells[bi]['recall (count)'], summary_cells[bi]['precision']
    xv = Reference(ws, min_col=7, min_row=int(rec_c[1:]),  max_row=int(rec_c[1:]))
    yv = Reference(ws, min_col=7, min_row=int(prec_c[1:]), max_row=int(prec_c[1:]))
    s = Series(yv, xv, title=f'Rule {rule["num"]}')
    s.marker.symbol, s.marker.size = 'circle', 7
    s.graphicalProperties.line.noFill = True
    chart.series.append(s)
chart.legend = None
ws.add_chart(chart, 'J3')
ws.freeze_panes = 'A9'

# ══════════════════════════════════════════════════════════════════════════════
# 5. GRID SEARCH SHEET (same engine as the Virginia workbook)
# ══════════════════════════════════════════════════════════════════════════════
DATA2D = f'{DQ}!$A$2:${LASTCOL}${MAX_ROW}'
HH     = dr('hh_group')
OVER   = dr('over_threshold')
ERRD   = dr('total_error_amount')

HELP0 = 100
GRID0 = HELP0 + 3*6 + 2
def h0(s):  return HELP0 + s*6
def gs(s):  return GRID0 + s*(MAXG + 3)
def rng(s, col):
    top = gs(s) + 2
    return f'${col}${top}:${col}${top + MAXG - 1}'
def sref(i, col): return f'${col}${7 + i}'

def dyn_pairs(lbl, thr_refs):
    parts = []
    for i in range(NSLOTS):
        parts.append(f'INDEX({DATA2D},0,{sref(i,"AF")})')
        parts.append(f'IF({sref(i,"A")}="","{lbl}",{sref(i,"B")}&{thr_refs[i]})')
    return ','.join(parts)

# ── Search bracket, per condition ─────────────────────────────────────────────
# The interactive sheet searches the SAME bracket the Tier-2 engine does: each
# threshold is scaled by the factor grid (fine for rules with <= 3 conditions,
# coarse beyond that, because the combinations grow exponentially in the
# condition count), never the 2nd-to-98th percentile of the state's own data as
# in v1. Structure is frozen; only thresholds move.
#
# A condition is collapsed to a single point (no search) when scaling it is
# meaningless: a shipped threshold of exactly 0, or a variable with <= 5 distinct
# observed values in the rule's own stratum (binary indicators, small counts).
# That mirrors the "low-cardinality variables (single cut each)" rule in
# state_threshold_gridsearch_v2.R.
NOSEARCH_MAX_LEVELS = 5

def condition_bracket(rule, c):
    fac = TCFG.factors_for(len(rule['conds']))
    thr = float(c['thr'])
    obs = df[df.hh_group == rule['hh']][c['var']].dropna().values.astype(float)
    levels = np.unique(obs[np.isfinite(obs)]).size
    if thr == 0.0 or levels <= NOSEARCH_MAX_LEVELS:
        return (thr, thr, max(base_step(c['var']), 1e-9), 1)
    lo, hi = round(thr * min(fac), 6), round(thr * max(fac), 6)
    step = (hi - lo) / (len(fac) - 1)
    if c['var'] in UNIT_VARS:                 # integer-valued: keep cuts whole
        step = max(1.0, round(step))
        lo, hi = math.floor(lo), math.ceil(hi)
    step = round(step, 6)
    npts = int(round((hi - lo) / step)) + 1 if step > 0 else 1
    return (lo, hi, step, npts)

for rule in RULES:
    rule['bracket'] = [condition_bracket(rule, c) for c in rule['conds']]
    combos = int(np.prod([b[3] for b in rule['bracket']]))
    assert combos <= MAXG, (f'rule {rule["num"]}: {combos} bracket combinations '
                            f'exceeds the {MAXG}-row grid capacity')
print('bracket combinations per rule: '
      f'min {min(int(np.prod([b[3] for b in r["bracket"]])) for r in RULES)}, '
      f'max {max(int(np.prod([b[3] for b in r["bracket"]])) for r in RULES)}')

ws_g = wb.create_sheet('Grid Search', 1)
ws_g.sheet_view.showGridLines = False
for col_letter, width in {'A':26,'B':13,'C':13,'D':13,'E':13,'F':12,'G':12,
                          'H':13,'I':13,'J':13,'K':12,'L':10,'M':8}.items():
    ws_g.column_dimensions[col_letter].width = width
for c in ('Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','L','M'):
    ws_g.column_dimensions[c].hidden = True

merge(ws_g,1,1,1,13, value=f'Grid Search — Optimize One Delivery Rule ({STATE_NAME})',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=16,color='FFFFFF'), align=center)
ws_g.row_dimensions[1].height = 32

# hidden preset tables: names in AA, terms in AB:AE; dropdown list in Y
NP = len(RULES)
set_cell(ws_g,1,25,'Custom')                          # Y1
for p, rule in enumerate(RULES):
    pname = f'Rule {rule["num"]} (HH {rule["hh"]})'
    set_cell(ws_g,2+p,27,pname)                       # AA
    set_cell(ws_g,2+p,25,pname)                       # Y (dropdown source)
    for i in range(NSLOTS):
        rr = 2 + p*NSLOTS + i
        if i < len(rule['conds']):
            c = rule['conds'][i]
            lo, hi, step, _ = rule['bracket'][i]
            set_cell(ws_g,rr,28,c['var']); set_cell(ws_g,rr,29,c['op'])
            set_cell(ws_g,rr,30,step); set_cell(ws_g,rr,31,c['thr'])
            set_cell(ws_g,rr,33,lo);    set_cell(ws_g,rr,34,hi)   # AG/AH: the bracket

merge(ws_g,4,1,4,13,
      value='Step 1 — Pick a rule preset, or pick "Custom" and fill the custom block on the right',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
set_cell(ws_g,5,1,'Rule preset', font=bold_font(), fill=GRAY, align=left)
set_cell(ws_g,5,2,f'Rule {RULES[0]["num"]} (HH {RULES[0]["hh"]})', fill=YELLOW,
         align=center, border=thin(), font=bold_font(), number_format='@')
set_cell(ws_g,5,4,'← delivery rules sorted by error dollars caught on this state\'s data, or "Custom"',
         font=Font(name=FONT,size=9,color='808080'), align=left)
merge(ws_g,5,8,5,11, value='Custom rule (used when preset = Custom)',
      fill=GRAY, font=bold_font(10), align=center)
for col, txt in [(1,'Variable'),(2,'Op'),(3,'Step'),(4,'Original threshold'),
                 (8,'Variable'),(9,'Op'),(10,'Step'),(11,'Original')]:
    set_cell(ws_g,6,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())

PLOOK = f'(MATCH($B$5,$AA$2:$AA${1+NP},0)-1)*{NSLOTS}'
PEND = 1 + NP*NSLOTS
for i in range(NSLOTS):
    r = 7 + i
    vlook = f'INDEX($AB$2:$AB${PEND},{PLOOK}+{i+1})'
    set_cell(ws_g,r,1,
             formula=f'=IF($B$5="Custom",IF($H{r}="","",$H{r}),IF({vlook}=0,"",{vlook}))',
             fill=BLUE_LIGHT, align=left, border=thin())
    for cc, pcol, ccol in [(2,'AC','I'),(3,'AD','J'),(4,'AE','K')]:
        set_cell(ws_g,r,cc,
                 formula=(f'=IF($A{r}="","",IF($B$5="Custom",${ccol}{r},'
                          f'INDEX(${pcol}$2:${pcol}${PEND},{PLOOK}+{i+1})))'),
                 fill=BLUE_LIGHT, align=center, border=thin())
    # empty slot fallback: point at hh_group so the degenerate condition is
    # the always-true stratum match
    set_cell(ws_g,r,32,formula=f'=IF($A{r}="",{RAW_OFF + DCOLS.index("hh_group") + 1},'
                               f'MATCH($A{r},{DQ}!$A$1:${LASTCOL}$1,0))')
    # AI/AJ: this slot's search bracket. Presets carry the bracket computed by
    # condition_bracket(); a Custom rule gets the same +/-25% window on its own
    # original value, so the interactive sheet can never search wider than Tier 2.
    for cc, pcol in ((35, 'AG'), (36, 'AH')):
        fpc = TCFG.factors_fine[0] if cc == 35 else TCFG.factors_fine[-1]
        set_cell(ws_g,r,cc,
                 formula=(f'=IF($A{r}="","",IF($B$5="Custom",ROUND($K{r}*{fpc},6),'
                          f'INDEX(${pcol}$2:${pcol}${PEND},{PLOOK}+{i+1})))'))
    if i < len(RULES[0]['conds']):
        c = RULES[0]['conds'][i]
        set_cell(ws_g,r,8, c['var'], fill=YELLOW, align=left,   border=thin(), font=bold_font())
        set_cell(ws_g,r,9, c['op'],  fill=YELLOW, align=center, border=thin(), font=bold_font())
        set_cell(ws_g,r,10,RULES[0]['bracket'][i][2], fill=YELLOW, align=center,
                 border=thin(), font=bold_font())
        set_cell(ws_g,r,11,c['thr'], fill=YELLOW, align=center, border=thin(), font=bold_font())
    else:
        for cc in (8,9,10,11):
            set_cell(ws_g,r,cc,None, fill=YELLOW, align=center, border=thin(), font=bold_font())

for i, v in enumerate(RULE_VARS, 2):
    set_cell(ws_g,i,26,v)                             # Z: variable dropdown source
dv_preset = DataValidation(type='list', formula1=f'=$Y$1:$Y${1+NP}', allow_blank=False)
ws_g.add_data_validation(dv_preset); dv_preset.add('B5')
dv_var = DataValidation(type='list', formula1=f'=$Z$2:$Z${len(RULE_VARS)+1}', allow_blank=True)
ws_g.add_data_validation(dv_var); dv_var.add('H7:H10')
dv_op = DataValidation(type='list', formula1='">=,<=,>,<,="', allow_blank=True)
ws_g.add_data_validation(dv_op); dv_op.add('I7:I10')

merge(ws_g,11,1,11,13, value='Step 2 — Settings',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
set_cell(ws_g,12,1,'Precision floor (on the BOUND)', font=bold_font(), fill=GRAY, align=left)
set_cell(ws_g,12,2,PREC_FLOOR, fill=YELLOW, align=center, border=thin(), font=bold_font(),
         number_format='0.0%')
set_cell(ws_g,12,4,'← applied to the lower confidence bound below, not to raw precision: '
                   'raw precision on the rows you searched is optimistic by construction',
         font=Font(name=FONT,size=9,color='808080'), align=left)
set_cell(ws_g,13,1,'Recall basis (maximized)', font=bold_font(), fill=GRAY, align=left)
set_cell(ws_g,13,2,'dollar', fill=YELLOW, align=center, border=thin(), font=bold_font())
set_cell(ws_g,13,4,'← what to maximize among qualifying combinations; fix this before you look, '
                   'never per rule',
         font=Font(name=FONT,size=9,color='808080'), align=left)
set_cell(ws_g,14,1,'Search alpha (per rule)', font=bold_font(), fill=GRAY, align=left)
set_cell(ws_g,14,2,TCFG.variant_gate_alpha, fill=YELLOW, align=center, border=thin(),
         font=bold_font(), number_format='0.00')
set_cell(ws_g,14,4,'← the bound\'s confidence is set from this divided by the number of '
                   'combinations searched, so a wider search must clear a higher bar',
         font=Font(name=FONT,size=9,color='808080'), align=left)
set_cell(ws_g,12,8,'Min flagged', font=bold_font(), fill=GRAY, align=left)
set_cell(ws_g,12,9,MIN_FLAGGED, fill=YELLOW, align=center, border=thin(), font=bold_font(),
         number_format='0')
set_cell(ws_g,12,10,'← hard support floor; below this a combination is too noisy to trust '
                    '(findings §9: at n >= 5 state-scale tuning collapsed)',
         font=Font(name=FONT,size=9,color='808080'), align=left)
set_cell(ws_g,13,8,'Recall floor', font=bold_font(), fill=GRAY, align=left)
set_cell(ws_g,13,9,RECALL_FLOOR, fill=YELLOW, align=center, border=thin(), font=bold_font(),
         number_format='0.0%')
set_cell(ws_g,13,10,'← optional extra floor on the share of error $ / cases caught',
         font=Font(name=FONT,size=9,color='808080'), align=left)
set_cell(ws_g,14,8,'Search bracket', font=bold_font(), fill=GRAY, align=left)
set_cell(ws_g,14,9,f'{TCFG.factors_fine[0]:g}x - {TCFG.factors_fine[-1]:g}x',
         fill=GRAY, align=center, border=thin(), font=bold_font())
set_cell(ws_g,14,10,'← fixed: every threshold is searched only inside this multiple of its '
                    'delivered value (coarser for rules with 4+ conditions). Structure never moves.',
         font=Font(name=FONT,size=9,color='808080'), align=left)
dv_floor = DataValidation(type='decimal', operator='between', formula1='0', formula2='1')
ws_g.add_data_validation(dv_floor); dv_floor.add('B12')
ws_g.add_data_validation(dv_floor2 := DataValidation(type='decimal', operator='between',
                                                     formula1='0', formula2='1'))
dv_floor2.add('I13')
dv_basis = DataValidation(type='list', formula1='"dollar,count"', allow_blank=False)
ws_g.add_data_validation(dv_basis); dv_basis.add('B13')
ws_g.add_data_validation(dv_alpha := DataValidation(type='decimal', operator='between',
                                                    formula1='0.0001', formula2='0.5'))
dv_alpha.add('B14')
dv_minf = DataValidation(type='whole', operator='greaterThanOrEqual', formula1='0')
ws_g.add_data_validation(dv_minf); dv_minf.add('I12')

merge(ws_g,15,1,15,13, value='Stratum Overview',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
for col, txt in [(1,'HH size'),(2,'Cases'),(3,'Errors'),(4,'Error $'),(5,'Combinations')]:
    set_cell(ws_g,16,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())
for s, lbl in enumerate(STRATA):
    r = 17 + s
    set_cell(ws_g,r,1,lbl, font=bold_font(), align=center, border=thin())
    set_cell(ws_g,r,2,formula=f'=COUNTIF({HH},"{lbl}")', align=center, border=thin(),
             number_format='#,##0', fill=BLUE_LIGHT)
    set_cell(ws_g,r,3,formula=f'=COUNTIFS({HH},"{lbl}",{OVER},1)', align=center,
             border=thin(), number_format='#,##0', fill=BLUE_LIGHT)
    set_cell(ws_g,r,4,formula=f'=SUMIFS({ERRD},{HH},"{lbl}",{OVER},1)', align=center,
             border=thin(), number_format='$#,##0', fill=BLUE_LIGHT)
    set_cell(ws_g,r,5,formula=f'=IF($A$7="","",$H${h0(s)})', align=center,
             border=thin(), number_format='#,##0', fill=BLUE_LIGHT)

F0 = 31
def fs(s): return F0 + s*18

merge(ws_g,21,1,21,13,
      value='Best per Stratum — max reach among combinations whose precision BOUND clears the '
            'floor and which flag at least the support minimum (vs delivered thresholds)',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
for col, txt in [(1,'HH size'),(2,'thr 1'),(3,'thr 2'),(4,'thr 3'),(5,'thr 4'),
                 (6,'precision'),(7,'recall'),(8,'dollar recall'),(9,'n_flagged'),
                 (10,'errors'),(11,'workload %'),(12,'precision bound')]:
    set_cell(ws_g,22,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())
BEST_FMTS = {6:'0.0%',7:'0.0%',8:'0.0%',9:'#,##0',10:'#,##0'}
for s, lbl in enumerate(STRATA):
    r = 23 + s
    M = f'$M${fs(s)+2}'
    set_cell(ws_g,r,1,f'{lbl} (best)', font=bold_font(), align=center, border=thin())
    for ci, col in enumerate('ABCD'):
        set_cell(ws_g,r,2+ci,
                 formula=f'=IF({M}="","",IF({sref(ci,"A")}="","",INDEX({rng(s,col)},{M})))',
                 fill=GREEN, align=center, border=thin(), font=bold_font())
    for cnum, gcol in [(6,'H'),(7,'I'),(8,'J'),(9,'E'),(10,'F')]:
        set_cell(ws_g,r,cnum,formula=f'=IF({M}="","",INDEX({rng(s,gcol)},{M}))',
                 fill=BLUE_LIGHT, align=center, border=thin(), number_format=BEST_FMTS[cnum])
    set_cell(ws_g,r,11,formula=f'=IF({M}="","",INDEX({rng(s,"E")},{M})/$B${17+s})',
             fill=BLUE_LIGHT, align=center, border=thin(), number_format='0.0%')
    set_cell(ws_g,r,12,formula=f'=IF({M}="","",INDEX({rng(s,"L")},{M}))',
             fill=BLUE_LIGHT, align=center, border=thin(), number_format='0.000')

ORIG_REFS = [sref(i,'D') for i in range(NSLOTS)]
for s, lbl in enumerate(STRATA):
    r = 26 + s
    gray = Font(name=FONT, color='808080')
    set_cell(ws_g,r,1,f'{lbl} (original)', align=center, border=thin(), font=gray)
    for i in range(NSLOTS):
        set_cell(ws_g,r,2+i,formula=f'=IF({sref(i,"A")}="","",{sref(i,"D")})',
                 align=center, border=thin(), font=gray)
    pairs = dyn_pairs(lbl, ORIG_REFS)
    set_cell(ws_g,r,9, formula=f'=IF($A$7="","",COUNTIFS({HH},"{lbl}",{pairs}))',
             align=center, border=thin(), number_format='#,##0', font=gray)
    set_cell(ws_g,r,10,formula=f'=IF($A$7="","",COUNTIFS({HH},"{lbl}",{pairs},{OVER},1))',
             align=center, border=thin(), number_format='#,##0', font=gray)
    set_cell(ws_g,r,12,formula=f'=IF($A$7="","",SUMIFS({ERRD},{HH},"{lbl}",{OVER},1,{pairs}))')
    set_cell(ws_g,r,6, formula=f'=IF($I{r}="","",IF($I{r}=0,0,$J{r}/$I{r}))',
             align=center, border=thin(), number_format='0.0%', font=gray)
    set_cell(ws_g,r,7, formula=f'=IF($J{r}="","",IF($C${17+s}=0,0,$J{r}/$C${17+s}))',
             align=center, border=thin(), number_format='0.0%', font=gray)
    set_cell(ws_g,r,8, formula=f'=IF($L{r}="","",IF($D${17+s}=0,0,$L{r}/$D${17+s}))',
             align=center, border=thin(), number_format='0.0%', font=gray)
    set_cell(ws_g,r,11,formula=f'=IF($I{r}="","",$I{r}/$B${17+s})',
             align=center, border=thin(), number_format='0.0%', font=gray)

merge(ws_g,30,1,30,13,
      value='Feasible Combinations — top 15 by reach per stratum (precision bound ≥ floor, '
            'support ≥ minimum)',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
FEAS_HEADS = [(1,'rank'),(2,'thr 1'),(3,'thr 2'),(4,'thr 3'),(5,'thr 4'),
              (6,'precision'),(7,'recall'),(8,'dollar recall'),(9,'n_flagged'),(10,'errors')]
for s, lbl in enumerate(STRATA):
    f0 = fs(s)
    merge(ws_g,f0,1,f0,13, value=f'HH size {lbl}',
          fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=10,color='FFFFFF'), align=left)
    for col, txt in FEAS_HEADS:
        set_cell(ws_g,f0+1,col,txt, font=bold_font(9), fill=GRAY, align=center, border=thin())
    for k in range(1, 16):
        r = f0 + 1 + k
        set_cell(ws_g,r,1,k, align=center, border=thin())
        set_cell(ws_g,r,13,
                 formula=f'=IFERROR(MATCH(LARGE({rng(s,"K")},{k}),{rng(s,"K")},0),"")')
        for ci, col in enumerate('ABCD'):
            set_cell(ws_g,r,2+ci,
                     formula=f'=IF($M{r}="","",IF({sref(ci,"A")}="","",INDEX({rng(s,col)},$M{r})))',
                     align=center, border=thin())
        for cnum, gcol, fmt in [(6,'H','0.0%'),(7,'I','0.0%'),(8,'J','0.0%'),
                                (9,'E','#,##0'),(10,'F','#,##0')]:
            set_cell(ws_g,r,cnum,
                     formula=f'=IF($M{r}="","",INDEX({rng(s,gcol)},$M{r}))',
                     align=center, border=thin(), number_format=fmt)

NOTES = [
    f'This sheet is an EXPLORATION aid, searching all years at once. Nothing in this workbook adopts what it finds: at public-QC sample sizes a searched threshold is mostly noise (the delivered thresholds always sit in the grid, and "leave it alone" is the expected outcome).',
    f'Each threshold is searched only inside {TCFG.factors_fine[0]:g}x-{TCFG.factors_fine[-1]:g}x its delivered value ({TCFG.factors_coarse[0]:g}x-{TCFG.factors_coarse[-1]:g}x for rules with more than {TCFG.fine_max_conds} conditions). Variables, operators, condition count and stratum never move: this perturbs an already-validated rule, it does not search for a new one.',
    'The stratum matching the rule\'s "HH" tag is the one the delivery list targets.',
    f'Qualification is the precision BOUND (column L, hidden) at a confidence set from the search alpha divided by the number of combinations searched, plus n_flagged >= the support floor. Raw precision on searched rows is optimistic by construction and is not the gate.',
    f'Best row = most reach (chosen basis) among qualifying combinations; ties by grid order. The delivered thresholds are always in the grid, so "leave it alone" is always available.',
    f'Capacity: up to {MAXG:,} combinations per stratum; the bracket keeps every rule well inside it.',
    f'The full scored grid lives in the blocks from row {GRID0} down (helper rows 100-119 are hidden).',
]
for i, txt in enumerate(NOTES):
    set_cell(ws_g,85+i,1,txt, font=Font(name=FONT,size=9,color='808080'), align=left)

for s, lbl in enumerate(STRATA):
    hh0 = h0(s)
    set_cell(ws_g,hh0,8,formula=f'=F{hh0+1}*F{hh0+2}*F{hh0+3}*F{hh0+4}')
    # the multiplicity-adjusted confidence z for this stratum's search: one-sided
    # alpha / (combinations searched). A wider search must clear a higher bar.
    # post-2007 functions must carry the _xlfn. prefix when written straight
    # into the XML, or Excel renders #NAME?
    set_cell(ws_g,hh0,9,
             formula=f'=IF($H${hh0}<=0,3,_xlfn.NORM.S.INV(1-MIN(0.5,$B$14/MAX($H${hh0},1))))')
    for i in range(NSLOTS):
        hr = hh0 + 1 + i
        V, S = sref(i,'A'), sref(i,'C')
        # bracket bounds, NOT data percentiles: the grid is the delivered
        # threshold's own +/-25% window (AI/AJ), so the interactive search can
        # never roam wider than the Tier-2 engine allows.
        ws_g.cell(row=hr, column=2).value = f'=IF({V}="","",{sref(i,"AI")})'
        ws_g.cell(row=hr, column=3).value = f'=IF({V}="","",{sref(i,"AJ")})'
        ws_g.cell(row=hr, column=4).value = f'=IF({V}="","",B{hr})'
        ws_g.cell(row=hr, column=5).value = f'=IF({V}="","",MAX(C{hr},D{hr}))'
        ws_g.cell(row=hr, column=6).value = (
            f'=IF({V}="",1,IF({S}<=0,1,ROUND((E{hr}-D{hr})/{S},0)+1))')
        ws_g.cell(row=hr, column=7).value = ('=1' if i == 0 else f'=F{hr-1}*G{hr-1}')
for rr in range(HELP0, GRID0):
    ws_g.row_dimensions[rr].hidden = True

for s, lbl in enumerate(STRATA):
    g0 = gs(s); data0 = g0 + 2; hh0 = h0(s)
    merge(ws_g,g0,1,g0,13, value=f'GRID — HH size {lbl}',
          fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=10,color='FFFFFF'), align=left)
    for ci, htxt in enumerate(['thr 1','thr 2','thr 3','thr 4','n_flagged','errors_caught',
                               'err_dollars','precision','recall_count','dollar_recall',
                               'sortkey','prec_lcb'], 1):
        set_cell(ws_g,g0+1,ci,htxt, font=bold_font(9), fill=GRAY, align=center)
    for r in range(data0, data0 + MAXG):
        for i in range(NSLOTS):
            hr = hh0 + 1 + i
            V, S = sref(i,'A'), sref(i,'C')
            ws_g.cell(row=r, column=1+i).value = (
                f'=IF($A$7="","",IF($H${hh0}>{MAXG},"",IF(ROW()-{g0+1}>$H${hh0},"",'
                f'IF({V}="","",ROUND($D${hr}+{S}*MOD(INT((ROW()-{data0})/$G${hr}),$F${hr}),6)))))')
        pairs = dyn_pairs(lbl, [f'$A{r}', f'$B{r}', f'$C{r}', f'$D{r}'])
        ws_g.cell(row=r, column=5).value  = f'=IF($A{r}="","",COUNTIFS({HH},"{lbl}",{pairs}))'
        ws_g.cell(row=r, column=6).value  = f'=IF($A{r}="","",COUNTIFS({HH},"{lbl}",{pairs},{OVER},1))'
        ws_g.cell(row=r, column=7).value  = f'=IF($A{r}="","",SUMIFS({ERRD},{HH},"{lbl}",{OVER},1,{pairs}))'
        ws_g.cell(row=r, column=8).value  = f'=IF($E{r}="","",IF($E{r}=0,0,$F{r}/$E{r}))'
        ws_g.cell(row=r, column=9).value  = f'=IF($E{r}="","",IF($C${17+s}=0,0,$F{r}/$C${17+s}))'
        ws_g.cell(row=r, column=10).value = f'=IF($E{r}="","",IF($D${17+s}=0,0,$G{r}/$D${17+s}))'
        # one-sided Wilson lower bound on this combination's precision, at the
        # multiplicity-adjusted z in $I$hh0: what the combination is worth once
        # you have paid for having searched the whole bracket
        Z = f'$I${hh0}'
        ws_g.cell(row=r, column=12).value = (
            f'=IF($E{r}="","",IF($E{r}=0,0,'
            f'MAX(0,($H{r}+{Z}^2/(2*$E{r})-{Z}*SQRT($H{r}*(1-$H{r})/$E{r}'
            f'+{Z}^2/(4*$E{r}^2)))/(1+{Z}^2/$E{r}))))')
        # qualification is the BOUND plus the support floor, never raw precision
        ws_g.cell(row=r, column=11).value = (
            f'=IF($H{r}="","",IF(AND($L{r}>=$B$12,$E{r}>=$I$12,'
            f'IF($B$13="dollar",$J{r},$I{r})>=$I$13),'
            f'IF($B$13="dollar",$J{r},$I{r})+({MAXG}-(ROW()-{g0+1}))*0.0000000001,""))')
        for cc in (8, 9, 10, 11, 12):
            ws_g.cell(row=r, column=cc).number_format = '0.000'

ws_g.freeze_panes = 'A21'
wb.calculation.fullCalcOnLoad = True

# ══════════════════════════════════════════════════════════════════════════════
# 6. SCORE THE DELIVERY LISTS, delivered thresholds as-is
#
# No tuning runs here or anywhere in the workbook (decision 2026-08-16): on a
# public QC sample the tiered tuning's support floor is never reachable, so it
# always refused, and it could not re-run inside Excel on a state's pasted
# data anyway. The workbook measures the delivered lists; it does not modify
# them. The tuning machinery stays in tuning.py for pipeline-side use on a
# state's internal data (methods/tuning_principles.md).
# ══════════════════════════════════════════════════════════════════════════════
is_err_all = (df['over_threshold'] == 1).values
ed_all = np.where(is_err_all, df['total_error_amount'].fillna(0).values, 0)
TOT_ERR = max(int(is_err_all.sum()), 1)
TOT_ED  = max(float(ed_all.sum()), 1e-9)


def score_list(rules_list):
    """Static per-rule metrics, each rule ALONE. Recall and $ Recall are
    shares of ALL errors / error dollars on the frame (decision 2026-08-18:
    grand-total denominators, comparable to the orange union row; rules
    overlap, so they deliberately do not sum to the union). Precision and
    workload stay within the rule's own flagged set / stratum. Must match
    make_live.rule_row's live formulas (risk R4)."""
    masks = [rule_selection.rule_mask(df, r) for r in rules_list]
    scores = []
    for m in masks:
        n, tp = int(m.sum()), int((m & is_err_all).sum())
        dol = float(ed_all[m].sum())
        scores.append({'prec': tp / n if n else 0.0, 'rec': tp / TOT_ERR,
                       'drec': dol / TOT_ED, 'n': n, 'tp': tp, 'dollars': dol})
    conds = [rule_selection.rule_text(r['conds']) for r in rules_list]
    return masks, scores, conds


orig_masks, blended_scores, blended_conds = score_list(RULES)

# ── hidden RuleFlags sheet: case x rule 0/1 hit matrices + live union columns ─
# row 1 rule nums | row 2 selected (from the rules tabs' Include? cols)
# row 4 target stratum | rows FLAG0.. one per Data case (only 1s written)
NR, NDATA = len(RULES), len(df)
NAT_ROW0 = 11                  # Blended Rules / National Rules row of the first rule
FLAG0 = 5

ws_f = wb.create_sheet('RuleFlags')
ws_f.sheet_state = 'hidden'
fcol = lambda j: get_column_letter(2 + j)                 # blended mask block B..
bsel = lambda j: get_column_letter(3 + NR + j)            # blended selection vector
NATU = get_column_letter(4 + 2*NR)                        # union under the Blended selection
set_cell(ws_f,1,1,'rule num'); set_cell(ws_f,2,1,'selected'); set_cell(ws_f,4,1,'stratum')
for j, rule in enumerate(RULES):
    set_cell(ws_f,1,2+j, rule['num'])
    set_cell(ws_f,4,2+j, rule['hh'])
    set_cell(ws_f,2,3+NR+j,
             formula=f"=IF('{BLENDED_SHEET}'!$L${NAT_ROW0+j}=TRUE,1,0)")
    for i in np.flatnonzero(orig_masks[j]):
        ws_f.cell(row=FLAG0+int(i), column=2+j).value = 1
HHRNG  = f'$B$4:${fcol(NR-1)}$4'
NATSEL = f'${bsel(0)}$2:${bsel(NR-1)}$2'
for i in range(NDATA):
    r = FLAG0 + i
    ws_f.cell(row=r, column=4+2*NR).value = (
        f'=IF(SUMPRODUCT({NATSEL},$B{r}:${fcol(NR-1)}{r})>0,1,0)')

# shared by the delivery tabs
cases_by = {lbl: int((df['hh_group'] == lbl).sum()) for lbl in STRATA}
D_HH  = f'{DQ}!${dc("hh_group")}$2:${dc("hh_group")}${1+NDATA}'
D_OV  = f'{DQ}!${dc("over_threshold")}$2:${dc("over_threshold")}${1+NDATA}'
D_AM  = f'{DQ}!${dc("total_error_amount")}$2:${dc("total_error_amount")}${1+NDATA}'
CHECKBOX_CELLS = {}


# ══════════════════════════════════════════════════════════════════════════════
# 7. THE RULE-LIST TAB — thresholds as-is, no tuning
#
# One tab: the effective rule list (the blended delivery list after the
# rule_selection transform), sorted by error dollars caught on the state frame.
# ══════════════════════════════════════════════════════════════════════════════
def delivery_list_tab(sheet_name, position, rules_list, scores, conds_text,
                      sel_rng, hh_rng, union_col, intro, title=None):
    title = title or sheet_name
    ws = wb.create_sheet(sheet_name, position)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '2F5496'
    LASTC = 13 + len(CHAR_COLS)          # ..., Include?(12), char block, Exact expression
    # D is wide because below 10 flagged cases the live build shows text
    # ('2 errors of 9 cases flagged') instead of a percentage
    for col_letter, width in {'A':9,'B':9,'C':60,'D':26,'E':11,'F':11,'G':10,'H':10,
                              'I':14,'J':12,'K':21,'L':9}.items():
        ws.column_dimensions[col_letter].width = width
    for ci, (_, _, _, w) in enumerate(CHAR_COLS, 13):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.column_dimensions[get_column_letter(LASTC)].width = 100
    merge(ws,1,1,1,LASTC, value=f'{title} — {STATE_NAME} FY{FY_LABEL}',
          fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=16,color='FFFFFF'), align=center)
    ws.row_dimensions[1].height = 32
    # instructions: black 11pt, merged A:K only (not the full sheet width) so
    # the text sits over the metric columns instead of stretching unreadably
    # far right (revision 2026-08-18)
    wrapped = Alignment(horizontal='left', vertical='top', wrap_text=True)
    merge(ws,2,1,2,11, value=intro,
          fill=GRAY, font=Font(name=FONT,size=11,color='000000'), align=wrapped)
    ws.row_dimensions[2].height = 60
    merge(ws,3,1,3,11,
          value=f'Rules are sorted by total error dollars caught on the data in the "{DATA_SHEET}" '
                'tab. For individual rules, Recall and $ Recall are that rule ALONE as a share of '
                'ALL errors / error dollars in the pasted data; rules overlap, so these columns do '
                'not sum to the combined rows above. (In the orange rows, Recall and $ Recall are '
                'within that row\'s household-size scope.) A rule\'s Workload % is the share of its '
                'own household-size stratum; in the orange summary rows it is the share of ALL '
                'cases. The exact machine expression of every rule is in the last column.',
          fill=GRAY, font=Font(name=FONT,size=11,color='000000'), align=wrapped)
    ws.row_dimensions[3].height = 60
    for col, txt in enumerate(['Rule','HH size','What the rule says','Precision','Recall',
                               '$ Recall','Flagged','Errors','Error $ caught','Workload %',
                               'Expected error $ by case','Include?'] +
                              [h for _, h, _, _ in CHAR_COLS] + ['Exact expression'], 1):
        set_cell(ws,4,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())

    merge(ws,5,1,5,LASTC,
          value='All rules combined (a case is flagged if ANY rule with Include? = TRUE flags it, '
                'at delivered thresholds)',
          fill=BLUE_LIGHT, font=bold_font(10), align=left)
    # the overall-results rows: orange (accent 6, lighter 60%), matching the
    # orange step tabs
    ORANGE60 = PatternFill('solid', fgColor='F8CBAD')
    u_rng = f'RuleFlags!${union_col}${FLAG0}:${union_col}${FLAG0+NDATA-1}'
    for i, (scope_hh, sel) in enumerate([('all', np.ones(len(df), bool))] +
                                        [(lbl, (df['hh_group'] == lbl).values) for lbl in STRATA]):
        r = 6 + i
        tot_err = max(int((is_err_all & sel).sum()), 1)
        tot_ed  = round(float(ed_all[sel].sum()), 2) or 1
        # workload in the summary rows is the share of ALL cases, so the
        # household-size rows are percentage-point portions of the 'all' row
        tot_n   = max(len(df), 1)
        sterm = '' if scope_hh == 'all' else f'*({D_HH}="{scope_hh}")'
        f_cnt = (f'=SUM(RuleFlags!{sel_rng})' if scope_hh == 'all'
                 else f'=SUMPRODUCT((RuleFlags!{sel_rng})*(RuleFlags!{hh_rng}="{scope_hh}"))')
        set_cell(ws,r,1,'All rules', align=center, border=thin(), fill=ORANGE60)
        set_cell(ws,r,2,scope_hh, align=center, border=thin(), fill=ORANGE60)
        set_cell(ws,r,3,formula=f_cnt, align=center, border=thin(), fill=ORANGE60, number_format='0')
        for col, f, fmt in [
            (4, f'=IF($G{r}=0,0,$H{r}/$G{r})',                            '0.0%'),
            (5, f'=$H{r}/{tot_err}',                                      '0.0%'),
            (6, f'=$I{r}/{tot_ed}',                                       '0.0%'),
            (7, f'=SUMPRODUCT(({u_rng}){sterm})',                         '#,##0'),
            (8, f'=SUMPRODUCT(({u_rng}){sterm}*({D_OV}))',                '#,##0'),
            (9, f'=SUMPRODUCT(({u_rng}){sterm}*({D_OV})*({D_AM}))',       '$#,##0'),
            (10,f'=$G{r}/{tot_n}',                                        '0.0%'),
            (11,f'=IFERROR(IF($G{r}=0,"",$I{r}/$G{r}),"")',               '$#,##0'),
        ]:
            set_cell(ws,r,col,formula=f, align=center, border=thin(),
                     number_format=fmt, fill=ORANGE60)
        set_cell(ws,r,12,None, align=center, border=thin(), fill=ORANGE60)

    merge(ws,10,1,10,LASTC, value='Individual rules — delivered thresholds, no tuning',
          fill=BLUE_LIGHT, font=bold_font(10), align=left)
    plain_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    for j, rule in enumerate(rules_list):   # already sorted by error $ caught
        r = NAT_ROW0 + j
        sc = scores[j]
        set_cell(ws,r,1,f'Rule {rule["num"]}', align=center, border=thin(), fill=GREEN)
        set_cell(ws,r,2,rule['hh'], align=center, border=thin(), fill=GREEN)
        set_cell(ws,r,3,render_plain(rule['conds']), align=plain_align, border=thin(),
                 font=Font(name=FONT,size=10))
        for col, key, fmt in [(4,'prec','0.0%'),(5,'rec','0.0%'),(6,'drec','0.0%'),
                              (7,'n','#,##0'),(8,'tp','#,##0'),(9,'dollars','$#,##0')]:
            set_cell(ws,r,col,round(float(sc[key]),4), align=center, border=thin(),
                     number_format=fmt, fill=GREEN)
        set_cell(ws,r,10,round(sc['n']/cases_by[rule['hh']],4), align=center, border=thin(),
                 number_format='0.0%', fill=GREEN)
        set_cell(ws,r,11,(round(sc['dollars']/sc['n'], 2) if sc['n'] else '—'),
                 align=center, border=thin(), number_format='$#,##0', fill=GREEN)
        set_cell(ws,r,12,True, fill=YELLOW, align=center, border=thin(), font=Font(name=FONT))
        for ci, (key, _, fmt, _) in enumerate(CHAR_COLS, 13):
            v = rule.get('char', {}).get(key)
            if v is None or (isinstance(v, float) and np.isnan(v)):
                v = ''
            elif fmt != '@':
                v = float(v)
            set_cell(ws,r,ci,v, align=(left if fmt == '@' else center), border=thin(),
                     number_format=fmt, font=Font(name=FONT,size=10))
        set_cell(ws,r,LASTC,conds_text[j], align=left,
                 font=Font(name=FONT,size=8,color='808080'))
    ws.freeze_panes = 'D5'
    # Include? is plain TRUE/FALSE text (no native checkboxes since
    # 2026-08-18); the dropdown constrains typing without needing Excel 365
    dv_incl = DataValidation(type='list', formula1='"TRUE,FALSE"', allow_blank=False)
    ws.add_data_validation(dv_incl)
    dv_incl.add(f'L{NAT_ROW0}:L{NAT_ROW0 + len(rules_list) - 1}')
    return ws


# row 2 carries the agreed workbook wording verbatim (revision note 2026-08-18)
ws_n = delivery_list_tab(
    BLENDED_SHEET, 1, RULES,
    scores=blended_scores,
    conds_text=blended_conds,
    sel_rng=NATSEL, hh_rng=HHRNG, union_col=NATU,
    title='Select Rules',
    intro=('A list of potential rules, selected from state and national rules, prioritized by '
           'precision and filled to a default 10% caseload. Expected error $ by case = error '
           'dollars caught / cases flagged. To remove a rule, set the text in the yellow '
           '"Include?" column to FALSE, which will decrease the workload % in cell J6. Columns '
           'to the right of "Include?" characterize each rule as applied to NATIONAL data (what '
           'error elements and natures it catches, who caused them, whether the error was '
           'discovered in the case file); they are fixed context and do not recompute from '
           'pasted data.'))

# ══════════════════════════════════════════════════════════════════════════════
# 8. SEE CASES FLAGGED BY A RULE — pick a rule, list the cases it flags, with
# a toggle between true errors only and every flagged case
# ══════════════════════════════════════════════════════════════════════════════
VIEWER = VIEWER_SHEET
ws_e = wb.create_sheet(VIEWER, 2)
ws_e.sheet_view.showGridLines = False
ws_e.sheet_properties.tabColor = '2F5496'
NCOL   = len(DCOLS)
NCOLV  = NCOL + 1                            # + the CASE_ID lead column
G0     = 3                                   # CASE_ID column (C); DCOLS follow
DROWC  = get_column_letter(G0 + NCOLV)       # data_row column
LISTN  = max(NR, 64)                         # rule-list slots: one per rule, so a
                                             # longer delivery list is never
                                             # silently truncated on this panel
GRIDN  = 60                                  # result rows
HELP_Y = get_column_letter(G0 + NCOLV + 1)   # rule index / dashboard row
HELP_Z = get_column_letter(G0 + NCOLV + 2)   # unused since 2026-08-18 (hit
HELP_A = get_column_letter(G0 + NCOLV + 3)   # matching moved into the Data
                                             # sheet's _view_* table columns);
                                             # kept so the panel columns to
                                             # the right keep their letters
SC_B   = get_column_letter(G0 + NCOLV + 4)   # rule score
SC_C   = get_column_letter(G0 + NCOLV + 5)   # rule match count
SC_D   = get_column_letter(G0 + NCOLV + 6)   # compacted sorted rule list
PRESETS = f"'Grid Search'!$AA$2:$AA${1+NR}"
for cl, w in {'A':20,'B':9}.items():
    ws_e.column_dimensions[cl].width = w
# wide enough for the full variable names in the rule readout (revision
# 2026-08-18)
ws_e.column_dimensions[get_column_letter(G0)].width = 26
for i in range(1, NCOLV):
    ws_e.column_dimensions[get_column_letter(G0+i)].width = 13 if i > 3 else 16
ws_e.column_dimensions[DROWC].width = 9
for cl in (HELP_Y, HELP_Z, HELP_A, SC_B, SC_C, SC_D):
    ws_e.column_dimensions[cl].hidden = True
TOG = f'${get_column_letter(G0+7)}$3'        # errors-only / all-flagged toggle

set_cell(ws_e,1,1,formula=f'=${SC_B}$1&" rules with matches"',
         fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=14,color='FFFFFF'), align=center)
ws_e.merge_cells('A1:B1'); ws_e.row_dimensions[1].height = 32
merge(ws_e,2,1,2,2, value='sorted by cases matched — updates live',
      fill=GRAY, font=Font(name=FONT,size=13,color='000000'), align=left)
ws_e.row_dimensions[2].height = 36
set_cell(ws_e,3,1,'Rule', font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,3,2,'matches', font=bold_font(10), fill=GRAY, align=center, border=thin())
for k in range(1, LISTN+1):
    r = 3 + k
    set_cell(ws_e,r,1,formula=f'=IF({k}>${SC_B}$1,"",INDEX(${SC_D}$2:${SC_D}${1+NR},{k}))',
             align=left, border=thin())
    set_cell(ws_e,r,2,formula=f'=IF({k}>${SC_B}$1,"",INT(LARGE(${SC_B}$2:${SC_B}${1+NR},{k})/1000))',
             align=center, border=thin())

merge(ws_e,1,G0,1,G0+11,
      value=f'{VIEWER} — the cases the selected rule flags, at current thresholds',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=16,color='FFFFFF'), align=center)
merge(ws_e,2,G0,2,G0+NCOLV-1,
      value='Pick a rule (the dropdown lists rules currently matching, sorted by matches) and '
            'choose whether to see only flagged cases that are true payment errors, or every '
            'flagged case. Pasted rows appear here too; the columns the rule uses are '
            f'highlighted in blue, and the first {GRIDN} matching cases are shown.',
      fill=GRAY, font=Font(name=FONT,size=13,color='000000'),
      align=Alignment(horizontal='left', vertical='top', wrap_text=True))
ws_e.row_dimensions[2].height = 36
set_cell(ws_e,3,G0,'Rule', font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,3,G0+1,f'Rule {RULES[0]["num"]} (HH {RULES[0]["hh"]})',
         fill=YELLOW, align=center, border=thin(), font=bold_font(), number_format='@')
ws_e.merge_cells(start_row=3, start_column=G0+1, end_row=3, end_column=G0+3)
set_cell(ws_e,3,G0+6,'Show', font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,3,G0+7,'true errors only', fill=YELLOW, align=center, border=thin(),
         font=bold_font(), number_format='@')
ws_e.merge_cells(start_row=3, start_column=G0+7, end_row=3, end_column=G0+8)
dv_tog = DataValidation(type='list', formula1='"true errors only,all flagged cases"',
                        allow_blank=False)
ws_e.add_data_validation(dv_tog)
dv_tog.add(f'{get_column_letter(G0+7)}3')
YREF = f'${HELP_Y}$1'; Y2REF = f'${HELP_Y}$2'
set_cell(ws_e,1,G0+NCOLV+1,
         formula=f'=MATCH(${get_column_letter(G0+1)}$3,{PRESETS},0)-1')
set_cell(ws_e,2,G0+NCOLV+1, formula=f'={BASE_ROW+2}+{YREF}*{BLOCK_HEIGHT}')
set_cell(ws_e,4,G0,'HH stratum', font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,4,G0+1,formula=f'=INDEX(RuleFlags!{HHRNG},1,{YREF}+1)',
         fill=BLUE_LIGHT, align=center, border=thin())
set_cell(ws_e,4,G0+3,'n_flagged', font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,4,G0+4,formula=f'=INDEX(Dashboard!$G:$G,{Y2REF})',
         fill=BLUE_LIGHT, align=center, border=thin())
set_cell(ws_e,4,G0+6,'errors caught', font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,4,G0+7,formula=f'=INDEX(Dashboard!$G:$G,{Y2REF}+1)',
         fill=BLUE_LIGHT, align=center, border=thin())
for off, txt in enumerate(['Variable','Op','Current threshold','(col ref)']):
    set_cell(ws_e,5,G0+off,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())
CV = get_column_letter(G0); OPV = get_column_letter(G0+1)
THV = get_column_letter(G0+2); IXV = get_column_letter(G0+3)
for i in range(NSLOTS):
    r = 6 + i
    set_cell(ws_e,r,G0,  formula=f'=IFERROR(IF(INDEX(Dashboard!$A:$A,{Y2REF}+{i})=0,"",'
                                 f'INDEX(Dashboard!$A:$A,{Y2REF}+{i})),"")',
             fill=BLUE_LIGHT, align=center, border=thin())
    set_cell(ws_e,r,G0+1,formula=f'=IFERROR(IF(INDEX(Dashboard!$B:$B,{Y2REF}+{i})=0,"",'
                                 f'INDEX(Dashboard!$B:$B,{Y2REF}+{i})),"")',
             fill=BLUE_LIGHT, align=center, border=thin())
    set_cell(ws_e,r,G0+2,formula=f'=IF(${CV}{r}="","",INDEX(Dashboard!$C:$C,{Y2REF}+{i}))',
             fill=BLUE_LIGHT, align=center, border=thin())
    set_cell(ws_e,r,G0+3,formula=f'=IF(${CV}{r}="","",IFERROR(MATCH(${CV}{r},'
                                 f'{DQ}!$A$1:${LASTCOL}$1,0),""))',
             fill=BLUE_LIGHT, align=center, border=thin())
set_cell(ws_e,10,G0,'CASE_ID', font=bold_font(10), fill=GRAY, align=center, border=thin())
for i, h in enumerate(DCOLS):
    set_cell(ws_e,10,G0+1+i,h, font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,10,G0+NCOLV,'data_row', font=bold_font(10), fill=GRAY, align=center, border=thin())
# blue-highlight the columns the selected rule uses (revision 2026-08-18):
# conditional formatting keyed to the rule's variable readouts in CV6:CV9 —
# the header turns BLUE_LIGHT when its name is among the rule's variables,
# and the grid cells below follow (only on rows actually showing a case).
# NB conditional-format fills read the dxf BACKGROUND color, hence bgColor.
HL_FILL  = PatternFill(bgColor='BDD7EE')
HL_FIRST = get_column_letter(G0 + 1)
HL_LAST  = get_column_letter(G0 + NCOL)
ws_e.conditional_formatting.add(
    f'{HL_FIRST}10:{HL_LAST}10',
    FormulaRule(formula=[f'COUNTIF(${CV}$6:${CV}$9,{HL_FIRST}$10)>0'],
                fill=HL_FILL))
ws_e.conditional_formatting.add(
    f'{HL_FIRST}11:{HL_LAST}{10 + GRIDN}',
    FormulaRule(formula=[f'AND(COUNTIF(${CV}$6:${CV}$9,{HL_FIRST}$10)>0,'
                         f'${DROWC}11<>"")'],
                fill=HL_FILL))
# ── per-case matching lives INSIDE the Data sheet (2026-08-18) ───────────────
# Two hidden columns appended right after the feature block, BEFORE make_live
# forms the CaseData table over the sheet — so they become table columns that
# Excel auto-fills when a state pastes rows, and the viewer covers pasted data
# (it used to read a fixed helper block sized to the shipped sample only).
#   _view_hit  does the currently selected rule flag this row (0/1, honoring
#              the stratum and the errors-only toggle)
#   _view_cum  running count of hits; the reference to the cell above is
#              wrapped in N() so the header text reads as 0 on the first row
VQ  = qref(VIEWER)
VHI = RAW_OFF + len(DCOLS) + 1
VHC, VCC = get_column_letter(VHI), get_column_letter(VHI + 1)
ws_data.cell(row=1, column=VHI, value='_view_hit')
ws_data.cell(row=1, column=VHI + 1, value='_view_cum')
HHC, OVC = dc('hh_group'), dc('over_threshold')
STRAT_REF = f'{VQ}!${get_column_letter(G0+1)}$4'
for i in range(NDATA):
    rr = 2 + i
    slots = '*'.join(
        f'IF({VQ}!${CV}${6+k}="",1,COUNTIF(INDEX($A{rr}:${LASTCOL}{rr},1,'
        f'{VQ}!${IXV}${6+k}),{VQ}!${OPV}${6+k}&{VQ}!${THV}${6+k}))'
        for k in range(NSLOTS))
    ws_data.cell(row=rr, column=VHI).value = (
        f'=IF(${HHC}{rr}<>{STRAT_REF},0,'
        f'IF(AND({VQ}!{TOG}="true errors only",${OVC}{rr}<>1),0,{slots}))')
    ws_data.cell(row=rr, column=VHI + 1).value = f'=${VHC}{rr}+N(${VCC}{rr-1})'
ws_data.column_dimensions[VHC].hidden = True
ws_data.column_dimensions[VCC].hidden = True

# the viewer grid looks hits up in the whole _view_cum column, so rows pasted
# past the shipped sample appear too (display still capped at GRIDN matches);
# MATCH over a whole column returns the sheet row directly
CUMC = f'{DQ}!${VCC}:${VCC}'
for r in range(11, 11+GRIDN):
    ws_e.cell(row=r, column=G0).value = (
        f'=IFERROR(INDEX({DQ}!$A:$A,MATCH(ROW()-10,{CUMC},0)),"")')
    for i in range(NCOL):
        ws_e.cell(row=r, column=G0+1+i).value = (
            f'=IFERROR(INDEX({DQ}!${FEAT0}:${LASTCOL},MATCH(ROW()-10,{CUMC},0),'
            f'COLUMN()-{G0}),"")')
    ws_e.cell(row=r, column=G0+NCOLV).value = f'=IFERROR(MATCH(ROW()-10,{CUMC},0),"")'
ws_e.cell(row=1, column=G0+NCOLV+4).value = f'=COUNTIF(${SC_C}$2:${SC_C}${1+NR},">0")'
for j in range(NR):
    r = 2 + j
    # metric row: n_flagged (BASE_ROW+2) in all-flagged mode, errors (+3) in
    # errors-only mode
    mrow = (f'{BASE_ROW+2}+{j}*{BLOCK_HEIGHT}'
            f'+IF({TOG}="true errors only",1,0)')
    ws_e.cell(row=r, column=G0+NCOLV+4).value = (
        f'=INDEX(Dashboard!$G:$G,{mrow})*1000+({NR}-{j})')
    ws_e.cell(row=r, column=G0+NCOLV+5).value = (
        f'=INDEX(Dashboard!$G:$G,{mrow})')
    ws_e.cell(row=r, column=G0+NCOLV+6).value = (
        f'=IF(ROW()-1>${SC_B}$1,"",INDEX({PRESETS},'
        f'{1+NR}-MOD(LARGE(${SC_B}$2:${SC_B}${1+NR},ROW()-1),1000)))')
wb.defined_names.add(DefinedName('RuleListLive',
    attr_text=(f"OFFSET('{VIEWER}'!${SC_D}$2,0,0,"
               f"MAX('{VIEWER}'!${SC_B}$1,1),1)")))
dv_live = DataValidation(type='list', formula1='RuleListLive', allow_blank=False)
ws_e.add_data_validation(dv_live)
dv_live.add(f'{get_column_letter(G0+1)}3')
ws_e.freeze_panes = f'{get_column_letter(G0)}11'

# ══════════════════════════════════════════════════════════════════════════════
# 9. STEP 4 — EXPORT RULES: the rules still set to TRUE on the Step 3 tab, in
# tab order, with their exact machine logic. Live: setting Include? to FALSE
# removes the rule here immediately. No dynamic arrays (Excel-2013-safe).
# ══════════════════════════════════════════════════════════════════════════════
BQ = qref(BLENDED_SHEET)
EXPR_L = get_column_letter(13 + len(CHAR_COLS))        # Step 3 exact-expression col
ws_x = wb.create_sheet(EXPORT_SHEET)
ws_x.sheet_view.showGridLines = False
# pale orange: the optional post-Step-3 tabs (Eric's WA scheme, 2026-08-19)
ws_x.sheet_properties.tabColor = 'FDE9D9'
for cl, w in {'A': 9, 'B': 9, 'C': 60, 'D': 100}.items():
    ws_x.column_dimensions[cl].width = w
for cl in ('H', 'I'):
    ws_x.column_dimensions[cl].hidden = True
merge(ws_x, 1, 1, 1, 4, value=f'Export Rules — {STATE_NAME}',
      fill=BLUE_DARK, font=Font(name=FONT, bold=True, size=16, color='FFFFFF'),
      align=center)
ws_x.row_dimensions[1].height = 32
merge(ws_x, 2, 1, 2, 4,
      value=f'The rules currently set to TRUE on the "{BLENDED_SHEET}" tab, in the same '
            'order, with their exact machine logic — this list updates as you change '
            'Include?. A few example uses: filter your caseload in Excel with these '
            'conditions, translate them into a query in your eligibility or case-management '
            'system, or send this sheet to your vendor. Variable definitions are on the '
            'dictionary tab.',
      fill=GRAY, font=Font(name=FONT, size=11, color='000000'),
      align=Alignment(horizontal='left', vertical='top', wrap_text=True))
ws_x.row_dimensions[2].height = 45
for col, txt in enumerate(['Rule', 'HH size', 'What the rule says',
                           'Exact rule logic'], 1):
    set_cell(ws_x, 3, col, txt, font=bold_font(10), fill=GRAY, align=center,
             border=thin())
# hidden helpers: H = 1 if the k-th Step 3 rule is TRUE, I = running count
XCUM = f'$I$4:$I${3 + NR}'
for k in range(1, NR + 1):
    r = 3 + k
    ws_x.cell(row=r, column=8).value = f'=IF({BQ}!$L${10 + k}=TRUE,1,0)'
    ws_x.cell(row=r, column=9).value = f'=$H{r}+N($I{r - 1})'
    idx = f'MATCH(ROW()-3,{XCUM},0)'
    ws_x.cell(row=r, column=1).value = (
        f'=IFERROR(INDEX({BQ}!$A$11:$A${10 + NR},{idx}),"")')
    ws_x.cell(row=r, column=2).value = (
        f'=IFERROR(INDEX({BQ}!$B$11:$B${10 + NR},{idx}),"")')
    c3 = set_cell(ws_x, r, 3,
                  formula=f'=IFERROR(INDEX({BQ}!$C$11:$C${10 + NR},{idx}),"")',
                  align=Alignment(horizontal='left', vertical='top', wrap_text=True),
                  font=Font(name=FONT, size=10))
    ws_x.cell(row=r, column=4).value = (
        f'=IFERROR(INDEX({BQ}!${EXPR_L}$11:${EXPR_L}${10 + NR},{idx}),"")')
    ws_x.cell(row=r, column=4).font = Font(name=FONT, size=9)
    ws_x.cell(row=r, column=1).alignment = center
    ws_x.cell(row=r, column=2).alignment = center
ws_x.freeze_panes = 'A4'

# ── final touches ─────────────────────────────────────────────────────────────
ws.sheet_state   = 'hidden'      # Dashboard  (engine)
ws_g.sheet_state = 'hidden'      # Grid Search(engine)

# ── Save ──────────────────────────────────────────────────────────────────────
os.makedirs(STATE_DIR, exist_ok=True)
wb.save(OUT)
import json
os.makedirs(BUILD_DIR, exist_ok=True)
# empty since 2026-08-18 (Include? is plain TRUE/FALSE, no native checkboxes);
# kept so postprocess_workbook's calc-chain stage keeps its handoff file
json.dump(CHECKBOX_CELLS, open(os.path.join(BUILD_DIR, 'checkbox_cells.json'), 'w'))
json.dump({'out': OUT, 'state': STATE_NAME, 'abbr': STATE_ABBR},
          open(os.path.join(BUILD_DIR, 'target.json'), 'w'))
print(f'Saved: {OUT}')
print(f'Rules: {len(RULES)} | Data rows: {len(df)}')
