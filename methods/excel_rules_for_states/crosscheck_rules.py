"""
Test that the rules as shown in the workbook's rules tab match an
independent R implementation of the same rules over the same years of
reg_model_data.rds.

    python crosscheck_rules.py WA [--workbook path.xlsx]

Runs crosscheck_rules.R against the workbook's EFFECTIVE rule list
(.build/effective_rules_<ABBR>.csv, written by the build via
rule_selection.py — the delivery CSV after the div-100 drop, bbce strip and
buffer promotion). The R side re-parses that CSV's rule text, re-applies
every rule with the miner's own prep_features(), and scores it; this
compares, per rule: n flagged, errors caught, error dollars, precision,
recall, dollar recall, workload — plus the all-rules union, recomputed from
the workbook's RuleFlags hit matrix, overall and per household-size stratum.
Because the R side re-derives each mask from the stripped rule text on the
state's own rows, a bbce strip that changed any flag set would surface here.

Exit code 0 = everything matches. Run it against the PLAIN build in
.build/out_<ABBR>/ (the default): its rules-tab values are static, so they
can be read without Excel. The delivered workbook holds the same numbers as
formulas; verify those by opening it in Excel (make_state.py does).
"""
import argparse
import os
import shutil
import subprocess
import sys
import tempfile

import numpy as np
import openpyxl
import pandas as pd

import states as STATE_REGISTRY
from workbook_layout import BLENDED_SHEET

PKG = os.path.dirname(os.path.abspath(__file__))
NAT_ROW0, FLAG0 = 10, 5


def find_repo(start):
    cands = [os.environ['SNAP_REPO']] if os.environ.get('SNAP_REPO') else []
    d = start
    for _ in range(6):
        cands += [d, os.path.join(d, 'snap_qc')]
        d = os.path.dirname(d)
    for c in cands:
        if (os.path.isfile(os.path.join(c, 'reg_model_data.rds'))
                and os.path.isfile(os.path.join(c, 'rule_mining_helpers.R'))):
            return os.path.abspath(c)
    raise SystemExit('cannot find the snap_qc checkout; set SNAP_REPO')


def rscript():
    for c in ([os.environ['RSCRIPT']] if os.environ.get('RSCRIPT') else []) + [
            os.path.join('C:' + os.sep, 'Program Files', 'R', 'R-4.5.1',
                         'bin', 'Rscript.exe'), 'Rscript']:
        if os.path.isabs(c) and os.path.isfile(c):
            return c
        if not os.path.isabs(c) and shutil.which(c):
            return shutil.which(c)
    raise SystemExit('Rscript not found; set RSCRIPT')


def run_r(cfg, repo, csv_path, tmp, tag):
    r_out = os.path.join(tmp, f'r_rules_{tag}.csv')
    # role "all": the effective-rules CSV already holds exactly the tab's
    # rules (transformed core + promoted buffer)
    cmd = [rscript(), os.path.join(PKG, 'crosscheck_rules.R'),
           '--state', cfg['name'], '--csv', csv_path, '--out', r_out,
           '--years', ','.join(str(y) for y in cfg.get('years', (2022, 2023, 2024))),
           '--repo', repo, '--role', 'all']
    print('$ ' + ' '.join(cmd))
    if subprocess.run(cmd, cwd=repo).returncode != 0:
        raise SystemExit('crosscheck_rules.R failed')
    return (pd.read_csv(r_out),
            pd.read_csv(r_out.replace('.csv', '_union.csv')))


def check_tab(wb, sheet_name, r_rules, r_union, mask_col0, cases):
    """Per-rule static values on `sheet_name` vs R, plus the union recomputed
    from the RuleFlags mask block starting at column `mask_col0`."""
    ws = wb[sheet_name]
    NR = len(r_rules)
    fails = 0
    for j in range(NR):
        r = NAT_ROW0 + j
        rk = int(str(ws.cell(row=r, column=1).value).split()[-1])
        rr = r_rules[r_rules['rank'] == rk]
        assert len(rr) == 1, f'{sheet_name}: rank {rk} missing from the R output'
        rr = rr.iloc[0]
        vals = [('hh', ws.cell(row=r, column=2).value, rr['hh'], None),
                ('precision', ws.cell(row=r, column=4).value, rr['precision'], 5e-5),
                ('recall', ws.cell(row=r, column=5).value, rr['recall'], 5e-5),
                ('dollar_recall', ws.cell(row=r, column=6).value, rr['dollar_recall'], 5e-5),
                ('n_flagged', ws.cell(row=r, column=7).value, rr['n_flagged'], 0),
                ('errors', ws.cell(row=r, column=8).value, rr['errors'], 0),
                ('dollars', ws.cell(row=r, column=9).value, rr['dollars'], 0.51),
                ('workload', ws.cell(row=r, column=10).value, rr['workload'], 5e-5)]
        for name, got, want, tol in vals:
            if tol is None:
                ok = str(got) == str(want)
            else:
                # the builder writes ratios as round(value, 4); compare to that
                ok = abs(float(got) - round(float(want), 4)) <= max(tol, 1e-9)
            if not ok:
                fails += 1
                print(f'MISMATCH {sheet_name} rule {rk} {name}: '
                      f'workbook={got!r} R={want!r}')
    print(f'{sheet_name}: {NR} rules x 8 fields checked, {fails} mismatches')

    rf = wb['RuleFlags']
    ndata = len(cases)
    hit = np.zeros(ndata, dtype=int)
    for j in range(NR):
        col = mask_col0 + j
        for i in range(ndata):
            if rf.cell(row=FLAG0 + i, column=col).value == 1:
                hit[i] = 1
    is_err = (cases['over_threshold'] == 1).values
    ed = np.where(is_err, cases['total_error_amount'].fillna(0).values, 0)
    ufails = 0
    for _, u in r_union.iterrows():
        sel = np.ones(ndata, bool) if u['scope'] == 'all' \
            else (cases['hh_group'].astype(str) == u['scope']).values
        got = (int((hit & sel).sum()), int(((hit == 1) & sel & is_err).sum()),
               float(ed[(hit == 1) & sel].sum()))
        want = (int(u['flagged']), int(u['errors']), float(u['dollars']))
        if got != want:
            ufails += 1
            print(f"UNION MISMATCH {sheet_name} scope {u['scope']}: "
                  f"workbook={got} R={want}")
        else:
            print(f"{sheet_name} union scope {u['scope']:>3}: flagged {got[0]}, "
                  f"errors {got[1]}, dollars {got[2]:.0f}  == R")
    print(f'{sheet_name} union: {ufails} mismatches')
    return fails + ufails


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('state', nargs='?', default='WA')
    ap.add_argument('--workbook')
    a = ap.parse_args()
    cfg = STATE_REGISTRY.get(a.state.upper())
    repo = find_repo(PKG)
    wbp = a.workbook or os.path.join(PKG, '.build', f'out_{cfg["abbr"]}',
                                     f'SNAP_flagging_rules_{cfg["abbr"]}.xlsx')
    cases = pd.read_csv(os.path.join(os.path.dirname(wbp),
                                     f'{cfg["abbr"].lower()}_cases.csv'))

    eff_csv = os.path.join(PKG, '.build', f'effective_rules_{cfg["abbr"]}.csv')
    assert os.path.isfile(eff_csv), (
        f'{eff_csv} not found — it is written by build_workbook_v2.py '
        '(rule_selection.effective_rules); rebuild the state first')

    wb = openpyxl.load_workbook(wbp)
    tmp = tempfile.mkdtemp(prefix='crosscheck_')

    blended, blended_u = run_r(cfg, repo, eff_csv, tmp, 'blended')
    # RuleFlags layout (build_workbook_v2): the rule list's hit-mask block
    # starts at column 2
    total = check_tab(wb, BLENDED_SHEET, blended, blended_u, 2, cases)

    sys.exit(1 if total else 0)


if __name__ == '__main__':
    main()
