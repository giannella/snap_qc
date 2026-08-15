"""
Stage 1 — build a state's SNAP QC dashboard from its delivery rule list.

Sheets produced:
  Data               reconstructed state QC case data
  Dashboard          one tuning block per rule + PR chart      (hidden engine)
  Grid Search        formula-based threshold search            (hidden engine)
  Summary            tuned vs original thresholds per rule, mixable via Include?
  Summary (National) delivery thresholds as-is, no grid search
  Error Cases        live list of rules catching errors + the cases they catch
  RuleFlags          case x rule hit matrices                  (hidden engine)

Pick the state with the SNAP_STATE environment variable (default WA); add new
states in states.py.  Run through make_state.py so the native checkboxes get
injected and the result is verified in Excel.
"""
import os
import re
import math
import itertools
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.formula import ArrayFormula   # needs openpyxl >= 3.1
import numpy as np
import pandas as pd
import pyreadstat

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG — state settings live in states.py; nothing here needs editing
# ══════════════════════════════════════════════════════════════════════════════
import states as STATE_REGISTRY

# ── WARNING (2026-07-29) ──────────────────────────────────────────────────────
# This builder reconstructs its features from the raw public .sav files. That
# reproduces the row universe exactly but NOT the feature values: it skips the
# pre-QC variable restoration the munging script performs
# (1_data_munging_..._for_using_public_qc_data.R, correct_variables <- TRUE),
# which is the scale the delivery rules were mined on. On Washington's 114-rule
# 10% list that cost 21 rules that never fired at all, and held-out precision
# 0.255 against 0.318. Use build_workbook_v2.py, which reads the munged frame.
import sys as _sys
print('WARNING: build_workbook.py scores rules on features rebuilt from the raw '
      '.sav files, not the munged (pre-QC-restored) values the rules were mined '
      'on. Prefer build_workbook_v2.py. See README "Known divergence".',
      file=_sys.stderr)

PKG          = os.path.dirname(os.path.abspath(__file__))
BASE         = os.path.dirname(PKG)                # project root
BUILD_DIR    = os.path.join(PKG, '.build')         # stage-to-stage handoff files
_cfg         = STATE_REGISTRY.get(os.environ.get('SNAP_STATE', 'WA'))

STATE_NAME   = _cfg['name']
STATE_ABBR   = _cfg['abbr']
STATE_FIPS   = _cfg['fips']
FY_LABEL     = _cfg['fy_label']
ROLE_FILTER  = _cfg['role_filter']
QC_FILES     = _cfg['qc_files']
def _find_repo(start):
    """Locate the snap_qc checkout by walking up (same rule as v2)."""
    cands = [os.environ['SNAP_REPO']] if os.environ.get('SNAP_REPO') else []
    d = start
    for _ in range(6):
        cands += [d, os.path.join(d, 'snap_qc')]
        d = os.path.dirname(d)
    for c in cands:
        if os.path.isdir(os.path.join(c, 'qc_data')):
            return os.path.abspath(c)
    raise SystemExit('cannot find the snap_qc checkout (needs qc_data/); set SNAP_REPO')


REPO         = _find_repo(PKG)
_dc          = os.path.join(REPO, _cfg['delivery_csv'])
DELIVERY_CSV = _dc if os.path.isfile(_dc) else os.path.join(BASE, _cfg['delivery_csv'])
STATE_DIR    = os.environ.get('SNAP_OUT_DIR') or os.path.join(
                   PKG, 'state_workbooks', STATE_ABBR)
CASES_CSV    = os.path.join(STATE_DIR, f'{STATE_ABBR.lower()}_cases.csv')
OUT          = os.path.join(STATE_DIR, f'snap_qc_dashboard_{STATE_ABBR}.xlsx')

MAX_ROW      = 3000            # Data-sheet formula range ceiling
PREC_FLOOR   = STATE_REGISTRY.PREC_FLOOR
MIN_FLAGGED  = STATE_REGISTRY.MIN_FLAGGED
RECALL_FLOOR = STATE_REGISTRY.RECALL_FLOOR
MAXG         = 2500            # grid rows allocated per stratum
STRATA       = ['1', '2-3', '4+']
NSLOTS       = 4               # condition slots in the grid engine

# ══════════════════════════════════════════════════════════════════════════════
# 1. BUILD THE STATE DATASET (mirrors build_virginia_dataset.py + rule variables)
# ══════════════════════════════════════════════════════════════════════════════
yd = pd.read_csv(os.path.join(REPO, 'additional_data/year_data.csv')).dropna(axis=1, how='all')
yd.columns = [c.strip() for c in yd.columns]
err_thr   = dict(zip(yd.year, yd.error_threshold))
max_shel  = dict(zip(yd.year, yd.max_shelter_deduction))
min_allot = dict(zip(yd.year, yd.min_allotment))

sd = pd.read_csv(os.path.join(REPO, 'additional_data/standard_deductions.csv')).dropna(axis=1, how='all')
ma = pd.read_csv(os.path.join(REPO, 'additional_data/max_allotments.csv')).dropna(axis=1, how='all')
sd.columns = [str(c).strip() for c in sd.columns]
ma.columns = [str(c).strip() for c in ma.columns]
sd = sd.set_index('year'); ma = ma.set_index('year')

def lk(tbl, yr, sz):
    sz = int(min(max(sz, 1), 20)); yr = int(yr)
    if yr not in tbl.index:
        yr = min(tbl.index, key=lambda y: abs(y - yr))
    return float(tbl.loc[yr, str(sz)])

frames = []
for f in QC_FILES:
    dfi, _ = pyreadstat.read_sav(os.path.join(REPO, f))
    dfi.columns = [c.upper() for c in dfi.columns]
    dfi = dfi[dfi['STATE'] == STATE_FIPS].copy()
    print(f'{f}: {len(dfi)} {STATE_ABBR} rows')
    frames.append(dfi)
d = pd.concat(frames, ignore_index=True)
print(f'{STATE_ABBR} rows total:', len(d))

def c0(col): return pd.to_numeric(d[col], errors='coerce').fillna(0.0)

yr = d['YRMONTH'].astype(str).str[:4].astype(int)
mo = d['YRMONTH'].astype(str).str[4:6].astype(int)
d['fiscal_year'] = np.where(mo >= 10, yr + 1, yr)
fy = d['fiscal_year'].values

size      = pd.to_numeric(d['FSUSIZE'], errors='coerce').fillna(0).astype(int).clip(lower=1)
rawearn   = c0('FSEARN')
rawunearn = c0('FSUNEARN')
rawgross  = rawearn + rawunearn
rawernded = rawearn * 0.2
rawdepded = c0('FSDEPDED')
rawcsded  = c0('FSCSDED')
rawmedded = c0('FSMEDDED')
rawrent   = c0('RENT')
rawutil   = c0('UTIL')
rawsltexp = rawrent + rawutil
elder     = c0('FSNELDER')
dis       = c0('FSNDIS')

rawstdded = np.array([lk(sd, fy[i], size.iloc[i]) for i in range(len(d))])
rawbenmax = np.array([lk(ma, fy[i], size.iloc[i]) for i in range(len(d))])
rawhomeless = c0('HOMELESS_DED') if 'HOMELESS_DED' in d.columns else np.zeros(len(d))

rawnet_bs = rawgross - (rawernded + rawdepded + rawmedded + rawcsded + rawstdded)
half      = np.maximum(rawnet_bs * 0.5, 0)
maxshel   = np.where((elder + dis) > 0, np.inf,
                     np.array([max_shel.get(int(y), np.inf) for y in fy]))
slt_unc   = np.floor(rawsltexp - half)
rawsltded = np.where(np.isinf(maxshel),
                     np.maximum(slt_unc, 0),
                     np.minimum(np.maximum(slt_unc, 0), maxshel))
rawnet_an  = np.floor(rawnet_bs - (rawsltded + rawhomeless))
rawben_unc = np.floor(rawbenmax - 0.3 * rawnet_an)
minben     = np.where(size < 3, np.array([min_allot.get(int(y), 0) for y in fy]), 0)
rawben_rec = np.minimum(np.maximum(rawben_unc, minben), rawbenmax)

months_since_cert = pd.to_numeric(d['LASTCERT'], errors='coerce')

certhhsz  = pd.to_numeric(d['CERTHHSZ'], errors='coerce').fillna(1).clip(lower=1)
abwd_cols = [c for c in d.columns if c.startswith('ABWDST')]
if abwd_cols:
    abawd_matrix = d[abwd_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
    pct_abawd = (abawd_matrix.isin([2, 3, 4, 5]).sum(axis=1) / certhhsz).values
else:
    pct_abawd = np.zeros(len(d))

expedited_i = (pd.to_numeric(d['EXPEDSER'], errors='coerce').fillna(3) < 3).astype(int)

# rule variables beyond the Virginia builder --------------------------------
total_deductions = (rawdepded + rawcsded + rawsltded + rawmedded + rawernded)
hh_size_n = size.values.astype(float)

rel_cols = [c for c in d.columns if re.fullmatch(r'REL\d+', c)]
rel = d[rel_cols].apply(pd.to_numeric, errors='coerce')
married = (rel == 2).any(axis=1).astype(int)          # any spouse relationship

children_i = (pd.to_numeric(d['FSNKID'], errors='coerce').fillna(0) > 0).astype(int)
homeless   = (pd.to_numeric(d['HOMEDED'], errors='coerce').fillna(1) != 1).astype(int)

INCOME_VARS = ['FSWAGES','FSSLFEMP','FSOTHERN','FSSSI','FSTANF','FSGA','FSSOCSEC',
               'FSUNEMP','FSVET','FSWCOMP','FSEDLOAN','FSCSUPRT','FSDEEM','FSCONT',
               'FSOTHGOV','FSOTHUN','FSDIVER','FSWGESUP','FSENERGY','FSEITC','FSFOSTER']
DEDUCT_VARS = ['FSSTDDED','FSERNDED','FSDEPDED','FSSLTDED','FSMEDDED','FSCSDED',
               'HOMELESS_DED']
div_cols = [c for c in INCOME_VARS + DEDUCT_VARS if c in d.columns]
divm = d[div_cols].apply(pd.to_numeric, errors='coerce')
count_div_100 = ((divm > 0) & (divm % 100 == 0)).sum(axis=1).astype(int)

out = pd.DataFrame({
    'fiscal_year':        fy,
    'hh_size_raw':        size.values,
    'hh_group':           np.where(size <= 1, '1', np.where(size <= 3, '2-3', '4+')),
    'HH_size_n':          hh_size_n,
    'rawben_rel_max':     (rawben_rec / rawbenmax).round(4),
    'unc_rawben_rel_max': (rawben_unc / rawbenmax).round(4),
    'medical_deductions': rawmedded.values.round(0),
    'total_deductions_by_hh_size':    (total_deductions / hh_size_n).round(2),
    'shelter_expenses_by_hh_size':    (rawsltexp / hh_size_n).round(2),
    'utilities':          rawutil.values.round(0),
    'months_since_cert_n': months_since_cert.values,
    'percent_abawd':      np.round(pct_abawd, 4),
    'expedited_i':        expedited_i.values,
    'elderly_disabled_i': ((elder > 0) | (dis > 0)).astype(int).values,
    'cat_elig':           pd.to_numeric(d['CAT_ELIG'], errors='coerce').fillna(0).values,
    'married':            married.values,
    'children_i':         children_i.values,
    'homeless':           homeless.values,
    'count_divisible_by_100': count_div_100.values,
})

absdiff = (pd.to_numeric(d['RAWBEN'], errors='coerce').fillna(0)
           - pd.to_numeric(d['FSBEN'], errors='coerce').fillna(0)).abs()
amterr  = pd.to_numeric(d['AMTERR'], errors='coerce').fillna(0)
thr     = np.array([err_thr.get(int(y), 37) for y in fy])
out['over_threshold']     = (absdiff.values > thr).astype(int)
out['total_error_amount'] = amterr.values.round(0)

# keep only rows where the benefit-error difference is consistent (as in R)
keep = (np.abs(absdiff.values - amterr.values) <= 5)
out  = out[keep].reset_index(drop=True)
print('after benefit-consistency filter:', len(out),
      '| error rate:', round(out.over_threshold.mean(), 3))
out.to_csv(CASES_CSV, index=False)
df = out

# ══════════════════════════════════════════════════════════════════════════════
# 2. PARSE THE DELIVERY RULES
# ══════════════════════════════════════════════════════════════════════════════
rules_df = pd.read_csv(DELIVERY_CSV)
rules_df = rules_df[rules_df['role'] == ROLE_FILTER].sort_values('rank').reset_index(drop=True)
COND_PAT = re.compile(r'([A-Za-z_][A-Za-z0-9_]*)\s*(>=|<=|>|<|==)\s*(-?[0-9.]+)')

RULES = []
for _, rr in rules_df.iterrows():
    conds = [{'var': v, 'op': op, 'thr': float(t)} for v, op, t in COND_PAT.findall(rr['rule'])]
    assert 1 <= len(conds) <= NSLOTS, rr['rule']
    RULES.append({'num': int(rr['rank']), 'hh': str(rr['hh']), 'conds': conds,
                  'prec_train': float(rr['precision_train']),
                  'prec_lcb': float(rr['precision_train_lcb']),
                  'engine': rr['engines'], 'frame': rr['mined_frames']})
print('rules implemented:', len(RULES))

RULE_VARS = sorted({c['var'] for r in RULES for c in r['conds']})
for v in RULE_VARS:
    assert v in df.columns, f'missing variable: {v}'

# snapped_grid steps by variable type (dollar 50, ratio 0.05, months/counts 1)
RATIO_VARS = {'rawben_rel_max', 'unc_rawben_rel_max', 'percent_abawd'}
UNIT_VARS  = {'months_since_cert_n', 'HH_size_n', 'count_divisible_by_100', 'cat_elig',
              'expedited_i', 'elderly_disabled_i', 'married', 'children_i', 'homeless'}
def base_step(v):
    if v in RATIO_VARS: return 0.05
    if v in UNIT_VARS:  return 1.0
    return 50.0

# ══════════════════════════════════════════════════════════════════════════════
# 3. WORKBOOK SCAFFOLDING
# ══════════════════════════════════════════════════════════════════════════════
DCOLS = ['fiscal_year', 'hh_size_raw', 'hh_group'] + RULE_VARS + \
        ['over_threshold', 'total_error_amount']
LASTCOL = get_column_letter(len(DCOLS))

def dc(varname): return get_column_letter(DCOLS.index(varname) + 1)
def dr(varname):
    c = dc(varname)
    return f'Data!${c}$2:${c}${MAX_ROW}'

YELLOW     = PatternFill('solid', fgColor='FFFF99')
BLUE_LIGHT = PatternFill('solid', fgColor='BDD7EE')
BLUE_DARK  = PatternFill('solid', fgColor='2F5496')
GRAY       = PatternFill('solid', fgColor='F2F2F2')
GREEN      = PatternFill('solid', fgColor='E2EFDA')
WHITE      = PatternFill('solid', fgColor='FFFFFF')
FONT = 'Calibri'
def bold_font(size=11, color='000000'): return Font(name=FONT, bold=True, size=size, color=color)
center = Alignment(horizontal='center', vertical='center')
left   = Alignment(horizontal='left',   vertical='center')

def thin():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value=None, formula=None, fill=None, font=None,
             align=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col)
    cell.value = formula if formula else value
    if fill:   cell.fill = fill
    if font:   cell.font = font
    if align:  cell.alignment = align
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell

def merge(ws, r1, c1, r2, c2, value=None, fill=None, font=None, align=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    if value is not None: cell.value = value
    if fill:  cell.fill = fill
    if font:  cell.font = font
    if align: cell.alignment = align
    return cell

wb = openpyxl.Workbook()

# ── Data sheet ────────────────────────────────────────────────────────────────
ws_data = wb.create_sheet('Data')
dfx = df[DCOLS]
for ci, col in enumerate(DCOLS, 1):
    c = ws_data.cell(row=1, column=ci, value=col)
    c.fill = BLUE_DARK
    c.font = Font(name=FONT, bold=True, color='FFFFFF')
    c.alignment = center
for ri, row in enumerate(dfx.itertuples(index=False), 2):
    for ci, val in enumerate(row, 1):
        ws_data.cell(row=ri, column=ci, value=val)
ws_data.freeze_panes = 'A2'

# ══════════════════════════════════════════════════════════════════════════════
# 4. DASHBOARD — one tuning block per delivery rule
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Dashboard'
ws.sheet_view.showGridLines = False
for col_letter, width in {'A':34,'B':8,'C':16,'D':14,'E':3,'F':18,'G':13,'H':13,'I':3,
                          'J':9,'K':7,'L':12,'M':12,'N':12,'O':12,'P':12,'Q':10}.items():
    ws.column_dimensions[col_letter].width = width

merge(ws,1,1,1,9,
      value=f'SNAP QC Inclusion Rule Tuner — {STATE_NAME} (FY2022–2024) — '
            f'{len(RULES)} core delivery rules',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=16,color='FFFFFF'), align=center)
ws.row_dimensions[1].height = 32

# stratum summary (referenced by every rule block)
merge(ws,3,1,3,9, value='Dataset Summary by Household Size',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
for col, txt in [(1,'HH size'),(2,''),(3,'Cases'),(4,'Errors'),(6,'Error $'),(7,'Base error rate')]:
    if txt:
        set_cell(ws,4,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())
S_ROWS = {}   # stratum -> summary row
for s, lbl in enumerate(STRATA):
    r = 5 + s
    S_ROWS[lbl] = r
    set_cell(ws,r,1,lbl, font=bold_font(), align=center, border=thin())
    set_cell(ws,r,3,formula=f'=COUNTIF({dr("hh_group")},"{lbl}")',
             fill=BLUE_LIGHT, align=center, border=thin(), number_format='#,##0')
    set_cell(ws,r,4,formula=f'=COUNTIFS({dr("hh_group")},"{lbl}",{dr("over_threshold")},1)',
             fill=BLUE_LIGHT, align=center, border=thin(), number_format='#,##0')
    set_cell(ws,r,6,formula=f'=SUMIFS({dr("total_error_amount")},{dr("hh_group")},"{lbl}",{dr("over_threshold")},1)',
             fill=BLUE_LIGHT, align=center, border=thin(), number_format='$#,##0')
    set_cell(ws,r,7,formula=f'=D{r}/C{r}', fill=BLUE_LIGHT, align=center,
             border=thin(), number_format='0.0%')

def rule_countifs(rule, thr_cells, add_error=False):
    parts = [f'{dr("hh_group")},"{rule["hh"]}"']
    for c, tc in zip(rule['conds'], thr_cells):
        parts.append(f'{dr(c["var"])},"{c["op"]}"&{tc}')
    if add_error:
        parts.append(f'{dr("over_threshold")},1')
    return f'COUNTIFS({",".join(parts)})'

def rule_sumifs(rule, thr_cells):
    parts = [f'{dr("total_error_amount")},{dr("hh_group")},"{rule["hh"]}",{dr("over_threshold")},1']
    for c, tc in zip(rule['conds'], thr_cells):
        parts.append(f'{dr(c["var"])},"{c["op"]}"&{tc}')
    return f'SUMIFS({",".join(parts)})'

METRIC_NAMES = ['n_flagged','errors caught','precision','recall (count)','dollar recall','lift']
METRIC_FMTS  = ['#,##0','#,##0','0.0%','0.0%','0.0%','0.00']
BLOCK_HEIGHT = 10
BASE_ROW = 9
summary_cells = []   # per rule: dict of metric cell refs

for bi, rule in enumerate(RULES):
    r0 = BASE_ROW + bi * BLOCK_HEIGHT
    merge(ws,r0,1,r0,9,
          value=(f'Rule {rule["num"]}  │  HH size {rule["hh"]}  │  {rule["engine"]}, '
                 f'{rule["frame"]}  │  train precision {rule["prec_train"]:.1%} '
                 f'(LCB {rule["prec_lcb"]:.1%})'),
          fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=11,color='FFFFFF'), align=left)
    r1 = r0 + 1
    for col, txt in [(1,'Variable'),(2,'Op'),(3,'Threshold ← EDIT'),(4,'Original'),
                     (6,'Metric'),(7,'This state'),(8,'Train (natl.)')]:
        set_cell(ws,r1,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())

    thr_cells = []
    for ci, cond in enumerate(rule['conds']):
        rc = r0 + 2 + ci
        set_cell(ws,rc,1,cond['var'], align=left, border=thin())
        set_cell(ws,rc,2,cond['op'], align=center, border=thin())
        set_cell(ws,rc,3,cond['thr'], fill=YELLOW, align=center, border=thin(), font=bold_font())
        set_cell(ws,rc,4,cond['thr'], align=center, border=thin(),
                 font=Font(name=FONT,color='808080'))
        thr_cells.append(f'C{rc}')

    srow = S_ROWS[rule['hh']]
    flag_f, err_f = rule_countifs(rule, thr_cells), rule_countifs(rule, thr_cells, True)
    dol_f = rule_sumifs(rule, thr_cells)
    formulas = [
        f'={flag_f}',
        f'={err_f}',
        f'=IFERROR({err_f}/{flag_f},0)',
        f'=IFERROR({err_f}/$D${srow},0)',
        f'=IFERROR({dol_f}/$F${srow},0)',
        f'=IFERROR(({err_f}/{flag_f})/($D${srow}/$C${srow}),0)',
    ]
    nat = [None, None, rule['prec_train'], None, None, None]
    cells = {}
    for mi, (mname, mfmt, mf) in enumerate(zip(METRIC_NAMES, METRIC_FMTS, formulas)):
        rm = r0 + 2 + mi
        set_cell(ws,rm,6,mname, font=bold_font(10), fill=GRAY, align=left, border=thin())
        set_cell(ws,rm,7,formula=mf, fill=BLUE_LIGHT, align=center, border=thin(),
                 number_format=mfmt)
        if nat[mi] is not None:
            set_cell(ws,rm,8,nat[mi], align=center, border=thin(), number_format=mfmt,
                     font=Font(name=FONT,color='808080'))
        cells[mname] = f'G{rm}'
    summary_cells.append(cells)

# rule comparison table (right side, below the chart)
SUMM_ROW = 26
merge(ws,SUMM_ROW,10,SUMM_ROW,17, value='Rule Comparison Summary',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
for off, hdr in enumerate(['Rule','HH','Train prec.','Precision','Recall','$ Recall',
                           'n_flagged','Errors']):
    set_cell(ws,SUMM_ROW+1,10+off,hdr, font=bold_font(10), fill=GRAY, align=center, border=thin())
for bi, rule in enumerate(RULES):
    sr = SUMM_ROW + 2 + bi
    cc = summary_cells[bi]
    fill_r = BLUE_LIGHT if bi % 2 == 0 else WHITE
    vals = [(f'Rule {rule["num"]}', None), (rule['hh'], None),
            (rule['prec_train'], '0.0%'), (f'={cc["precision"]}', '0.0%'),
            (f'={cc["recall (count)"]}', '0.0%'), (f'={cc["dollar recall"]}', '0.0%'),
            (f'={cc["n_flagged"]}', '#,##0'), (f'={cc["errors caught"]}', '#,##0')]
    for off, (v, fmt) in enumerate(vals):
        c = set_cell(ws,sr,10+off,formula=v if isinstance(v,str) and v.startswith('=') else None,
                     value=None if isinstance(v,str) and v.startswith('=') else v,
                     fill=fill_r, align=center, border=thin())
        if fmt: c.number_format = fmt

# precision-recall chart (79 points; labels off — identify via the table)
chart = ScatterChart()
chart.title = 'Precision vs Recall (tuned thresholds)'
chart.x_axis.title = 'Recall'; chart.y_axis.title = 'Precision'
chart.x_axis.numFmt = '0%';   chart.y_axis.numFmt = '0%'
chart.width, chart.height = 15, 12
for bi, rule in enumerate(RULES):
    rec_c, prec_c = summary_cells[bi]['recall (count)'], summary_cells[bi]['precision']
    xv = Reference(ws, min_col=7, min_row=int(rec_c[1:]),  max_row=int(rec_c[1:]))
    yv = Reference(ws, min_col=7, min_row=int(prec_c[1:]), max_row=int(prec_c[1:]))
    s = Series(yv, xv, title=f'Rule {rule["num"]}')
    s.marker.symbol, s.marker.size = 'circle', 7
    s.graphicalProperties.line.noFill = True
    chart.series.append(s)
chart.legend = None
ws.add_chart(chart, 'J3')
ws.freeze_panes = 'A9'

# ══════════════════════════════════════════════════════════════════════════════
# 5. GRID SEARCH SHEET (same engine as the Virginia workbook)
# ══════════════════════════════════════════════════════════════════════════════
DATA2D = f'Data!$A$2:${LASTCOL}${MAX_ROW}'
HH     = dr('hh_group')
OVER   = dr('over_threshold')
ERRD   = dr('total_error_amount')

HELP0 = 100
GRID0 = HELP0 + 3*6 + 2
def h0(s):  return HELP0 + s*6
def gs(s):  return GRID0 + s*(MAXG + 3)
def rng(s, col):
    top = gs(s) + 2
    return f'${col}${top}:${col}${top + MAXG - 1}'
def sref(i, col): return f'${col}${7 + i}'

def dyn_pairs(lbl, thr_refs):
    parts = []
    for i in range(NSLOTS):
        parts.append(f'INDEX({DATA2D},0,{sref(i,"AF")})')
        parts.append(f'IF({sref(i,"A")}="","{lbl}",{sref(i,"B")}&{thr_refs[i]})')
    return ','.join(parts)

# preset steps: double the widest non-binary grid until every stratum fits MAXG
def py_grid(x, step):
    x = x[np.isfinite(x)]
    lo = math.floor(np.percentile(x, 2) / step) * step
    hi = max(-math.floor(-np.percentile(x, 98) / step) * step, lo + step)
    n = int(round((hi - lo) / step)) + 1
    return [round(lo + j * step, 6) for j in range(n)]

for rule in RULES:
    steps = {c['var']: base_step(c['var']) for c in rule['conds']}
    def worst():
        return max(np.prod([len(py_grid(df[df.hh_group == g][c['var']].dropna().values.astype(float),
                                        steps[c['var']])) for c in rule['conds']])
                   for g in STRATA)
    while worst() > MAXG:
        cand = {c['var']: max(len(py_grid(df[df.hh_group == g][c['var']].dropna().values.astype(float),
                                          steps[c['var']])) for g in STRATA)
                for c in rule['conds'] if c['var'] not in UNIT_VARS}
        if not cand: break
        v = max(cand, key=cand.get)
        steps[v] *= 2
    rule['steps'] = steps

ws_g = wb.create_sheet('Grid Search', 1)
ws_g.sheet_view.showGridLines = False
for col_letter, width in {'A':26,'B':13,'C':13,'D':13,'E':13,'F':12,'G':12,
                          'H':13,'I':13,'J':13,'K':12,'L':10,'M':8}.items():
    ws_g.column_dimensions[col_letter].width = width
for c in ('Y','Z','AA','AB','AC','AD','AE','AF','L','M'):
    ws_g.column_dimensions[c].hidden = True

merge(ws_g,1,1,1,13, value=f'Grid Search — Optimize One Delivery Rule ({STATE_NAME})',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=16,color='FFFFFF'), align=center)
ws_g.row_dimensions[1].height = 32

# hidden preset tables: names in AA, terms in AB:AE; dropdown list in Y
NP = len(RULES)
set_cell(ws_g,1,25,'Custom')                          # Y1
for p, rule in enumerate(RULES):
    pname = f'Rule {rule["num"]} (HH {rule["hh"]})'
    set_cell(ws_g,2+p,27,pname)                       # AA
    set_cell(ws_g,2+p,25,pname)                       # Y (dropdown source)
    for i in range(NSLOTS):
        rr = 2 + p*NSLOTS + i
        if i < len(rule['conds']):
            c = rule['conds'][i]
            set_cell(ws_g,rr,28,c['var']); set_cell(ws_g,rr,29,c['op'])
            set_cell(ws_g,rr,30,rule['steps'][c['var']]); set_cell(ws_g,rr,31,c['thr'])

merge(ws_g,4,1,4,13,
      value='Step 1 — Pick a rule preset, or pick "Custom" and fill the custom block on the right',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
set_cell(ws_g,5,1,'Rule preset', font=bold_font(), fill=GRAY, align=left)
set_cell(ws_g,5,2,f'Rule {RULES[0]["num"]} (HH {RULES[0]["hh"]})', fill=YELLOW,
         align=center, border=thin(), font=bold_font(), number_format='@')
set_cell(ws_g,5,4,'← delivery rules ranked by train precision LCB, or "Custom"',
         font=Font(name=FONT,size=9,color='808080'), align=left)
merge(ws_g,5,8,5,11, value='Custom rule (used when preset = Custom)',
      fill=GRAY, font=bold_font(10), align=center)
for col, txt in [(1,'Variable'),(2,'Op'),(3,'Step'),(4,'Original threshold'),
                 (8,'Variable'),(9,'Op'),(10,'Step'),(11,'Original')]:
    set_cell(ws_g,6,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())

PLOOK = f'(MATCH($B$5,$AA$2:$AA${1+NP},0)-1)*{NSLOTS}'
PEND = 1 + NP*NSLOTS
for i in range(NSLOTS):
    r = 7 + i
    vlook = f'INDEX($AB$2:$AB${PEND},{PLOOK}+{i+1})'
    set_cell(ws_g,r,1,
             formula=f'=IF($B$5="Custom",IF($H{r}="","",$H{r}),IF({vlook}=0,"",{vlook}))',
             fill=BLUE_LIGHT, align=left, border=thin())
    for cc, pcol, ccol in [(2,'AC','I'),(3,'AD','J'),(4,'AE','K')]:
        set_cell(ws_g,r,cc,
                 formula=(f'=IF($A{r}="","",IF($B$5="Custom",${ccol}{r},'
                          f'INDEX(${pcol}$2:${pcol}${PEND},{PLOOK}+{i+1})))'),
                 fill=BLUE_LIGHT, align=center, border=thin())
    set_cell(ws_g,r,32,formula=f'=IF($A{r}="",3,MATCH($A{r},Data!$A$1:${LASTCOL}$1,0))')
    if i < len(RULES[0]['conds']):
        c = RULES[0]['conds'][i]
        set_cell(ws_g,r,8, c['var'], fill=YELLOW, align=left,   border=thin(), font=bold_font())
        set_cell(ws_g,r,9, c['op'],  fill=YELLOW, align=center, border=thin(), font=bold_font())
        set_cell(ws_g,r,10,RULES[0]['steps'][c['var']], fill=YELLOW, align=center,
                 border=thin(), font=bold_font())
        set_cell(ws_g,r,11,c['thr'], fill=YELLOW, align=center, border=thin(), font=bold_font())
    else:
        for cc in (8,9,10,11):
            set_cell(ws_g,r,cc,None, fill=YELLOW, align=center, border=thin(), font=bold_font())

for i, v in enumerate(RULE_VARS, 2):
    set_cell(ws_g,i,26,v)                             # Z: variable dropdown source
dv_preset = DataValidation(type='list', formula1=f'=$Y$1:$Y${1+NP}', allow_blank=False)
ws_g.add_data_validation(dv_preset); dv_preset.add('B5')
dv_var = DataValidation(type='list', formula1=f'=$Z$2:$Z${len(RULE_VARS)+1}', allow_blank=True)
ws_g.add_data_validation(dv_var); dv_var.add('H7:H10')
dv_op = DataValidation(type='list', formula1='">=,<=,>,<,="', allow_blank=True)
ws_g.add_data_validation(dv_op); dv_op.add('I7:I10')

merge(ws_g,11,1,11,13, value='Step 2 — Settings',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
set_cell(ws_g,12,1,'Precision floor', font=bold_font(), fill=GRAY, align=left)
set_cell(ws_g,12,2,PREC_FLOOR, fill=YELLOW, align=center, border=thin(), font=bold_font(),
         number_format='0.0%')
set_cell(ws_g,12,4,'← minimum precision: at least this share of flagged cases must be true errors',
         font=Font(name=FONT,size=9,color='808080'), align=left)
set_cell(ws_g,13,1,'Recall basis (maximized)', font=bold_font(), fill=GRAY, align=left)
set_cell(ws_g,13,2,'dollar', fill=YELLOW, align=center, border=thin(), font=bold_font())
set_cell(ws_g,13,4,'← what to maximize: "dollar" = share of error $ caught, "count" = share of error cases',
         font=Font(name=FONT,size=9,color='808080'), align=left)
set_cell(ws_g,12,8,'Min flagged', font=bold_font(), fill=GRAY, align=left)
set_cell(ws_g,12,9,MIN_FLAGGED, fill=YELLOW, align=center, border=thin(), font=bold_font(),
         number_format='0')
set_cell(ws_g,12,10,'← ignore combos flagging fewer cases',
         font=Font(name=FONT,size=9,color='808080'), align=left)
set_cell(ws_g,13,8,'Recall floor', font=bold_font(), fill=GRAY, align=left)
set_cell(ws_g,13,9,RECALL_FLOOR, fill=YELLOW, align=center, border=thin(), font=bold_font(),
         number_format='0.0%')
set_cell(ws_g,13,10,'← combo must catch at least this share of error $ / cases (chosen basis)',
         font=Font(name=FONT,size=9,color='808080'), align=left)
dv_floor = DataValidation(type='decimal', operator='between', formula1='0', formula2='1')
ws_g.add_data_validation(dv_floor); dv_floor.add('B12')
ws_g.add_data_validation(dv_floor2 := DataValidation(type='decimal', operator='between',
                                                     formula1='0', formula2='1'))
dv_floor2.add('I13')
dv_basis = DataValidation(type='list', formula1='"dollar,count"', allow_blank=False)
ws_g.add_data_validation(dv_basis); dv_basis.add('B13')
dv_minf = DataValidation(type='whole', operator='greaterThanOrEqual', formula1='0')
ws_g.add_data_validation(dv_minf); dv_minf.add('I12')

merge(ws_g,15,1,15,13, value='Stratum Overview',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
for col, txt in [(1,'HH size'),(2,'Cases'),(3,'Errors'),(4,'Error $'),(5,'Combinations')]:
    set_cell(ws_g,16,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())
for s, lbl in enumerate(STRATA):
    r = 17 + s
    set_cell(ws_g,r,1,lbl, font=bold_font(), align=center, border=thin())
    set_cell(ws_g,r,2,formula=f'=COUNTIF({HH},"{lbl}")', align=center, border=thin(),
             number_format='#,##0', fill=BLUE_LIGHT)
    set_cell(ws_g,r,3,formula=f'=COUNTIFS({HH},"{lbl}",{OVER},1)', align=center,
             border=thin(), number_format='#,##0', fill=BLUE_LIGHT)
    set_cell(ws_g,r,4,formula=f'=SUMIFS({ERRD},{HH},"{lbl}",{OVER},1)', align=center,
             border=thin(), number_format='$#,##0', fill=BLUE_LIGHT)
    set_cell(ws_g,r,5,formula=f'=IF($A$7="","",$H${h0(s)})', align=center,
             border=thin(), number_format='#,##0', fill=BLUE_LIGHT)

F0 = 31
def fs(s): return F0 + s*18

merge(ws_g,21,1,21,13,
      value='Best per Stratum — max recall with precision ≥ floor (vs original thresholds)',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
for col, txt in [(1,'HH size'),(2,'thr 1'),(3,'thr 2'),(4,'thr 3'),(5,'thr 4'),
                 (6,'precision'),(7,'recall'),(8,'dollar recall'),(9,'n_flagged'),
                 (10,'errors'),(11,'workload %')]:
    set_cell(ws_g,22,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())
BEST_FMTS = {6:'0.0%',7:'0.0%',8:'0.0%',9:'#,##0',10:'#,##0'}
for s, lbl in enumerate(STRATA):
    r = 23 + s
    M = f'$M${fs(s)+2}'
    set_cell(ws_g,r,1,f'{lbl} (best)', font=bold_font(), align=center, border=thin())
    for ci, col in enumerate('ABCD'):
        set_cell(ws_g,r,2+ci,
                 formula=f'=IF({M}="","",IF({sref(ci,"A")}="","",INDEX({rng(s,col)},{M})))',
                 fill=GREEN, align=center, border=thin(), font=bold_font())
    for cnum, gcol in [(6,'H'),(7,'I'),(8,'J'),(9,'E'),(10,'F')]:
        set_cell(ws_g,r,cnum,formula=f'=IF({M}="","",INDEX({rng(s,gcol)},{M}))',
                 fill=BLUE_LIGHT, align=center, border=thin(), number_format=BEST_FMTS[cnum])
    set_cell(ws_g,r,11,formula=f'=IF({M}="","",INDEX({rng(s,"E")},{M})/$B${17+s})',
             fill=BLUE_LIGHT, align=center, border=thin(), number_format='0.0%')

ORIG_REFS = [sref(i,'D') for i in range(NSLOTS)]
for s, lbl in enumerate(STRATA):
    r = 26 + s
    gray = Font(name=FONT, color='808080')
    set_cell(ws_g,r,1,f'{lbl} (original)', align=center, border=thin(), font=gray)
    for i in range(NSLOTS):
        set_cell(ws_g,r,2+i,formula=f'=IF({sref(i,"A")}="","",{sref(i,"D")})',
                 align=center, border=thin(), font=gray)
    pairs = dyn_pairs(lbl, ORIG_REFS)
    set_cell(ws_g,r,9, formula=f'=IF($A$7="","",COUNTIFS({HH},"{lbl}",{pairs}))',
             align=center, border=thin(), number_format='#,##0', font=gray)
    set_cell(ws_g,r,10,formula=f'=IF($A$7="","",COUNTIFS({HH},"{lbl}",{pairs},{OVER},1))',
             align=center, border=thin(), number_format='#,##0', font=gray)
    set_cell(ws_g,r,12,formula=f'=IF($A$7="","",SUMIFS({ERRD},{HH},"{lbl}",{OVER},1,{pairs}))')
    set_cell(ws_g,r,6, formula=f'=IF($I{r}="","",IF($I{r}=0,0,$J{r}/$I{r}))',
             align=center, border=thin(), number_format='0.0%', font=gray)
    set_cell(ws_g,r,7, formula=f'=IF($J{r}="","",IF($C${17+s}=0,0,$J{r}/$C${17+s}))',
             align=center, border=thin(), number_format='0.0%', font=gray)
    set_cell(ws_g,r,8, formula=f'=IF($L{r}="","",IF($D${17+s}=0,0,$L{r}/$D${17+s}))',
             align=center, border=thin(), number_format='0.0%', font=gray)
    set_cell(ws_g,r,11,formula=f'=IF($I{r}="","",$I{r}/$B${17+s})',
             align=center, border=thin(), number_format='0.0%', font=gray)

merge(ws_g,30,1,30,13,
      value='Feasible Combinations — top 15 by recall per stratum (meeting the precision floor)',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=12,color='FFFFFF'), align=left)
FEAS_HEADS = [(1,'rank'),(2,'thr 1'),(3,'thr 2'),(4,'thr 3'),(5,'thr 4'),
              (6,'precision'),(7,'recall'),(8,'dollar recall'),(9,'n_flagged'),(10,'errors')]
for s, lbl in enumerate(STRATA):
    f0 = fs(s)
    merge(ws_g,f0,1,f0,13, value=f'HH size {lbl}',
          fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=10,color='FFFFFF'), align=left)
    for col, txt in FEAS_HEADS:
        set_cell(ws_g,f0+1,col,txt, font=bold_font(9), fill=GRAY, align=center, border=thin())
    for k in range(1, 16):
        r = f0 + 1 + k
        set_cell(ws_g,r,1,k, align=center, border=thin())
        set_cell(ws_g,r,13,
                 formula=f'=IFERROR(MATCH(LARGE({rng(s,"K")},{k}),{rng(s,"K")},0),"")')
        for ci, col in enumerate('ABCD'):
            set_cell(ws_g,r,2+ci,
                     formula=f'=IF($M{r}="","",IF({sref(ci,"A")}="","",INDEX({rng(s,col)},$M{r})))',
                     align=center, border=thin())
        for cnum, gcol, fmt in [(6,'H','0.0%'),(7,'I','0.0%'),(8,'J','0.0%'),
                                (9,'E','#,##0'),(10,'F','#,##0')]:
            set_cell(ws_g,r,cnum,
                     formula=f'=IF($M{r}="","",INDEX({rng(s,gcol)},$M{r}))',
                     align=center, border=thin(), number_format=fmt)

NOTES = [
    'Each preset loads one delivery rule with pre-tuned snapped_grid steps; the search runs in all three strata.',
    'The stratum matching the rule\'s "HH" tag is the one the delivery list targets.',
    f'Capacity: up to {MAXG:,} combinations per stratum; over capacity everything goes blank — increase steps.',
    'Best row = highest recall (chosen basis) among combos with precision >= floor, n_flagged >= min flagged, recall >= recall floor; ties by grid order.',
    f'The full scored grid lives in the blocks from row {GRID0} down (helper rows 100-119 are hidden).',
]
for i, txt in enumerate(NOTES):
    set_cell(ws_g,85+i,1,txt, font=Font(name=FONT,size=9,color='808080'), align=left)

for s, lbl in enumerate(STRATA):
    hh0 = h0(s)
    set_cell(ws_g,hh0,8,formula=f'=F{hh0+1}*F{hh0+2}*F{hh0+3}*F{hh0+4}')
    for i in range(NSLOTS):
        hr = hh0 + 1 + i
        V, S, C = sref(i,'A'), sref(i,'C'), sref(i,'AF')
        q02 = f'=IF({V}="","",PERCENTILE(IF({HH}="{lbl}",INDEX({DATA2D},0,{C})),0.02))'
        q98 = f'=IF({V}="","",PERCENTILE(IF({HH}="{lbl}",INDEX({DATA2D},0,{C})),0.98))'
        ws_g.cell(row=hr, column=2).value = ArrayFormula(f'B{hr}', q02)
        ws_g.cell(row=hr, column=3).value = ArrayFormula(f'C{hr}', q98)
        ws_g.cell(row=hr, column=4).value = f'=IF({V}="","",INT(B{hr}/{S})*{S})'
        ws_g.cell(row=hr, column=5).value = f'=IF({V}="","",MAX(-INT(-C{hr}/{S})*{S},D{hr}+{S}))'
        ws_g.cell(row=hr, column=6).value = f'=IF({V}="",1,ROUND((E{hr}-D{hr})/{S},0)+1)'
        ws_g.cell(row=hr, column=7).value = ('=1' if i == 0 else f'=F{hr-1}*G{hr-1}')
for rr in range(HELP0, GRID0):
    ws_g.row_dimensions[rr].hidden = True

for s, lbl in enumerate(STRATA):
    g0 = gs(s); data0 = g0 + 2; hh0 = h0(s)
    merge(ws_g,g0,1,g0,13, value=f'GRID — HH size {lbl}',
          fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=10,color='FFFFFF'), align=left)
    for ci, htxt in enumerate(['thr 1','thr 2','thr 3','thr 4','n_flagged','errors_caught',
                               'err_dollars','precision','recall_count','dollar_recall',
                               'sortkey'], 1):
        set_cell(ws_g,g0+1,ci,htxt, font=bold_font(9), fill=GRAY, align=center)
    for r in range(data0, data0 + MAXG):
        for i in range(NSLOTS):
            hr = hh0 + 1 + i
            V, S = sref(i,'A'), sref(i,'C')
            ws_g.cell(row=r, column=1+i).value = (
                f'=IF($A$7="","",IF($H${hh0}>{MAXG},"",IF(ROW()-{g0+1}>$H${hh0},"",'
                f'IF({V}="","",ROUND($D${hr}+{S}*MOD(INT((ROW()-{data0})/$G${hr}),$F${hr}),6)))))')
        pairs = dyn_pairs(lbl, [f'$A{r}', f'$B{r}', f'$C{r}', f'$D{r}'])
        ws_g.cell(row=r, column=5).value  = f'=IF($A{r}="","",COUNTIFS({HH},"{lbl}",{pairs}))'
        ws_g.cell(row=r, column=6).value  = f'=IF($A{r}="","",COUNTIFS({HH},"{lbl}",{pairs},{OVER},1))'
        ws_g.cell(row=r, column=7).value  = f'=IF($A{r}="","",SUMIFS({ERRD},{HH},"{lbl}",{OVER},1,{pairs}))'
        ws_g.cell(row=r, column=8).value  = f'=IF($E{r}="","",IF($E{r}=0,0,$F{r}/$E{r}))'
        ws_g.cell(row=r, column=9).value  = f'=IF($E{r}="","",IF($C${17+s}=0,0,$F{r}/$C${17+s}))'
        ws_g.cell(row=r, column=10).value = f'=IF($E{r}="","",IF($D${17+s}=0,0,$G{r}/$D${17+s}))'
        ws_g.cell(row=r, column=11).value = (
            f'=IF($H{r}="","",IF(AND($H{r}>=$B$12,$E{r}>=$I$12,'
            f'IF($B$13="dollar",$J{r},$I{r})>=$I$13),'
            f'IF($B$13="dollar",$J{r},$I{r})+({MAXG}-(ROW()-{g0+1}))*0.0000000001,""))')
        for cc in (8, 9, 10, 11):
            ws_g.cell(row=r, column=cc).number_format = '0.000'

ws_g.freeze_panes = 'A21'
wb.calculation.fullCalcOnLoad = True

# ══════════════════════════════════════════════════════════════════════════════
# 6. SUMMARY SHEET — optimal vs original per rule, on its designated stratum
# ══════════════════════════════════════════════════════════════════════════════
def py_score(dd, conds, thresholds):
    is_err = (dd['over_threshold'] == 1).values
    ed = np.where(is_err, dd['total_error_amount'].fillna(0).values, 0)
    m = np.ones(len(dd), bool)
    for c, t in zip(conds, thresholds):
        xv = dd[c['var']].values.astype(float)
        if   c['op'] == '>=': cm = xv >= t
        elif c['op'] == '>':  cm = xv > t
        elif c['op'] == '<=': cm = xv <= t
        else:                 cm = xv < t
        m &= np.where(np.isnan(xv), False, cm)
    n, tp = int(m.sum()), int((m & is_err).sum())
    return {'prec': tp / n if n else 0.0, 'rec': tp / max(is_err.sum(), 1),
            'drec': ed[m].sum() / max(ed.sum(), 1e-9), 'n': n, 'tp': tp,
            'dollars': ed[m].sum()}

summary_rows = []
rule_best_thr = {}   # rule index -> optimal thresholds tuple (None if none meets floor)
for ri, rule in enumerate(RULES):
    dd = df[df['hh_group'] == rule['hh']]
    grids = [py_grid(dd[c['var']].dropna().values.astype(float), rule['steps'][c['var']])
             for c in rule['conds']]
    best = None
    for combo in itertools.product(*reversed(grids)):
        combo = tuple(reversed(combo))
        sc = py_score(dd, rule['conds'], combo)
        if (sc['prec'] >= PREC_FLOOR and sc['n'] >= MIN_FLAGGED
                and sc['drec'] >= RECALL_FLOOR
                and (best is None or sc['drec'] > best[1]['drec'])):
            best = (combo, sc)
    base = py_score(dd, rule['conds'], [c['thr'] for c in rule['conds']])
    rule_best_thr[ri] = best[0] if best else None
    opt_cond = (' & '.join(f'{c["var"]} {c["op"]} {t:g}'
                           for c, t in zip(rule['conds'], best[0])) if best
                else 'no combination meets the floors')
    orig_cond = ' & '.join(f'{c["var"]} {c["op"]} {c["thr"]:g}' for c in rule['conds'])
    summary_rows.append((rule, 'optimal', best[1] if best else None, opt_cond))
    summary_rows.append((rule, 'original', base, orig_cond))

# ── union of all rules (a case is flagged if ANY rule flags it on its stratum) ─
def rule_mask_full(rule, thresholds):
    m = (df['hh_group'] == rule['hh']).values.copy()
    for c, t in zip(rule['conds'], thresholds):
        xv = df[c['var']].values.astype(float)
        if   c['op'] == '>=': cm = xv >= t
        elif c['op'] == '>':  cm = xv > t
        elif c['op'] == '<=': cm = xv <= t
        else:                 cm = xv < t
        m &= np.where(np.isnan(xv), False, cm)
    return m

is_err_all = (df['over_threshold'] == 1).values
ed_all = np.where(is_err_all, df['total_error_amount'].fillna(0).values, 0)
opt_masks  = [rule_mask_full(r, rule_best_thr[ri]) if rule_best_thr[ri] is not None else None
              for ri, r in enumerate(RULES)]
orig_masks = [rule_mask_full(r, [c['thr'] for c in r['conds']]) for r in RULES]

# ── hidden RuleFlags sheet: case x rule 0/1 matrices + live union columns ─────
# row 1 rule nums | row 2 selected (from Summary Include? col) | row 3 has-optimal
# row 4 target stratum | rows FLAG0.. one per Data case (only 1s written)
NR, NDATA = len(RULES), len(df)
RULE_ROW0 = 15                 # Summary row of the first rule's "optimal" line
NAT_ROW0  = 11                 # Summary (National) row of the first rule (delivery-rank order)
FLAG0 = 5

# Individual rules are listed by expected error $ per flagged case, best first;
# rules with no floor-compliant combination sink to the bottom.
def _exp_per_case(j):
    sc = summary_rows[2*j][2]
    return (sc['dollars'] / sc['n']) if (sc and sc['n']) else None
rule_order = sorted(range(NR), key=lambda j: (0 if _exp_per_case(j) is not None else 1,
                                              -(_exp_per_case(j) or 0), RULES[j]['num']))
rule_row = {j: RULE_ROW0 + 2*pos for pos, j in enumerate(rule_order)}   # rule idx -> Summary row
ws_f = wb.create_sheet('RuleFlags')
ws_f.sheet_state = 'hidden'
fcol_opt  = lambda j: get_column_letter(2 + j)            # optimal block B..
fcol_orig = lambda j: get_column_letter(3 + NR + j)       # original block (1 col gap)
UCOL, VCOL = get_column_letter(4 + 2*NR), get_column_letter(5 + 2*NR)
NATU  = get_column_letter(7 + 2*NR)                       # union under the National selection
natsel = lambda j: get_column_letter(9 + 2*NR + j)        # National selection vector (row 2)
set_cell(ws_f,1,1,'rule num'); set_cell(ws_f,2,1,'selected')
set_cell(ws_f,3,1,'has optimal'); set_cell(ws_f,4,1,'stratum')
for j, rule in enumerate(RULES):
    set_cell(ws_f,1,2+j, rule['num'])
    set_cell(ws_f,2,2+j, formula=f'=IF(Summary!$M${rule_row[j]}=TRUE,1,0)')
    if rule_best_thr[j] is not None:
        set_cell(ws_f,3,2+j, 1)
    set_cell(ws_f,4,2+j, rule['hh'])
    set_cell(ws_f,1,3+NR+j, rule['num'])
    set_cell(ws_f,2,9+2*NR+j,
             formula=f"=IF('Summary (National)'!$L${NAT_ROW0+j}=TRUE,1,0)")
for j in range(NR):
    if opt_masks[j] is not None:
        for i in np.flatnonzero(opt_masks[j]):
            ws_f.cell(row=FLAG0+int(i), column=2+j).value = 1
    for i in np.flatnonzero(orig_masks[j]):
        ws_f.cell(row=FLAG0+int(i), column=3+NR+j).value = 1
SELRNG = f'$B$2:${fcol_opt(NR-1)}$2'
HASRNG = f'$B$3:${fcol_opt(NR-1)}$3'
HHRNG  = f'$B$4:${fcol_opt(NR-1)}$4'
NATSEL = f'${natsel(0)}$2:${natsel(NR-1)}$2'
for i in range(NDATA):
    r = FLAG0 + i
    ws_f.cell(row=r, column=4+2*NR).value = (
        f'=IF(SUMPRODUCT({SELRNG},$B{r}:${fcol_opt(NR-1)}{r})>0,1,0)')
    ws_f.cell(row=r, column=5+2*NR).value = (
        f'=IF(SUMPRODUCT({SELRNG},${fcol_orig(0)}{r}:${fcol_orig(NR-1)}{r})>0,1,0)')
    ws_f.cell(row=r, column=7+2*NR).value = (
        f'=IF(SUMPRODUCT({NATSEL},${fcol_orig(0)}{r}:${fcol_orig(NR-1)}{r})>0,1,0)')

ws_s = wb.create_sheet('Summary', 1)
ws_s.sheet_view.showGridLines = False
for col_letter, width in {'A':9,'B':9,'C':10,'D':9,'E':11,'F':10,'G':11,
                          'H':10,'I':9,'J':13,'K':11,'L':13,'M':9,'N':115}.items():
    ws_s.column_dimensions[col_letter].width = width
merge(ws_s,1,1,1,12, value=f'Summary — Optimal Rule Parameters for {STATE_NAME}',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=16,color='FFFFFF'), align=center)
ws_s.row_dimensions[1].height = 32

merge(ws_s,2,1,2,12,
      value=f'Reading guide: "optimal" = tuned thresholds that catch the most error dollars, subject to '
            f'precision ≥ {PREC_FLOOR:.0%}, ≥ {MIN_FLAGGED} cases flagged, and ≥ {RECALL_FLOOR:.0%} of '
            f'error $ caught (matching the R optimizer), searched on the rule\'s own stratum.',
      fill=GRAY, font=Font(name=FONT,size=9,color='808080'), align=left)
merge(ws_s,3,1,3,12,
      value='"original" = the delivery-list thresholds on the same data, for comparison; dashes = the '
            'rule meets no floor-compliant combination. Check/uncheck Include? to add/remove a rule — '
            'the combined rows recompute live. Individual rules are ordered by expected error $ per '
            'flagged case.',
      fill=GRAY, font=Font(name=FONT,size=9,color='808080'), align=left)

set_cell(ws_s,2,13,False, fill=YELLOW, align=center, border=thin(), font=Font(name=FONT))
set_cell(ws_s,2,14,'← check to show "original" rows (hidden by default)',
         font=Font(name=FONT,size=9,color='808080'), align=left)

for col, txt in enumerate(['Rule','HH size','Version','Rules count','Precision','Recall',
                           '$ Recall','Flagged','Errors','Error $ caught','Workload %',
                           'Expected error $ by case','Include?','Conditions'], 1):
    set_cell(ws_s,4,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())

cases_by = {lbl: int((df['hh_group'] == lbl).sum()) for lbl in STRATA}

# ── overall block: union of all rules, overall + per stratum ──────────────────
row = 5
merge(ws_s,row,1,row,12, value='All rules combined (a case is flagged if ANY '
      'Include?-checked rule flags it) — recomputes as you tick/untick rules',
      fill=BLUE_LIGHT, font=bold_font(10), align=left)
row += 1
U_RNG = f'RuleFlags!${UCOL}${FLAG0}:${UCOL}${FLAG0+NDATA-1}'
V_RNG = f'RuleFlags!${VCOL}${FLAG0}:${VCOL}${FLAG0+NDATA-1}'
D_HH  = f'Data!${dc("hh_group")}$2:${dc("hh_group")}${1+NDATA}'
D_OV  = f'Data!${dc("over_threshold")}$2:${dc("over_threshold")}${1+NDATA}'
D_AM  = f'Data!${dc("total_error_amount")}$2:${dc("total_error_amount")}${1+NDATA}'
scopes = [('Overall', 'all', np.ones(len(df), bool))] + \
         [(f'HH {lbl}', lbl, (df['hh_group'] == lbl).values) for lbl in STRATA]
for scope_name, scope_hh, sel in scopes:
    tot_err = max(int((is_err_all & sel).sum()), 1)
    tot_ed  = round(float(ed_all[sel].sum()), 2) or 1
    tot_n   = max(int(sel.sum()), 1)
    sterm = '' if scope_hh == 'all' else f'*({D_HH}="{scope_hh}")'
    for kind, flags in [('optimal', U_RNG), ('original', V_RNG)]:
        is_opt = (kind == 'optimal')
        fnt = Font(name=FONT) if is_opt else Font(name=FONT, color='808080')
        fill = GREEN if is_opt else WHITE
        if scope_hh == 'all':
            f_cnt = (f'=SUMPRODUCT(RuleFlags!{SELRNG},RuleFlags!{HASRNG})' if is_opt
                     else f'=SUM(RuleFlags!{SELRNG})')
        else:
            hterm = f'*(RuleFlags!{HHRNG}="{scope_hh}")'
            f_cnt = (f'=SUMPRODUCT((RuleFlags!{SELRNG})*(RuleFlags!{HASRNG}){hterm})' if is_opt
                     else f'=SUMPRODUCT((RuleFlags!{SELRNG}){hterm})')
        set_cell(ws_s,row,1,'All rules', font=fnt, align=center, border=thin(), fill=fill)
        set_cell(ws_s,row,2,scope_hh, font=fnt, align=center, border=thin(), fill=fill)
        set_cell(ws_s,row,3,kind, font=fnt, align=center, border=thin(), fill=fill)
        set_cell(ws_s,row,4,formula=f_cnt, font=fnt, align=center, border=thin(), fill=fill,
                 number_format='0')
        for col, f, fmt in [
            (5, f'=IF($H{row}=0,0,$I{row}/$H{row})', '0.0%'),
            (6, f'=$I{row}/{tot_err}',               '0.0%'),
            (7, f'=$J{row}/{tot_ed}',                '0.0%'),
            (8, f'=SUMPRODUCT(({flags}){sterm})',                        '#,##0'),
            (9, f'=SUMPRODUCT(({flags}){sterm}*({D_OV}))',               '#,##0'),
            (10,f'=SUMPRODUCT(({flags}){sterm}*({D_OV})*({D_AM}))',      '$#,##0'),
            (11,f'=$H{row}/{tot_n}',                 '0.0%'),
            (12,f'=IFERROR(IF($H{row}=0,"",$J{row}/$H{row}),"")', '$#,##0'),
        ]:
            set_cell(ws_s,row,col,formula=f, font=fnt, align=center,
                     border=thin(), number_format=fmt, fill=fill)
        row += 1

merge(ws_s,row,1,row,12, value='Individual rules',
      fill=BLUE_LIGHT, font=bold_font(10), align=left)
row += 1
assert row == RULE_ROW0, f'first rule row {row} != RULE_ROW0 {RULE_ROW0}'
for j in rule_order:
    for off, (rule, kind, sc, cond) in enumerate(summary_rows[2*j:2*j+2]):
        is_opt = (kind == 'optimal')
        fnt = Font(name=FONT) if is_opt else Font(name=FONT, color='808080')
        fill = GREEN if is_opt else WHITE
        if is_opt:
            assert row == rule_row[j], f'rule {rule["num"]}: row {row} != {rule_row[j]}'
            set_cell(ws_s,row,13,rule_best_thr[j] is not None,
                     fill=YELLOW, align=center, border=thin(), font=Font(name=FONT))
        set_cell(ws_s,row,1,f'Rule {rule["num"]}', font=fnt, align=center, border=thin(), fill=fill)
        set_cell(ws_s,row,2,rule['hh'], font=fnt, align=center, border=thin(), fill=fill)
        set_cell(ws_s,row,3,kind, font=fnt, align=center, border=thin(), fill=fill)
        set_cell(ws_s,row,4,None, font=fnt, align=center, border=thin(), fill=fill)
        if sc:
            for col, key, fmt in [(5,'prec','0.0%'),(6,'rec','0.0%'),(7,'drec','0.0%'),
                                  (8,'n','#,##0'),(9,'tp','#,##0'),(10,'dollars','$#,##0')]:
                set_cell(ws_s,row,col,round(float(sc[key]),4), font=fnt, align=center,
                         border=thin(), number_format=fmt, fill=fill)
            set_cell(ws_s,row,11,round(sc['n']/cases_by[rule['hh']],4), font=fnt, align=center,
                     border=thin(), number_format='0.0%', fill=fill)
            set_cell(ws_s,row,12,(round(sc['dollars']/sc['n'], 2) if sc['n'] else '—'),
                     font=fnt, align=center, border=thin(), number_format='$#,##0', fill=fill)
        else:
            for col in range(5,13):
                set_cell(ws_s,row,col,'—', font=fnt, align=center, border=thin(), fill=fill)
        set_cell(ws_s,row,14,cond, align=left,
                 font=Font(name=FONT,size=10) if is_opt else Font(name=FONT,size=10,color='808080'))
        row += 1

LAST_ROW = row - 1
for rng_ in (f'A5:L{LAST_ROW}', f'N5:N{LAST_ROW}'):
    ws_s.conditional_formatting.add(
        rng_,
        FormulaRule(formula=['AND($C5="original",$M$2<>TRUE)'],
                    font=Font(name=FONT, color='FFFFFF'), stopIfTrue=True))
ws_s.freeze_panes = 'A5'
CHECKBOX_CELLS = {'Summary': ['M2'] + [f'M{rule_row[j]}' for j in range(NR)]}

# ══════════════════════════════════════════════════════════════════════════════
# 7. SUMMARY (NATIONAL) — delivery thresholds as-is, no grid search
# ══════════════════════════════════════════════════════════════════════════════
ws_n = wb.create_sheet('Summary (National)', 2)
ws_n.sheet_view.showGridLines = False
for col_letter, width in {'A':9,'B':9,'C':11,'D':11,'E':11,'F':11,'G':10,'H':10,
                          'I':14,'J':12,'K':21,'L':9,'M':115}.items():
    ws_n.column_dimensions[col_letter].width = width
merge(ws_n,1,1,1,13, value=f'Summary (National Thresholds) — {STATE_NAME} FY{FY_LABEL}',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=16,color='FFFFFF'), align=center)
ws_n.row_dimensions[1].height = 32
merge(ws_n,2,1,2,13,
      value='This tab reports the delivery rules exactly as mined on national QC data: thresholds are '
            f'used AS-IS, with no state-level grid search or tuning. Every metric below is computed on '
            f'the {STATE_NAME} QC cases in the Data tab.',
      fill=GRAY, font=Font(name=FONT,size=9,color='808080'), align=left)
merge(ws_n,3,1,3,13,
      value='Recall, $ Recall and Workload % are measured within each rule\'s own household-size '
            'stratum. Expected error $ by case = error dollars caught / cases flagged. Rules are listed '
            'in delivery-list rank order, exactly as delivered. Tick or untick Include? to add or '
            'remove a rule from the combined rows above.',
      fill=GRAY, font=Font(name=FONT,size=9,color='808080'), align=left)
for col, txt in enumerate(['Rule','HH size','Rules count','Precision','Recall','$ Recall',
                           'Flagged','Errors','Error $ caught','Workload %',
                           'Expected error $ by case','Include?',
                           'Conditions (national thresholds, as delivered)'], 1):
    set_cell(ws_n,4,col,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())

merge(ws_n,5,1,5,13,
      value='All rules combined (a case is flagged if ANY Include?-checked rule flags it, '
            'at national thresholds)',
      fill=BLUE_LIGHT, font=bold_font(10), align=left)
NAT_RNG = f'RuleFlags!${NATU}${FLAG0}:${NATU}${FLAG0+NDATA-1}'
for i, (scope_hh, sel) in enumerate([('all', np.ones(len(df), bool))] +
                                    [(lbl, (df['hh_group'] == lbl).values) for lbl in STRATA]):
    r = 6 + i
    tot_err = max(int((is_err_all & sel).sum()), 1)
    tot_ed  = round(float(ed_all[sel].sum()), 2) or 1
    tot_n   = max(int(sel.sum()), 1)
    sterm = '' if scope_hh == 'all' else f'*({D_HH}="{scope_hh}")'
    f_cnt = (f'=SUM(RuleFlags!{NATSEL})' if scope_hh == 'all'
             else f'=SUMPRODUCT((RuleFlags!{NATSEL})*(RuleFlags!{HHRNG}="{scope_hh}"))')
    set_cell(ws_n,r,1,'All rules', align=center, border=thin(), fill=GREEN)
    set_cell(ws_n,r,2,scope_hh, align=center, border=thin(), fill=GREEN)
    set_cell(ws_n,r,3,formula=f_cnt, align=center, border=thin(), fill=GREEN, number_format='0')
    for col, f, fmt in [
        (4, f'=IF($G{r}=0,0,$H{r}/$G{r})',                            '0.0%'),
        (5, f'=$H{r}/{tot_err}',                                      '0.0%'),
        (6, f'=$I{r}/{tot_ed}',                                       '0.0%'),
        (7, f'=SUMPRODUCT(({NAT_RNG}){sterm})',                       '#,##0'),
        (8, f'=SUMPRODUCT(({NAT_RNG}){sterm}*({D_OV}))',              '#,##0'),
        (9, f'=SUMPRODUCT(({NAT_RNG}){sterm}*({D_OV})*({D_AM}))',     '$#,##0'),
        (10,f'=$G{r}/{tot_n}',                                        '0.0%'),
        (11,f'=IFERROR(IF($G{r}=0,"",$I{r}/$G{r}),"")',               '$#,##0'),
    ]:
        set_cell(ws_n,r,col,formula=f, align=center, border=thin(),
                 number_format=fmt, fill=GREEN)
    set_cell(ws_n,r,12,None, align=center, border=thin(), fill=GREEN)
    set_cell(ws_n,r,13,'union of the checked rules at their national thresholds',
             align=left, font=Font(name=FONT,size=10))

merge(ws_n,10,1,10,13, value='Individual rules — national thresholds, no tuning',
      fill=BLUE_LIGHT, font=bold_font(10), align=left)
for j, rule in enumerate(RULES):                      # RULES is already in delivery-rank order
    r = NAT_ROW0 + j
    sc, cond = summary_rows[2*j+1][2], summary_rows[2*j+1][3]
    set_cell(ws_n,r,1,f'Rule {rule["num"]}', align=center, border=thin(), fill=GREEN)
    set_cell(ws_n,r,2,rule['hh'], align=center, border=thin(), fill=GREEN)
    set_cell(ws_n,r,3,None, align=center, border=thin(), fill=GREEN)
    for col, key, fmt in [(4,'prec','0.0%'),(5,'rec','0.0%'),(6,'drec','0.0%'),
                          (7,'n','#,##0'),(8,'tp','#,##0'),(9,'dollars','$#,##0')]:
        set_cell(ws_n,r,col,round(float(sc[key]),4), align=center, border=thin(),
                 number_format=fmt, fill=GREEN)
    set_cell(ws_n,r,10,round(sc['n']/cases_by[rule['hh']],4), align=center, border=thin(),
             number_format='0.0%', fill=GREEN)
    set_cell(ws_n,r,11,(round(sc['dollars']/sc['n'], 2) if sc['n'] else '—'),
             align=center, border=thin(), number_format='$#,##0', fill=GREEN)
    set_cell(ws_n,r,12,True, fill=YELLOW, align=center, border=thin(), font=Font(name=FONT))
    set_cell(ws_n,r,13,cond, align=left, font=Font(name=FONT,size=10))
ws_n.freeze_panes = 'A5'
CHECKBOX_CELLS['Summary (National)'] = [f'L{NAT_ROW0+j}' for j in range(NR)]

# ══════════════════════════════════════════════════════════════════════════════
# 8. ERROR CASES — live list of rules catching errors + the cases they catch
# ══════════════════════════════════════════════════════════════════════════════
ws_e = wb.create_sheet('Error Cases', 4)
ws_e.sheet_view.showGridLines = False
NCOL   = len(DCOLS)
G0     = 3                                   # first grid column (C)
GLAST  = get_column_letter(G0 + NCOL - 1)
DROWC  = get_column_letter(G0 + NCOL)        # data_row column
LISTN  = max(NR, 64)                         # rule-list slots: one per rule, so a
                                             # longer delivery list is never
                                             # silently truncated on this panel
GRIDN  = 60                                  # result rows
HELP_Z = get_column_letter(G0 + NCOL + 2)    # per-case hit flag
HELP_A = get_column_letter(G0 + NCOL + 3)    # cumulative count
HELP_Y = get_column_letter(G0 + NCOL + 1)    # rule index / dashboard row
SC_B   = get_column_letter(G0 + NCOL + 4)    # rule score
SC_C   = get_column_letter(G0 + NCOL + 5)    # rule errors
SC_D   = get_column_letter(G0 + NCOL + 6)    # compacted sorted rule list
PRESETS = f"'Grid Search'!$AA$2:$AA${1+NR}"
for cl, w in {'A':20,'B':9}.items():
    ws_e.column_dimensions[cl].width = w
for i in range(NCOL):
    ws_e.column_dimensions[get_column_letter(G0+i)].width = 13 if i > 2 else 16
ws_e.column_dimensions[DROWC].width = 9
for cl in (HELP_Y, HELP_Z, HELP_A, SC_B, SC_C, SC_D):
    ws_e.column_dimensions[cl].hidden = True

set_cell(ws_e,1,1,formula=f'=${SC_B}$1&" rules with errors"',
         fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=14,color='FFFFFF'), align=center)
ws_e.merge_cells('A1:B1'); ws_e.row_dimensions[1].height = 32
merge(ws_e,2,1,2,2, value='sorted by errors caught — updates live',
      fill=GRAY, font=Font(name=FONT,size=9,color='808080'), align=left)
set_cell(ws_e,3,1,'Rule', font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,3,2,'errors', font=bold_font(10), fill=GRAY, align=center, border=thin())
for k in range(1, LISTN+1):
    r = 3 + k
    set_cell(ws_e,r,1,formula=f'=IF({k}>${SC_B}$1,"",INDEX(${SC_D}$2:${SC_D}${1+NR},{k}))',
             align=left, border=thin())
    set_cell(ws_e,r,2,formula=f'=IF({k}>${SC_B}$1,"",INT(LARGE(${SC_B}$2:${SC_B}${1+NR},{k})/1000))',
             align=center, border=thin())

merge(ws_e,1,G0,1,G0+11,
      value='Error Cases — true errors caught by the selected rule (live Dashboard thresholds)',
      fill=BLUE_DARK, font=Font(name=FONT,bold=True,size=16,color='FFFFFF'), align=center)
merge(ws_e,2,G0,2,G0+NCOL-1,
      value='Pick a rule (the dropdown lists only rules currently catching errors, sorted by errors '
            'caught). The table lists every case the rule flags at its CURRENT Dashboard thresholds '
            'that is a true payment error (over_threshold = 1).',
      fill=GRAY, font=Font(name=FONT,size=9,color='808080'), align=left)
set_cell(ws_e,3,G0,'Rule', font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,3,G0+1,f'Rule {RULES[0]["num"]} (HH {RULES[0]["hh"]})',
         fill=YELLOW, align=center, border=thin(), font=bold_font(), number_format='@')
ws_e.merge_cells(start_row=3, start_column=G0+1, end_row=3, end_column=G0+3)
set_cell(ws_e,3,G0+4,'← dropdown lists only the rules currently catching errors',
         font=Font(name=FONT,size=9,color='808080'), align=left)
YREF = f'${HELP_Y}$1'; Y2REF = f'${HELP_Y}$2'
set_cell(ws_e,1,G0+NCOL+1,
         formula=f'=MATCH(${get_column_letter(G0+1)}$3,{PRESETS},0)-1')
set_cell(ws_e,2,G0+NCOL+1, formula=f'={BASE_ROW+2}+{YREF}*{BLOCK_HEIGHT}')
set_cell(ws_e,4,G0,'HH stratum', font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,4,G0+1,formula=f'=INDEX(RuleFlags!{HHRNG},1,{YREF}+1)',
         fill=BLUE_LIGHT, align=center, border=thin())
set_cell(ws_e,4,G0+3,'n_flagged', font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,4,G0+4,formula=f'=INDEX(Dashboard!$G:$G,{Y2REF})',
         fill=BLUE_LIGHT, align=center, border=thin())
set_cell(ws_e,4,G0+6,'errors caught', font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,4,G0+7,formula=f'=INDEX(Dashboard!$G:$G,{Y2REF}+1)',
         fill=BLUE_LIGHT, align=center, border=thin())
for off, txt in enumerate(['Variable','Op','Dashboard threshold','Data col #']):
    set_cell(ws_e,5,G0+off,txt, font=bold_font(10), fill=GRAY, align=center, border=thin())
CV = get_column_letter(G0); OPV = get_column_letter(G0+1)
THV = get_column_letter(G0+2); IXV = get_column_letter(G0+3)
for i in range(NSLOTS):
    r = 6 + i
    set_cell(ws_e,r,G0,  formula=f'=IFERROR(IF(INDEX(Dashboard!$A:$A,{Y2REF}+{i})=0,"",'
                                 f'INDEX(Dashboard!$A:$A,{Y2REF}+{i})),"")',
             fill=BLUE_LIGHT, align=center, border=thin())
    set_cell(ws_e,r,G0+1,formula=f'=IFERROR(IF(INDEX(Dashboard!$B:$B,{Y2REF}+{i})=0,"",'
                                 f'INDEX(Dashboard!$B:$B,{Y2REF}+{i})),"")',
             fill=BLUE_LIGHT, align=center, border=thin())
    set_cell(ws_e,r,G0+2,formula=f'=IF(${CV}{r}="","",INDEX(Dashboard!$C:$C,{Y2REF}+{i}))',
             fill=BLUE_LIGHT, align=center, border=thin())
    set_cell(ws_e,r,G0+3,formula=f'=IF(${CV}{r}="","",IFERROR(MATCH(${CV}{r},'
                                 f'Data!$A$1:${LASTCOL}$1,0),""))',
             fill=BLUE_LIGHT, align=center, border=thin())
for i, h in enumerate(DCOLS):
    set_cell(ws_e,10,G0+i,h, font=bold_font(10), fill=GRAY, align=center, border=thin())
set_cell(ws_e,10,G0+NCOL,'data_row', font=bold_font(10), fill=GRAY, align=center, border=thin())
CUM = f'${HELP_A}$12:${HELP_A}${11+NDATA}'
for r in range(11, 11+GRIDN):
    for i in range(NCOL):
        ws_e.cell(row=r, column=G0+i).value = (
            f'=IFERROR(INDEX(Data!$A$2:${LASTCOL}${1+NDATA},MATCH(ROW()-10,{CUM},0),'
            f'COLUMN()-{G0-1}),"")')
    ws_e.cell(row=r, column=G0+NCOL).value = f'=IFERROR(MATCH(ROW()-10,{CUM},0)+1,"")'
for i in range(NDATA):
    r, dr_ = 12 + i, 2 + i
    slots = '*'.join(
        f'IF(${CV}${6+k}="",1,COUNTIF(INDEX(Data!$A{dr_}:${LASTCOL}{dr_},1,${IXV}${6+k}),'
        f'${OPV}${6+k}&${THV}${6+k}))' for k in range(NSLOTS))
    ws_e.cell(row=r, column=G0+NCOL+2).value = (
        f'=IF(Data!${dc("hh_group")}{dr_}<>${get_column_letter(G0+1)}$4,0,'
        f'IF(Data!${dc("over_threshold")}{dr_}<>1,0,{slots}))')
    ws_e.cell(row=r, column=G0+NCOL+3).value = (
        f'=${HELP_Z}$12' if i == 0 else f'=${HELP_A}{r-1}+${HELP_Z}{r}')
ws_e.cell(row=1, column=G0+NCOL+4).value = f'=COUNTIF(${SC_C}$2:${SC_C}${1+NR},">0")'
for j in range(NR):
    r = 2 + j
    ws_e.cell(row=r, column=G0+NCOL+4).value = (
        f'=INDEX(Dashboard!$G:$G,{BASE_ROW+3}+{j}*{BLOCK_HEIGHT})*1000+({NR}-{j})')
    ws_e.cell(row=r, column=G0+NCOL+5).value = (
        f'=INDEX(Dashboard!$G:$G,{BASE_ROW+3}+{j}*{BLOCK_HEIGHT})')
    ws_e.cell(row=r, column=G0+NCOL+6).value = (
        f'=IF(ROW()-1>${SC_B}$1,"",INDEX({PRESETS},'
        f'{1+NR}-MOD(LARGE(${SC_B}$2:${SC_B}${1+NR},ROW()-1),1000)))')
wb.defined_names.add(DefinedName('RuleListLive',
    attr_text=f"OFFSET('Error Cases'!${SC_D}$2,0,0,MAX('Error Cases'!${SC_B}$1,1),1)"))
dv_live = DataValidation(type='list', formula1='RuleListLive', allow_blank=False)
ws_e.add_data_validation(dv_live)
dv_live.add(f'{get_column_letter(G0+1)}3')
ws_e.freeze_panes = f'{get_column_letter(G0)}11'

# ── final touches ─────────────────────────────────────────────────────────────
ws.sheet_state   = 'hidden'      # Dashboard  (engine)
ws_g.sheet_state = 'hidden'      # Grid Search(engine)

# ── Save ──────────────────────────────────────────────────────────────────────
os.makedirs(STATE_DIR, exist_ok=True)
wb.save(OUT)
import json
os.makedirs(BUILD_DIR, exist_ok=True)
json.dump(CHECKBOX_CELLS, open(os.path.join(BUILD_DIR, 'checkbox_cells.json'), 'w'))
json.dump({'out': OUT, 'state': STATE_NAME, 'abbr': STATE_ABBR},
          open(os.path.join(BUILD_DIR, 'target.json'), 'w'))
print(f'Saved: {OUT}')
print(f'Rules: {len(RULES)} | Data rows: {len(df)}')
print(f'Checkbox cells recorded for post-processing: '
      f'{sum(len(v) for v in CHECKBOX_CELLS.values())}')
