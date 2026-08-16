"""
Reconstruction stage: the workbook accepts RAW case fields and computes every
model feature with live formulas, so a state that can run neither Python nor R
can still use it. The input columns carry generic reported-value concept names
in features.R's state_col_map style (HOUSEHOLD_SIZE, EARNED_INCOME, RENT, ...)
that a state maps its own system's fields onto; the Data Dictionary tab
carries the crosswalk to the SNAP QC technical documentation's variables.

Data tab layout (one Excel table, columns bound by name):
  [ feature columns ]   the model features, in the SAME columns the built
                        workbook put them (A..), now formulas — positional
                        references from the other sheets stay valid
  [ hit columns ]       per-case rule tests, carried over from the LIVE build
  [ helper columns ]    the benefit-recomputation chain, hidden
  [ input contract ]    what a state pastes (RAW_COLS below): unit counts and
                        indicators, income totals, deductions, shelter costs,
                        and the QC review outcome (ERROR_FLAG, ERROR_AMOUNT)

Federal parameter tables (standard deduction, max allotment, shelter cap,
minimum allotment, by fiscal year x household size) live on a hidden
FederalTables sheet; formulas look them up by year and size.

The demo input block carries the research frame's RECONSTRUCTED
(pre-QC-review) values — the same scale the rules were mined on — so the
workbook's figures match the research pipeline's. A state pasting its own
data supplies its ordinary as-reported fields (internal data carries no QC
corrections, so nothing needs reconstructing there). The built-in validation
prints the per-feature match rate against the frame on every build.

Usage:  python make_recon.py <LIVE_workbook.xlsx> -o <out.xlsx> [--state WA]
Then:   python postprocess_workbook.py <out.xlsx> <checkbox_cells.json>
"""
import argparse
import os
import shutil

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter as CL
from openpyxl.worksheet.table import Table, TableStyleInfo

import states as STATE_REGISTRY

TABLE = 'CaseData'
PKG = os.path.dirname(os.path.abspath(__file__))


def find_repo(start):
    cands = [os.environ['SNAP_REPO']] if os.environ.get('SNAP_REPO') else []
    d = start
    for _ in range(6):
        cands += [d, os.path.join(d, 'snap_qc')]
        d = os.path.dirname(d)
    for c in cands:
        if (os.path.isfile(os.path.join(c, 'reg_model_data.rds'))
                and os.path.isfile(os.path.join(c, 'rule_mining_helpers.R'))):
            return os.path.abspath(c)
    raise SystemExit('cannot find the snap_qc checkout; set SNAP_REPO')


REPO = find_repo(PKG)

# ── the raw input contract, one row per reviewed case ─────────────────────────
# Column names follow features.R's state_col_map style: generic reported-value
# concepts a state maps its own system's fields onto. The Data Dictionary tab
# carries the crosswalk to the QC technical manual's variables. NB the QC
# outcome is named ERROR_FLAG, not OVER_THRESHOLD: Excel table column names
# are case-insensitively unique and the feature column over_threshold already
# claims that name.
RAW_COLS = [
    'CASE_ID', 'REVIEW_FISCAL_YEAR',
    'HOUSEHOLD_SIZE',
    'NUM_CHILDREN', 'NUM_ELDERLY', 'NUM_DISABLED', 'NUM_ABAWD',
    'MARRIED_FLAG', 'EXPEDITED', 'CATEGORICALLY_ELIGIBLE', 'HOMELESS_FLAG',
    'MONTHS_SINCE_CERT',
    'EARNED_INCOME', 'UNEARNED_INCOME',
    'MEDICAL_DEDUCTION', 'DEPENDENT_CARE_DEDUCTION', 'CHILD_SUPPORT_DEDUCTION',
    'HOMELESS_DEDUCTION',
    'RENT', 'UTILITY_COSTS',
    'NUM_AMOUNTS_DIVISIBLE_BY_100',   # state-precomputed; definition in the dictionary
    'ERROR_FLAG', 'ERROR_AMOUNT',
]


def raw_frame(cfg, frame_csv):
    """The input block for exactly the frame's rows, in the frame's row order.

    The demo carries the research frame's RECONSTRUCTED (pre-QC-review)
    values — the same scale the rules were mined on — so the workbook's
    figures match the research pipeline's (decision 2026-08-16). A state
    pasting its own data supplies its ordinary as-reported fields; internal
    data carries no QC corrections, so no reconstruction is needed there.
    Everything comes from the frame export; the .sav files are not read.
    """
    frame = pd.read_csv(frame_csv, dtype={'hhldno': str, 'stratum': str})
    need = ['rawearn', 'rawunearn', 'rawdepded', 'rawcsded', 'rawrent',
            'rawhomeless_ded', 'fsnkid', 'fsnelder', 'fsndis',
            'count_abawd', 'cat_elig']
    missing = [c for c in need if c not in frame.columns]
    assert not missing, (f'frame export lacks reconstructed input fields '
                         f'{missing}; re-export it (make_state.py --refresh)')

    g = lambda c: pd.to_numeric(frame[c], errors='coerce')
    out = pd.DataFrame({
        'CASE_ID': frame['hhldno'].astype(str).str.strip(),
        'REVIEW_FISCAL_YEAR': g('fiscal_year'),
        'HOUSEHOLD_SIZE': g('hh_size_raw'),
        'NUM_CHILDREN': g('fsnkid'),
        'NUM_ELDERLY': g('fsnelder'),
        'NUM_DISABLED': g('fsndis'),
        'NUM_ABAWD': g('count_abawd'),
        'MARRIED_FLAG': g('married').fillna(0).astype(int),
        'EXPEDITED': g('expedited_i').fillna(0).astype(int),
        'CATEGORICALLY_ELIGIBLE': (g('cat_elig') >= 1).fillna(False).astype(int),
        'HOMELESS_FLAG': g('homeless').fillna(0).astype(int),
        'MONTHS_SINCE_CERT': g('months_since_cert_n'),
        'EARNED_INCOME': g('rawearn'),
        'UNEARNED_INCOME': g('rawunearn'),
        'MEDICAL_DEDUCTION': g('medical_deductions'),
        'DEPENDENT_CARE_DEDUCTION': g('rawdepded'),
        'CHILD_SUPPORT_DEDUCTION': g('rawcsded'),
        'HOMELESS_DEDUCTION': g('rawhomeless_ded'),
        'RENT': g('rawrent'),
        'UTILITY_COSTS': g('utilities'),
        # per its mined definition this counts the QC file's component fields,
        # which the totals-based contract does not collect; the demo carries
        # the frame's own value and a state supplies its own (see dictionary)
        'NUM_AMOUNTS_DIVISIBLE_BY_100': g('count_divisible_by_100'),
        'ERROR_FLAG': frame['is_error'].astype(int).values,
        'ERROR_AMOUNT': g('total_error_amount'),
    })
    assert list(out.columns) == RAW_COLS
    # with reconstructed inputs there is no restoration gap; keep the
    # element-free split trivially all-True for the validation report
    return out, np.ones(len(frame), bool), frame


def federal_tables(wb, holdout_year):
    """Hidden sheet with the year x size parameter tables + settings."""
    ad = os.path.join(REPO, 'additional_data')
    yd = pd.read_csv(os.path.join(ad, 'year_data.csv')).dropna(axis=1, how='all')
    yd.columns = [str(c).strip() for c in yd.columns]
    sd = pd.read_csv(os.path.join(ad, 'standard_deductions.csv')).dropna(axis=1, how='all')
    ma = pd.read_csv(os.path.join(ad, 'max_allotments.csv')).dropna(axis=1, how='all')
    for t in (sd, ma):
        t.columns = [str(c).strip() for c in t.columns]

    ws = wb.create_sheet('FederalTables')
    ws.sheet_state = 'hidden'
    ws['A1'] = ('Federal SNAP parameters by fiscal year. Append a row per table '
                'each new fiscal year; every formula updates automatically.')
    ws['A3'], ws['B3'] = 'holdout_year (split label only)', holdout_year
    ws['A5'], ws['B5'], ws['C5'] = 'year', 'max_shelter', 'min_allotment'
    years = sorted(yd['year'])
    for i, y in enumerate(years):
        r = yd[yd.year == y].iloc[0]
        ws.cell(row=6 + i, column=1, value=int(y))
        ws.cell(row=6 + i, column=2, value=float(r['max_shelter_deduction']))
        ws.cell(row=6 + i, column=3, value=float(r['min_allotment']))
    ny = len(years)

    def block(df, row0, label):
        ws.cell(row=row0, column=5, value=label)
        ws.cell(row=row0 + 1, column=5, value='year')
        for s in range(1, 21):
            ws.cell(row=row0 + 1, column=5 + s, value=s)
        yrs = sorted(df['year'])
        for i, y in enumerate(yrs):
            ws.cell(row=row0 + 2 + i, column=5, value=int(y))
            r = df[df.year == y].iloc[0]
            for s in range(1, 21):
                ws.cell(row=row0 + 2 + i, column=5 + s, value=float(r[str(s)]))
        return len(yrs), row0 + 2
    n_sd, sd0 = block(sd.reset_index(), 5, 'standard deduction')
    n_ma, ma0 = block(ma.reset_index(), 5 + n_sd + 4, 'max allotment')
    return {
        'YEARS':  f'FederalTables!$A$6:$A${5 + ny}',
        'MAXSH':  f'FederalTables!$B$6:$B${5 + ny}',
        'MINAL':  f'FederalTables!$C$6:$C${5 + ny}',
        'SDYRS':  f'FederalTables!$E${sd0}:$E${sd0 + n_sd - 1}',
        'SDBLK':  f'FederalTables!$F${sd0}:$Y${sd0 + n_sd - 1}',
        'MAYRS':  f'FederalTables!$E${ma0}:$E${ma0 + n_ma - 1}',
        'MABLK':  f'FederalTables!$F${ma0}:$Y${ma0 + n_ma - 1}',
        'HOLDOUT': 'FederalTables!$B$3',
    }


# ── Start Here: KPIs, how-to, and background ─────────────────────────────────
GITHUB = 'https://github.com/giannella/snap_qc'


def background_tab(wb, state_name):
    ws = wb.create_sheet('Start Here', 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '2F5496'
    blue = PatternFill('solid', fgColor='2F5496')
    green = PatternFill('solid', fgColor='E2EFDA')
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 58
    ws.merge_cells('A1:B1')
    c = ws['A1']; c.value = f'SNAP QC payment-error rule lists — {state_name}'
    c.fill = blue; c.font = Font(bold=True, size=16, color='FFFFFF')
    ws.row_dimensions[1].height = 30

    r = 3

    def para(txt, bold=False, height=None):
        nonlocal r
        ws.merge_cells(f'A{r}:B{r}')
        c = ws.cell(row=r, column=1, value=txt)
        c.font = Font(bold=bold, size=11 if not bold else 12)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = height or (18 if bold else 60)
        r += 2

    def link(label, url):
        nonlocal r
        c = ws.cell(row=r, column=1, value=label)
        c.hyperlink = url
        c.font = Font(color='0563C1', underline='single')
        r += 1

    # live KPI block: reads the Blended Rules union row, recomputes on paste
    kpis = [
        ('Cases in this workbook',
         f'=COUNTA({TABLE}[CASE_ID])', '#,##0'),
        ('Flagged by the blended rules',
         "='Blended Rules'!G6", '#,##0'),
        ('Precision: share of flagged cases that are true errors',
         "='Blended Rules'!D6", '0.0%'),
        ('Base error rate, all cases (compare precision against this)',
         f'=COUNTIF({TABLE}[over_threshold],1)/COUNTA({TABLE}[CASE_ID])', '0.0%'),
        ('Error $ caught by the blended rules',
         "='Blended Rules'!I6", '$#,##0'),
    ]
    for label, f, fmt in kpis:
        ws.cell(row=r, column=1, value=label).font = Font(size=11)
        c = ws.cell(row=r, column=2)
        c.value = f
        c.font = Font(bold=True, size=18)
        c.number_format = fmt
        c.fill = green
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1

    para('How to use this workbook', bold=True)
    for step in [
        '1.  Read the Blended Rules tab: the combined rows at the top are the headline; '
        'each row below is one rule, in plain language, with its results here and its '
        'national evidence.',
        '2.  If there are rules that you do not want to use, untick the box in the '
        'Include? Column (or change the value to FALSE, if you don\'t see a checkbox).',
        '3.  Paste your own cases into the Data tab\'s amber columns (definitions on the '
        'Data Dictionary tab) — every figure on this tab and the rules tabs recomputes '
        'on your data.',
    ]:
        ws.merge_cells(f'A{r}:B{r}')
        c = ws.cell(row=r, column=1, value=step)
        c.font = Font(size=11)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 28
        r += 1
    r += 1

    para('What this is', bold=True)
    para('Rule lists for finding SNAP cases at higher risk of a payment error, evaluated '
         f'against {state_name}\'s cases in the public USDA SNAP Quality Control (QC) files '
         'for FY2022-2024. Each rule is a short, readable condition on case fields (for '
         'example: household size, income per person, deductions, benefit relative to the '
         'maximum). The rules tabs are POTENTIAL lists to evaluate: the blended list '
         '(national + this state\'s own mined rules) and the national-only list.')
    para('Why to use internal data, not just the public file', bold=True)
    para('Everything here is computed on the public QC file, which is a small audit sample '
         '(roughly one to a few thousand reviewed cases per state per year) and misses '
         'entire classes of cases: across states, the public files contain only 43-81% of '
         'each state\'s QC error cases. At that size, most individual rules flag a handful '
         'of cases, so per-rule figures carry wide uncertainty. Pasting the state\'s own '
         'internal case data into the Data tab -- the same fields, full caseload -- makes '
         'every figure recompute on far more cases, so rule performance becomes estimable.')
    para('How the rule lists were made', bold=True)
    para('Candidate rules are read off tree ensembles (xgboost and ranger) fitted to the '
         'national QC frame within household-size strata (1 / 2-3 / 4+), then filtered on a '
         'conservative lower confidence bound of training precision, admitted by a '
         'false-discovery-rate test against the stratum base rate, ranked, and filled to a '
         '10% review budget. The blended list additionally merges rules mined on this '
         'state\'s own QC rows into the national pool on a common ranking scale. Full '
         'methods and code:')
    link('The pipeline that builds these workbooks (GitHub)',
         f'{GITHUB}/tree/main/methods/excel_rules_for_states')
    link('The delivery rule lists, one CSV per state (GitHub)',
         f'{GITHUB}/tree/main/state_delivery_lists')
    link('The finished state workbooks (GitHub)',
         f'{GITHUB}/tree/main/methods/excel_rules_for_states/state_workbooks')
    r += 1
    para('Where the numbers on the rules tabs come from', bold=True)
    para('The rules were mined on a research frame whose fields are restored to their '
         'pre-QC-review values, and this workbook\'s Data tab carries those same '
         'reconstructed values, so the figures here sit on the scale the rules were mined '
         'on. A state pasting its own data supplies its ordinary as-reported case fields; '
         'internal data carries no QC corrections, so nothing needs reconstructing. The '
         'Data Dictionary tab defines every column; the Data tab\'s amber block is what a '
         'state supplies.')
    ws.freeze_panes = 'A2'
    return ws


# ── data dictionary: one line per column on the Data tab ─────────────────────
# Input names follow features.R's state_col_map style; each description cross-
# references the SNAP QC technical documentation's variables (Origin R =
# reported on the QC Review Schedule, C = constructed during editing). A state
# supplies its own as-reported value for each concept; the public demo carries
# the research frame's RECONSTRUCTED (pre-QC-review) value, which puts the
# workbook's figures on the same scale the rules were mined on.
RAW_DESC = {
    'CASE_ID':   'case / review identifier — row identity only. QC manual: HHLDNO.',
    'REVIEW_FISCAL_YEAR': 'federal fiscal year of the review month (Oct-Sep).',
    'HOUSEHOLD_SIZE': 'SNAP unit size as reported. QC manual: RAWHSIZE is the reported '
                      '(Origin R) version, FSUSIZE the QC-corrected one; the demo carries '
                      'the reconstructed pre-QC-review size.',
    'NUM_CHILDREN': 'children in the unit. QC manual: FSNKID.',
    'NUM_ELDERLY': 'members aged 60+. QC manual: FSNELDER.',
    'NUM_DISABLED': 'disabled members. QC manual: FSNDIS.',
    'NUM_ABAWD': 'members with ABAWD status. QC manual: count of ABWDST1-18 coded 2..5.',
    'MARRIED_FLAG':   '1 if a spouse is present in the unit. QC manual: any REL1-16 = 2.',
    'EXPEDITED': '1 if the case received expedited service. QC manual: EXPEDSER in (1, 2).',
    'CATEGORICALLY_ELIGIBLE': '1 if the unit is categorically eligible. '
                              'QC manual: CAT_ELIG >= 1.',
    'HOMELESS_FLAG':  '1 if the unit is homeless. QC manual: HOMEDED present and not 1.',
    'MONTHS_SINCE_CERT': 'months since the last certification. QC manual: LASTCERT.',
    'EARNED_INCOME': 'total earned income, $/month, as reported. QC manual: FSEARN '
                     '(= FSWAGES + FSSLFEMP + FSOTHERN) is the QC-corrected total; the '
                     'demo carries the reconstructed pre-QC-review value.',
    'UNEARNED_INCOME': 'total unearned income, $/month, as reported. QC manual: FSUNEARN '
                       '(the sum of the ~30 unearned income-type fields, FY2024 '
                       'definition) is the QC-corrected total; the demo carries the '
                       'reconstructed pre-QC-review value.',
    'MEDICAL_DEDUCTION': 'medical expense deduction, $/month. QC manual: FSMEDDED; the '
                         'demo carries the reconstructed pre-QC-review value.',
    'DEPENDENT_CARE_DEDUCTION': 'dependent care deduction, $/month. QC manual: FSDEPDED.',
    'CHILD_SUPPORT_DEDUCTION': 'child support payment deduction, $/month. QC manual: FSCSDED.',
    'HOMELESS_DEDUCTION': 'homeless household shelter deduction, $/month. '
                          'QC manual: HOMELESS_DED.',
    'RENT':      'rent / mortgage, $/month. QC manual: RENT; the demo carries the '
                 'reconstructed pre-QC-review value.',
    'UTILITY_COSTS': 'utilities, $/month. QC manual: UTIL; the demo carries the '
                     'reconstructed pre-QC-review value.',
    'NUM_AMOUNTS_DIVISIBLE_BY_100': 'how many of the case\'s reported income-type and '
                              'deduction-type amounts are a positive exact multiple of '
                              '$100 (QC manual: counted over the 21 income and 7 deduction '
                              'fields). Supplied precomputed: the totals above cannot '
                              'reproduce it. Demo carries the research frame\'s value.',
    'ERROR_FLAG': 'QC review outcome: 1 = payment error over the federal threshold.',
    'ERROR_AMOUNT': 'QC review outcome: benefit amount in error, $. QC manual: AMTERR.',
}
FEAT_DESC = {
    'fiscal_year': 'REVIEW_FISCAL_YEAR, unchanged',
    'split':       'label only: earlier fiscal years vs the most recent one (no tuning uses it)',
    'hh_size_raw': 'HOUSEHOLD_SIZE, unchanged',
    'hh_group':    'household-size stratum: 1 / 2-3 / 4+ (every rule applies within one stratum)',
    'HH_size_n':   'HOUSEHOLD_SIZE as a number, used inside rules',
    'bbce_state_i': 'state-year flag: 1 when at least half the year\'s cases are '
                    'CATEGORICALLY_ELIGIBLE (the state runs Broad-Based Categorical '
                    'Eligibility)',
    'children_i':  '1 if NUM_CHILDREN > 0',
    'count_divisible_by_100': 'NUM_AMOUNTS_DIVISIBLE_BY_100, unchanged (supplied precomputed; '
                              'see its input definition)',
    'elderly_disabled_i': '1 if NUM_ELDERLY + NUM_DISABLED > 0',
    'expedited_i': 'EXPEDITED, unchanged',
    'homeless':    'HOMELESS_FLAG, unchanged',
    'married':     'MARRIED_FLAG, unchanged',
    'medical_deductions': 'MEDICAL_DEDUCTION, unchanged',
    'months_since_cert_n': 'MONTHS_SINCE_CERT, unchanged',
    'percent_abawd': 'NUM_ABAWD / HOUSEHOLD_SIZE (the research frame divided by the '
                     'household person count, CERTHHSZ, which this contract does not collect)',
    'earned_by_hh_size':   'EARNED_INCOME / HOUSEHOLD_SIZE',
    'unearned_by_hh_size': 'UNEARNED_INCOME / HOUSEHOLD_SIZE',
    'gross_by_hh_size':    '(EARNED_INCOME + UNEARNED_INCOME) / HOUSEHOLD_SIZE',
    'rawben_rel_max':      'recomputed benefit / maximum allotment for the unit size '
                           '(via the hidden benefit-recomputation chain and FederalTables)',
    'unc_rawben_rel_max':  'recomputed benefit BEFORE the minimum/maximum caps / maximum allotment',
    'shelter_expenses_by_hh_size': '(RENT + UTILITY_COSTS) / HOUSEHOLD_SIZE',
    'total_deductions_by_hh_size': '(dependent care + child support + recomputed shelter + medical '
                                   '+ earned-income deductions) / HOUSEHOLD_SIZE',
    'utilities':   'UTILITY_COSTS, unchanged',
    'over_threshold':     'ERROR_FLAG, unchanged (what the rules aim to catch)',
    'total_error_amount': 'ABS(ERROR_AMOUNT), rounded to whole dollars',
}
HELPER_NOTE = ('_c_* columns (hidden): the benefit-recomputation chain — fiscal year, '
               'gross income, earned-income deduction, standard deduction and maximum '
               'allotment looked up per year x size from the hidden FederalTables sheet, '
               'net income before and after shelter, the shelter deduction with its '
               'elderly/disabled uncapping, and the recomputed benefit.')


def data_dictionary(wb, hdr):
    """A visible tab documenting every Data column: the input fields a state
    supplies and the constructed model variables computed from them, each
    cross-referenced to the SNAP QC technical documentation's variables."""
    ws = wb.create_sheet('Data Dictionary', wb.sheetnames.index('Data') + 1)
    ws.sheet_view.showGridLines = False
    blue = PatternFill('solid', fgColor='2F5496')
    gray = PatternFill('solid', fgColor='F2F2F2')
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 120
    ws.merge_cells('A1:C1')
    c = ws['A1']; c.value = 'Data Dictionary — every column on the Data tab'
    c.fill = blue; c.font = Font(bold=True, size=14, color='FFFFFF')
    ws.row_dimensions[1].height = 28

    def header(r, txt):
        ws.merge_cells(f'A{r}:C{r}')
        c = ws.cell(row=r, column=1, value=txt)
        c.fill = blue; c.font = Font(bold=True, color='FFFFFF')
        return r + 1

    def cols(r):
        for ci, t in enumerate(('column', 'type', 'definition'), 1):
            c = ws.cell(row=r, column=ci, value=t)
            c.fill = gray; c.font = Font(bold=True, size=10)
        return r + 1

    r = header(3, 'Input fields (amber block) — what a state supplies: its own '
                  'as-reported values, mapped onto these columns. Each '
                  'definition cross-references the SNAP QC technical '
                  'documentation; this public-data copy carries the research '
                  'frame\'s reconstructed pre-QC-review values.')
    r = cols(r)
    for name in RAW_COLS:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value='input value')
        ws.cell(row=r, column=3, value=RAW_DESC.get(name, ''))
        r += 1
    r = header(r + 1, 'Constructed variables (blue block) — computed by formula from the '
                      'input fields; the rules reference these. Do not paste over them.')
    r = cols(r)
    for name in hdr:
        if name.startswith('_') or name not in FEAT_DESC:
            continue
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value='formula')
        ws.cell(row=r, column=3, value=FEAT_DESC[name])
        r += 1
    r += 1
    ws.merge_cells(f'A{r}:C{r}')
    c = ws.cell(row=r, column=1, value=HELPER_NOTE)
    c.font = Font(size=9, color='808080')
    ws.freeze_panes = 'A2'
    return ws


def T(col):
    return f'{TABLE}[[#This Row],[{col}]]'


def feature_formulas(R):
    """name -> formula for every feature the current delivery vocabulary uses,
    mirroring 1_data_munging_..._for_using_public_qc_data.R. Helpers are
    prefixed '_c_'; order matters (left to right)."""
    sz = f'MIN(MAX({T("HOUSEHOLD_SIZE")},1),20)'
    hh = f'MAX({T("HOUSEHOLD_SIZE")},1)'
    helpers = [
        ('_c_fy',      f'={T("REVIEW_FISCAL_YEAR")}'),
        ('_c_eld',     f'=IF({T("NUM_ELDERLY")}+{T("NUM_DISABLED")}>0,1,0)'),
        ('_c_gross',   f'={T("EARNED_INCOME")}+{T("UNEARNED_INCOME")}'),
        # _xlfn. prefix: post-2007 functions written straight into the XML
        # need it, or Excel renders #NAME?
        ('_c_ernded',  f'=_xlfn.FLOOR.MATH({T("EARNED_INCOME")}*0.2)'),
        ('_c_stdded',  f'=INDEX({R["SDBLK"]},MATCH({T("_c_fy")},{R["SDYRS"]},1),{sz})'),
        ('_c_benmax',  f'=INDEX({R["MABLK"]},MATCH({T("_c_fy")},{R["MAYRS"]},1),{sz})'),
        ('_c_netbs',   f'={T("_c_gross")}-({T("_c_ernded")}+{T("DEPENDENT_CARE_DEDUCTION")}'
                       f'+{T("MEDICAL_DEDUCTION")}+{T("CHILD_SUPPORT_DEDUCTION")}+{T("_c_stdded")})'),
        ('_c_maxsh',   f'=IF({T("_c_eld")}=1,1000000000,'
                       f'INDEX({R["MAXSH"]},MATCH({T("_c_fy")},{R["YEARS"]},1)))'),
        # NB: the munging script does NOT floor the shelter deduction; only the
        # net incomes and the benefit are floored (calculate_raw_benefits)
        ('_c_sltded',  f'=MIN(MAX({T("RENT")}+{T("UTILITY_COSTS")}'
                       f'-MAX({T("_c_netbs")}*0.5,0),0),{T("_c_maxsh")})'),
        ('_c_netan',   f'=_xlfn.FLOOR.MATH({T("_c_netbs")}-({T("_c_sltded")}'
                       f'+{T("HOMELESS_DEDUCTION")}))'),
        ('_c_benunc',  f'=_xlfn.FLOOR.MATH({T("_c_benmax")}-0.3*{T("_c_netan")})'),
        ('_c_benrec',  f'=MIN(MAX({T("_c_benunc")},IF({T("HOUSEHOLD_SIZE")}<3,'
                       f'INDEX({R["MINAL"]},MATCH({T("_c_fy")},{R["YEARS"]},1)),0)),'
                       f'{T("_c_benmax")})'),
    ]
    feats = {
        'fiscal_year': f'={T("REVIEW_FISCAL_YEAR")}',
        'split':       f'=IF({T("REVIEW_FISCAL_YEAR")}>={R["HOLDOUT"]},"holdout","tune")',
        'hh_size_raw': f'={T("HOUSEHOLD_SIZE")}',
        'hh_group':    f'=IF({T("HOUSEHOLD_SIZE")}>=4,"4+",'
                       f'IF({T("HOUSEHOLD_SIZE")}>=2,"2-3","1"))',
        'HH_size_n':   f'={T("HOUSEHOLD_SIZE")}',
        'bbce_state_i':
            # state-year regime flag: share of the year's cases categorically
            # eligible reaching 0.5 (munging script, 2026-08-13 decision)
            f'=IF(COUNTIFS({TABLE}[REVIEW_FISCAL_YEAR],{T("REVIEW_FISCAL_YEAR")},'
            f'{TABLE}[CATEGORICALLY_ELIGIBLE],1)'
            f'/COUNTIFS({TABLE}[REVIEW_FISCAL_YEAR],{T("REVIEW_FISCAL_YEAR")})>=0.5,1,0)',
        'children_i':  f'=IF({T("NUM_CHILDREN")}>0,1,0)',
        'count_divisible_by_100': f'={T("NUM_AMOUNTS_DIVISIBLE_BY_100")}',
        'elderly_disabled_i': f'={T("_c_eld")}',
        'expedited_i': f'={T("EXPEDITED")}',
        'homeless':    f'={T("HOMELESS_FLAG")}',
        'married':     f'={T("MARRIED_FLAG")}',
        'medical_deductions':  f'={T("MEDICAL_DEDUCTION")}',
        'months_since_cert_n': f'={T("MONTHS_SINCE_CERT")}',
        'percent_abawd': f'={T("NUM_ABAWD")}/{hh}',
        'earned_by_hh_size':   f'={T("EARNED_INCOME")}/{hh}',
        'unearned_by_hh_size': f'={T("UNEARNED_INCOME")}/{hh}',
        'gross_by_hh_size':    f'=({T("EARNED_INCOME")}+{T("UNEARNED_INCOME")})/{hh}',
        'rawben_rel_max':      f'={T("_c_benrec")}/{T("_c_benmax")}',
        'shelter_expenses_by_hh_size': f'=({T("RENT")}+{T("UTILITY_COSTS")})/{hh}',
        'total_deductions_by_hh_size':
            f'=({T("DEPENDENT_CARE_DEDUCTION")}+{T("CHILD_SUPPORT_DEDUCTION")}'
            f'+{T("_c_sltded")}+{T("MEDICAL_DEDUCTION")}+{T("_c_ernded")})/{hh}',
        'unc_rawben_rel_max': f'={T("_c_benunc")}/{T("_c_benmax")}',
        'utilities':   f'={T("UTILITY_COSTS")}',
        'over_threshold':     f'={T("ERROR_FLAG")}',
        'total_error_amount': f'=ROUND(ABS({T("ERROR_AMOUNT")}),0)',
    }
    return helpers, feats


# ── validation: mirror every formula in pandas, score against the frame ──────
def _excel_floor(x, sig=1.0):
    return np.floor(np.asarray(x, float) / sig) * sig


def mirror_features(raw, ftabs):
    """Compute what the Excel formulas will produce, from the raw block."""
    g = lambda c: raw[c].fillna(0).astype(float).values     # Excel blank -> 0
    yd, sd, ma = ftabs
    fy = g('REVIEW_FISCAL_YEAR').astype(int)
    sz20 = np.clip(g('HOUSEHOLD_SIZE'), 1, 20).astype(int)
    hh = np.maximum(g('HOUSEHOLD_SIZE'), 1)
    lk = lambda tbl, col: np.array([tbl.loc[tbl.year <= y, col].iloc[-1] for y in fy])
    stdded = np.array([sd.loc[sd.year <= y, str(s)].iloc[-1] for y, s in zip(fy, sz20)])
    benmax = np.array([ma.loc[ma.year <= y, str(s)].iloc[-1] for y, s in zip(fy, sz20)])
    eld = ((g('NUM_ELDERLY') + g('NUM_DISABLED')) > 0).astype(int)
    ernded = _excel_floor(g('EARNED_INCOME') * 0.2)
    netbs = (g('EARNED_INCOME') + g('UNEARNED_INCOME')
             - (ernded + g('DEPENDENT_CARE_DEDUCTION') + g('MEDICAL_DEDUCTION')
                + g('CHILD_SUPPORT_DEDUCTION') + stdded))
    maxsh = np.where(eld == 1, 1e9, lk(yd, 'max_shelter_deduction'))
    sltded = np.minimum(np.maximum(g('RENT') + g('UTILITY_COSTS')
                                   - np.maximum(netbs * 0.5, 0), 0), maxsh)
    netan = _excel_floor(netbs - (sltded + g('HOMELESS_DEDUCTION')))
    benunc = _excel_floor(benmax - 0.3 * netan)
    minal = lk(yd, 'min_allotment')
    benrec = np.minimum(np.maximum(
        benunc, np.where(g('HOUSEHOLD_SIZE') < 3, minal, 0)), benmax)
    fyshare = pd.Series(g('CATEGORICALLY_ELIGIBLE')).groupby(fy).transform('mean')
    out = {
        'fiscal_year': fy, 'hh_size_raw': g('HOUSEHOLD_SIZE'),
        'HH_size_n': g('HOUSEHOLD_SIZE'),
        'hh_group': np.where(g('HOUSEHOLD_SIZE') >= 4, '4+',
                             np.where(g('HOUSEHOLD_SIZE') >= 2, '2-3', '1')),
        'bbce_state_i': (fyshare.values >= 0.5).astype(int),
        'children_i': (g('NUM_CHILDREN') > 0).astype(int),
        'count_divisible_by_100': g('NUM_AMOUNTS_DIVISIBLE_BY_100'),
        'elderly_disabled_i': eld,
        'expedited_i': g('EXPEDITED'),
        'homeless': g('HOMELESS_FLAG'),
        'married': g('MARRIED_FLAG'),
        'medical_deductions': g('MEDICAL_DEDUCTION'),
        'months_since_cert_n': g('MONTHS_SINCE_CERT'),
        'percent_abawd': g('NUM_ABAWD') / hh,
        'earned_by_hh_size': g('EARNED_INCOME') / hh,
        'unearned_by_hh_size': g('UNEARNED_INCOME') / hh,
        'gross_by_hh_size': (g('EARNED_INCOME') + g('UNEARNED_INCOME')) / hh,
        'rawben_rel_max': benrec / benmax,
        'shelter_expenses_by_hh_size': (g('RENT') + g('UTILITY_COSTS')) / hh,
        'total_deductions_by_hh_size': (g('DEPENDENT_CARE_DEDUCTION')
                                        + g('CHILD_SUPPORT_DEDUCTION') + sltded
                                        + g('MEDICAL_DEDUCTION') + ernded) / hh,
        'unc_rawben_rel_max': benunc / benmax,
        'utilities': g('UTILITY_COSTS'),
        'over_threshold': g('ERROR_FLAG'),
        'total_error_amount': np.round(np.abs(g('ERROR_AMOUNT'))),
    }
    return pd.DataFrame(out)


def validate(raw, frame, elem_free, feat_names):
    ad = os.path.join(REPO, 'additional_data')
    yd = pd.read_csv(os.path.join(ad, 'year_data.csv')).sort_values('year')
    sd = pd.read_csv(os.path.join(ad, 'standard_deductions.csv')).sort_values('year')
    ma = pd.read_csv(os.path.join(ad, 'max_allotments.csv')).sort_values('year')
    for t in (sd, ma):
        t.columns = [str(c).strip() for c in t.columns]
    mir = mirror_features(raw, (yd, sd, ma))
    print(f'\nformula validation vs the munged frame '
          f'({len(frame)} rows, reconstructed pre-QC-review inputs):')
    print(f'  {"feature":32s} {"all rows":>9s}')
    worst = []
    for c in feat_names:
        if c not in mir.columns or c not in frame.columns:
            continue
        if c == 'hh_group':
            ok = mir[c].astype(str).values == frame[c].astype(str).values
        else:
            a = mir[c].astype(float).values
            # over_threshold / total_error_amount live in the workbook in
            # build_workbook_v2's transformed form; compare against that
            if c == 'over_threshold':
                b = frame['is_error'].astype(float).values
            elif c == 'total_error_amount':
                b = np.round(np.abs(pd.to_numeric(
                    frame[c], errors='coerce').fillna(0).values))
            else:
                b = pd.to_numeric(frame[c], errors='coerce').fillna(0).values
            ok = np.isclose(a, b, atol=1e-6, rtol=1e-9)
        r_all = ok.mean()
        print(f'  {c:32s} {r_all:9.1%}')
        worst.append((c, r_all))
    bad = [c for c, r in worst if r < 0.995]
    if bad:
        print(f'  WARNING: match below 99.5% for: {", ".join(bad)}')
    return not bad


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('live_workbook')
    ap.add_argument('-o', '--out', required=True)
    ap.add_argument('--state', default=os.environ.get('SNAP_STATE', 'WA'))
    a = ap.parse_args()
    cfg = STATE_REGISTRY.get(a.state.upper())
    frame_csv = os.path.join(PKG, '.frames', f'{cfg["abbr"].lower()}_frame.csv')

    raw, elem_free, frame = raw_frame(cfg, frame_csv)
    print(f'raw extract: {len(raw)} rows x {len(RAW_COLS)} input fields')

    shutil.copy(a.live_workbook, a.out)
    wb = openpyxl.load_workbook(a.out)
    # the Dashboard (per-rule threshold tuner) stays hidden: engine plumbing;
    # unhide it in Excel to tune thresholds interactively
    dat = wb['Data']
    NROW = dat.max_row
    assert NROW - 1 == len(raw), f'workbook has {NROW-1} cases, extract {len(raw)}'

    hdr = [c.value for c in next(dat.iter_rows(min_row=1, max_row=1))]
    ncol0 = len(hdr)
    holdout_year = int(pd.to_numeric(
        frame['fiscal_year'], errors='coerce').max())    # split label only
    R = federal_tables(wb, holdout_year)
    helpers, feats = feature_formulas(R)

    missing = [h for h in hdr if not h.startswith('_') and h not in feats]
    assert not missing, f'no formula for Data columns: {missing}'
    # Excel table column names are case-insensitively unique; a collision makes
    # Excel reject the whole file as damaged, so fail here instead
    names = hdr + [h for h, _ in helpers] + RAW_COLS
    low = [n.lower() for n in names]
    dups = sorted({n for n in low if low.count(n) > 1})
    assert not dups, f'case-insensitive duplicate Data columns: {dups}'
    ok = validate(raw, frame, elem_free, [h for h in hdr if h in feats])
    if not ok:
        # HARD GATE (2026-08-16): the demo must sit on the reconstructed
        # pre-QC-review scale the rules were mined on — a state's internal
        # data is effectively that scale, and scoring rules against
        # QC-corrected values misleads (v1 post-mortem: 21 of 114 WA rules
        # never fired). With reconstructed inputs the match is 100%; any
        # shortfall means corrected values crept back into the input block.
        if os.environ.get('SNAP_ALLOW_VALIDATION_MISMATCH') == '1':
            print('  SNAP_ALLOW_VALIDATION_MISMATCH=1: continuing despite '
                  'validation failures (debug only)')
        else:
            raise SystemExit(
                'validation failed: formula features do not reproduce the '
                'research frame. The Data-tab demo must carry RECONSTRUCTED '
                '(pre-QC-review) values, never QC-corrected ones — see the '
                'columns flagged above. Set SNAP_ALLOW_VALIDATION_MISMATCH=1 '
                'only to debug.')

    # 1. helper columns, then the raw contract, appended AFTER the existing
    #    feature + hit columns so every positional reference stays valid
    c = ncol0
    helper_idx = []
    for name, formula in helpers:
        c += 1
        helper_idx.append(c)
        dat.cell(row=1, column=c, value=name)
        for rr in range(2, NROW + 1):
            dat.cell(row=rr, column=c, value=formula)
    raw0 = c + 1
    for name in RAW_COLS:
        c += 1
        dat.cell(row=1, column=c, value=name)
        vals = raw[name].tolist()
        for i, v in enumerate(vals):
            dat.cell(row=2 + i, column=c,
                     value=None if (isinstance(v, float) and np.isnan(v)) else v)
    LAST = CL(c)

    # 2. the feature columns become formulas (values overwritten in place)
    n_swapped = 0
    for ci, name in enumerate(hdr, 1):
        if name in feats:
            for rr in range(2, NROW + 1):
                dat.cell(row=rr, column=ci, value=feats[name])
            n_swapped += 1
    print(f'features -> formulas: {n_swapped} columns | helpers: {len(helpers)} | '
          f'raw: {len(RAW_COLS)} -> Data!A1:{LAST}{NROW}')

    # 3. rebind the table over the widened range
    for t in list(dat.tables):
        del dat.tables[t]
    tbl = Table(displayName=TABLE, ref=f'A1:{LAST}{NROW}')
    tbl.tableStyleInfo = TableStyleInfo(name='TableStyleLight1', showRowStripes=False)
    dat.add_table(tbl)

    # 4. presentation: helpers hidden; features blue; raw amber, collapsible
    feat_fill = PatternFill('solid', fgColor='DDEBF7')
    raw_fill = PatternFill('solid', fgColor='FCE4D6')
    for i in range(1, ncol0 + 1):
        dat.cell(row=1, column=i).fill = feat_fill
    for i in helper_idx:
        dat.column_dimensions[CL(i)].hidden = True
    for i in range(raw0, c + 1):
        dat.cell(row=1, column=i).fill = raw_fill
    # collapse button sits left of the group (summary-left): the feature block
    # collapses to leave the raw block, and vice versa
    dat.sheet_properties.outlinePr.summaryRight = False
    for i in range(2, ncol0 + 1):
        if not hdr[i - 1].startswith('_'):
            dat.column_dimensions[CL(i)].outline_level = 1
    for i in range(raw0 + 2, c + 1):
        dat.column_dimensions[CL(i)].outline_level = 1
    note = ('Feature columns (blue) are FORMULAS computed from the input '
            'fields (amber block to the right); do not paste values over them. '
            'A state supplies ONLY the amber columns, mapping its own reported '
            'case fields onto them. See the Data Dictionary tab for every '
            'definition and the QC technical-manual crosswalk.')
    dat.cell(row=1, column=1).comment = Comment(note, 'snap_dashboard')
    dat.freeze_panes = 'B2'

    data_dictionary(wb, hdr)
    background_tab(wb, cfg['name'])

    # fill the viewer's CASE_ID lead column, now that Data carries CASE_ID
    vname = 'See cases flagged by a rule'
    if vname in wb.sheetnames:
        wsv = wb[vname]
        hdr10 = {wsv.cell(row=10, column=ci).value: ci for ci in range(1, 200)}
        cid, drc = hdr10.get('CASE_ID'), hdr10.get('data_row')
        if cid and drc:
            cum = f'${CL(drc + 3)}$12:${CL(drc + 3)}${10 + NROW}'
            for rr in range(11, 71):
                wsv.cell(row=rr, column=cid).value = (
                    f'=IFERROR(INDEX({TABLE}[CASE_ID],MATCH(ROW()-10,{cum},0)),"")')
            print('viewer CASE_ID column wired to the Data table')

    # tab colors: blue = read, amber = supply data
    for name, color in (('Data', 'C55A11'), ('Data Dictionary', 'C55A11')):
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color

    wb.save(a.out)
    print('saved', a.out)


if __name__ == '__main__':
    main()
