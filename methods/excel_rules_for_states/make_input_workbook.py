"""
Reconstruction stage: the workbook accepts RAW case fields and computes every
model feature with live formulas, so a state that can run neither Python nor R
can still use it. The input columns carry generic reported-value concept names
in features.R's state_col_map style (HOUSEHOLD_SIZE, EARNED_INCOME, RENT, ...)
that a state maps its own system's fields onto; the Data Dictionary tab
carries the crosswalk to the SNAP QC technical documentation's variables.

Data tab layout (one Excel table, columns bound by name; see workbook_layout):
  [ input contract ]    what a state pastes (workbook_layout.RAW_COLS, on the
                        LEFT): unit counts and indicators, income totals,
                        deductions, shelter costs, and the QC review outcome
                        (ORIGINAL_BENEFIT_AMOUNT, CORRECTED_BENEFIT_AMOUNT,
                        STATUS). The error amount and the over-threshold flag
                        are computed from the benefit pair against the
                        FederalTables error-threshold table. Headers exist
                        from the first build stage; this stage fills the
                        values.
  [ feature columns ]   the model features, in the SAME columns the built
                        workbook put them, now formulas — positional
                        references from the other sheets stay valid
  [ hit columns ]       per-case rule tests, carried over from the LIVE build
  [ helper columns ]    the benefit-recomputation chain, appended here, hidden

Federal parameter tables (standard deduction, max allotment, shelter cap,
minimum allotment and QC error threshold, by fiscal year x household size)
live on a hidden FederalTables sheet; formulas look them up by year and
size. The sheet also carries the state's USDA state-options rows (BBCE and
the other options) as reference; unhide it to inspect or extend.

The demo input block carries the research frame's RECONSTRUCTED
(pre-QC-review) values — the same scale the rules were mined on — so the
workbook's figures match the research pipeline's. A state pasting its own
data supplies its ordinary as-reported fields (internal data carries no QC
corrections, so nothing needs reconstructing there). The built-in validation
prints the per-feature match rate against the frame on every build.

Usage:  python make_input_workbook.py <LIVE_workbook.xlsx> -o <out.xlsx> [--state WA]
Then:   python postprocess_workbook.py <out.xlsx>   (drops the stale calc chain)
"""
import argparse
import os
import shutil

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter as CL
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula

import states as STATE_REGISTRY
from workbook_layout import (DATA_SHEET, BLENDED_SHEET, DICT_SHEET,
                             EXPORT_SHEET, SCREEN_SHEET, FLAGGED_SHEET,
                             SHARE_SHEET, VIEWER_SHEET, RAW_COLS, qref)
from live_formulas import (countifs, make_table, read_delivery_tab,
                           rule_term, selection_refs)

TABLE = 'CaseData'
PKG = os.path.dirname(os.path.abspath(__file__))


def find_repo(start):
    cands = [os.environ['SNAP_REPO']] if os.environ.get('SNAP_REPO') else []
    d = start
    for _ in range(6):
        cands += [d, os.path.join(d, 'snap_qc')]
        d = os.path.dirname(d)
    for c in cands:
        if (os.path.isfile(os.path.join(c, 'reg_model_data.rds'))
                and os.path.isfile(os.path.join(c, 'rule_mining_helpers.R'))):
            return os.path.abspath(c)
    raise SystemExit('cannot find the snap_qc checkout; set SNAP_REPO')


REPO = find_repo(PKG)

# The raw input contract (one row per reviewed case) lives in
# workbook_layout.RAW_COLS, shared with build_workbook_v2, which writes the
# headers into the Data tab's left block; this stage fills the values.


def raw_frame(cfg, frame_csv):
    """The input block for exactly the frame's rows, in the frame's row order.

    The demo carries the research frame's RECONSTRUCTED (pre-QC-review)
    values — the same scale the rules were mined on — so the workbook's
    figures match the research pipeline's (decision 2026-08-16). A state
    pasting its own data supplies its ordinary as-reported fields; internal
    data carries no QC corrections, so no reconstruction is needed there.
    Everything comes from the frame export; the .sav files are not read.
    """
    frame = pd.read_csv(frame_csv, dtype={'hhldno': str, 'stratum': str})
    need = ['rawearn', 'rawunearn', 'rawdepded', 'rawcsded', 'rawrent',
            'rawhomeless_ded', 'fsnkid', 'fsnelder', 'fsndis',
            'count_abawd', 'cat_elig', 'rawben', 'benefit_amount_FS', 'status',
            'fscsexp']
    missing = [c for c in need if c not in frame.columns]
    assert not missing, (f'frame export lacks reconstructed input fields '
                         f'{missing}; re-export it (make_state.py --refresh)')

    g = lambda c: pd.to_numeric(frame[c], errors='coerce')
    out = pd.DataFrame({
        'CASE_ID': frame['hhldno'].astype(str).str.strip(),
        'REVIEW_FISCAL_YEAR': g('fiscal_year'),
        'HOUSEHOLD_SIZE': g('hh_size_raw'),
        'NUM_CHILDREN': g('fsnkid'),
        'NUM_ELDERLY': g('fsnelder'),
        'NUM_DISABLED': g('fsndis'),
        'NUM_ABAWD': g('count_abawd'),
        'MARRIED_FLAG': g('married').fillna(0).astype(int),
        'EXPEDITED': g('expedited_i').fillna(0).astype(int),
        'CATEGORICALLY_ELIGIBLE': (g('cat_elig') >= 1).fillna(False).astype(int),
        'HOMELESS_FLAG': g('homeless').fillna(0).astype(int),
        'MONTHS_SINCE_CERT': g('months_since_cert_n'),
        'EARNED_INCOME': g('rawearn'),
        'UNEARNED_INCOME': g('rawunearn'),
        'MEDICAL_DEDUCTION': g('medical_deductions'),
        'DEPENDENT_CARE_DEDUCTION': g('rawdepded'),
        'CHILD_SUPPORT_DEDUCTION': g('rawcsded'),
        # informational input (no rule reads it): the QC expense field, for
        # exclusion-state reconciliation — see its dictionary entry
        'CHILD_SUPPORT_EXPENSES': g('fscsexp'),
        'HOMELESS_DEDUCTION': g('rawhomeless_ded'),
        'RENT': g('rawrent'),
        'UTILITY_COSTS': g('utilities'),
        # QC outcome: the benefit as issued (QC manual RAWBEN — the reported
        # value, exactly what a state's own system holds) and the QC-corrected
        # benefit (FSBEN). Their rounded absolute difference IS the frame's
        # total_error_amount by construction, so the workbook's recomputed
        # outcome matches the frame 100% (R1 check, 2026-08-18).
        'ORIGINAL_BENEFIT_AMOUNT': g('rawben'),
        'CORRECTED_BENEFIT_AMOUNT': g('benefit_amount_FS'),
        # review disposition: 2 = overissuance, 3 = underissuance,
        # 4 = ineligible (never present in public files); 1 = correct
        'STATUS': g('status'),
    })
    assert list(out.columns) == RAW_COLS
    # the demo block must carry COMPLETE rows: in the workbook a blank cell
    # means missing (the per-row feature guards blank the case), while the
    # frame's reconstruction already zero-imputes its missing inputs — write
    # those zeros explicitly so the shipped figures match the frame. STATUS
    # is excluded: 0 is not a code, blank = unknown.
    for c in out.columns:
        if c not in ('CASE_ID', 'STATUS'):
            out[c] = out[c].fillna(0)
    # with reconstructed inputs there is no restoration gap; keep the
    # element-free split trivially all-True for the validation report
    return out, np.ones(len(frame), bool), frame


def federal_tables(wb, state_name=None):
    """Hidden reference sheet (hidden again 2026-08-19 by request): the
    year-level parameters (max shelter, minimum allotment, QC error
    threshold), the year x size tables (standard deduction, max allotment),
    and the state's USDA state-options rows (BBCE and the other options).
    Formulas look values up by year with MATCH(..,1), so unhiding and
    appending a row per new fiscal year updates everything."""
    ad = os.path.join(REPO, 'additional_data')
    yd = pd.read_csv(os.path.join(ad, 'year_data.csv')).dropna(axis=1, how='all')
    yd.columns = [str(c).strip() for c in yd.columns]
    sd = pd.read_csv(os.path.join(ad, 'standard_deductions.csv')).dropna(axis=1, how='all')
    ma = pd.read_csv(os.path.join(ad, 'max_allotments.csv')).dropna(axis=1, how='all')
    for t in (sd, ma):
        t.columns = [str(c).strip() for c in t.columns]

    ws = wb.create_sheet('FederalTables')
    ws.sheet_state = 'hidden'          # reference plumbing; unhide to inspect
                                       # or to append a new fiscal year
    ws.sheet_view.showGridLines = False
    blue = PatternFill('solid', fgColor='2F5496')
    gray = PatternFill('solid', fgColor='F2F2F2')
    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = 'Federal SNAP parameters by fiscal year'
    c.fill = blue; c.font = Font(bold=True, size=14, color='FFFFFF')
    ws.row_dimensions[1].height = 26
    ws['A2'] = ('Every formula in the workbook looks these up by fiscal year. To use a '
                'new fiscal year, append one row to each table below; everything '
                'updates automatically. error_threshold is the federal QC tolerance: a '
                'case is a payment error when its error amount EXCEEDS this dollar '
                'amount for its review year.')
    ws['A2'].font = Font(size=10)
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A2:J2')
    ws.row_dimensions[2].height = 42
    for ci, h in enumerate(('year', 'max_shelter', 'min_allotment',
                            'error_threshold'), 1):
        hc = ws.cell(row=5, column=ci, value=h)
        hc.fill = gray; hc.font = Font(bold=True, size=10)
    years = sorted(yd['year'])
    for i, y in enumerate(years):
        r = yd[yd.year == y].iloc[0]
        ws.cell(row=6 + i, column=1, value=int(y))
        ws.cell(row=6 + i, column=2, value=float(r['max_shelter_deduction']))
        ws.cell(row=6 + i, column=3, value=float(r['min_allotment']))
        ws.cell(row=6 + i, column=4, value=float(r['error_threshold']))
    ny = len(years)
    for cl, w in (('A', 8), ('B', 12), ('C', 13), ('D', 15)):
        ws.column_dimensions[cl].width = w

    def block(df, row0, label):
        hc = ws.cell(row=row0, column=5, value=label)
        hc.font = Font(bold=True, size=10)
        ws.cell(row=row0 + 1, column=5, value='year').fill = gray
        for s in range(1, 21):
            sc = ws.cell(row=row0 + 1, column=5 + s, value=s)
            sc.fill = gray; sc.font = Font(bold=True, size=10)
        yrs = sorted(df['year'])
        for i, y in enumerate(yrs):
            ws.cell(row=row0 + 2 + i, column=5, value=int(y))
            r = df[df.year == y].iloc[0]
            for s in range(1, 21):
                ws.cell(row=row0 + 2 + i, column=5 + s, value=float(r[str(s)]))
        return len(yrs), row0 + 2
    n_sd, sd0 = block(sd.reset_index(), 5, 'standard deduction')
    n_ma, ma0 = block(ma.reset_index(), 5 + n_sd + 4, 'max allotment')

    # the state's rows from the USDA state-options panel, as reference only
    # (nothing computes from them; BBCE membership is already baked into the
    # delivered rules by the rule_selection transform)
    opt0 = ma0 + n_ma + 3
    try:
        so = pd.read_csv(os.path.join(ad, 'snap_state_options_all_years.csv'))
        rows = so[so['State'] == state_name] if state_name else so.iloc[0:0]
    except OSError:
        rows = None
    if rows is not None and len(rows):
        hc = ws.cell(row=opt0, column=1,
                     value=f'USDA state options — {state_name} (reference only; '
                           'BBCE membership is already reflected in the rule list)')
        hc.font = Font(bold=True, size=10)
        ws.merge_cells(start_row=opt0, start_column=1, end_row=opt0, end_column=10)
        for ci, h in enumerate(rows.columns, 1):
            hc = ws.cell(row=opt0 + 1, column=ci, value=str(h))
            hc.fill = gray; hc.font = Font(bold=True, size=9)
        for ri, (_, rr) in enumerate(rows.iterrows()):
            for ci, h in enumerate(rows.columns, 1):
                v = rr[h]
                if pd.isna(v):
                    v = None
                elif hasattr(v, 'item'):        # numpy scalar -> python scalar
                    v = v.item()
                ws.cell(row=opt0 + 2 + ri, column=ci, value=v).font = Font(size=9)
    # hidden per-year SUA-mode block (2026-08-22): one ARRAY-entered
    # MODE.SNGL per fiscal year over the CaseData columns, so the mode is
    # evaluated as an array exactly once per year and each Data row reads
    # it with a plain INDEX/MATCH (see feature_formulas: an in-row array
    # silently collapses). Years span the FederalTables year list so any
    # pasted year resolves; a year with no positive utilities reads 0.
    # Written AFTER the CaseData table exists (main() calls
    # write_sua_mode_block once the table is rebound), since the array
    # formulas reference the table's columns.
    sua0 = opt0 + (len(rows) + 3 if rows is not None and len(rows) else 0) + 2
    ws.cell(row=sua0, column=12, value='SUA mode by fiscal year (hidden helper; '
            'array formulas over the pasted data)').font = Font(bold=True, size=9)
    ws.cell(row=sua0 + 1, column=12, value='year').fill = gray
    ws.cell(row=sua0 + 1, column=13, value='mode of positive UTILITY_COSTS').fill = gray
    for i, y in enumerate(years):
        ws.cell(row=sua0 + 2 + i, column=12, value=int(y))
    ws.column_dimensions['L'].width = 8
    ws.column_dimensions['M'].width = 30
    ws.freeze_panes = 'A3'
    return {
        'SUAYRS':  f'FederalTables!$L${sua0 + 2}:$L${sua0 + 1 + ny}',
        'SUAMODE': f'FederalTables!$M${sua0 + 2}:$M${sua0 + 1 + ny}',
        'SUAROW0': sua0 + 2,
        'YEARS':  f'FederalTables!$A$6:$A${5 + ny}',
        'MAXSH':  f'FederalTables!$B$6:$B${5 + ny}',
        'MINAL':  f'FederalTables!$C$6:$C${5 + ny}',
        'ERRTHR': f'FederalTables!$D$6:$D${5 + ny}',
        'SDYRS':  f'FederalTables!$E${sd0}:$E${sd0 + n_sd - 1}',
        'SDBLK':  f'FederalTables!$F${sd0}:$Y${sd0 + n_sd - 1}',
        'MAYRS':  f'FederalTables!$E${ma0}:$E${ma0 + n_ma - 1}',
        'MABLK':  f'FederalTables!$F${ma0}:$Y${ma0 + n_ma - 1}',
        'HOLDOUT': 'FederalTables!$B$3',
    }


def write_sua_mode_block(wb, R, nrow):
    """The per-year SUA-mode cells on FederalTables, ARRAY-entered so the
    IF(...) inside MODE.SNGL evaluates as an array (a plain table cell would
    collapse it). References are absolute Data-sheet column ranges rather
    than structured refs: Excel is free to rewrite structured refs inside an
    array formula, and a fixed generous range (rows 2..MAXR) covers pasted
    rows; blanks contribute neither to the year test nor the positive test.
    MAXR is sized to the 250k-row design ceiling."""
    from openpyxl.utils import get_column_letter
    ws = wb['FederalTables']
    dat = wb[DATA_SHEET]
    hdr = [c.value for c in next(dat.iter_rows(min_row=1, max_row=1))]
    cy = get_column_letter(hdr.index('REVIEW_FISCAL_YEAR') + 1)
    cu = get_column_letter(hdr.index('UTILITY_COSTS') + 1)
    dq = qref(DATA_SHEET)
    MAXR = 250000
    yrs = f'{dq}!${cy}$2:${cy}${MAXR}'
    uts = f'{dq}!${cu}$2:${cu}${MAXR}'
    r0 = R['SUAROW0']
    n = 0
    r = r0
    while ws.cell(row=r, column=12).value is not None:
        ref = f'M{r}'
        ws[ref] = ArrayFormula(
            ref,
            f'=IFERROR(_xlfn.MODE.SNGL(IF(({yrs}=$L{r})*({uts}>0),ROUND({uts},0))),0)')
        n += 1
        r += 1
    return n


# ── Start Here: KPIs, how-to, and background ─────────────────────────────────
GITHUB = 'https://github.com/giannella/snap_qc'


def background_tab(wb, state_name):
    ws = wb.create_sheet('Start Here', 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '2F5496'
    blue = PatternFill('solid', fgColor='2F5496')
    # orange (accent 6, lighter 60%): matches the rules tabs' overall-results rows
    accent = PatternFill('solid', fgColor='F8CBAD')
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 58
    ws.merge_cells('A1:B1')
    c = ws['A1']; c.value = f'SNAP QC payment-error rule lists — {state_name}'
    c.fill = blue; c.font = Font(bold=True, size=16, color='FFFFFF')
    ws.row_dimensions[1].height = 30

    r = 3

    def para(txt, bold=False, height=None):
        nonlocal r
        ws.merge_cells(f'A{r}:B{r}')
        c = ws.cell(row=r, column=1, value=txt)
        # explainer text at 12pt black (feedback 2026-08-21)
        c.font = Font(bold=bold, size=14 if bold else 12)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        # ~100 wrapped chars per line across the A:B merge at 12pt, 15pt per line
        ws.row_dimensions[r].height = height or (
            20 if bold else max(30, 15 * -(-len(txt) // 100)))
        r += 2

    def link(label, url):
        nonlocal r
        c = ws.cell(row=r, column=1, value=label)
        c.hyperlink = url
        c.font = Font(color='0563C1', underline='single')
        r += 1

    # live KPI block: reads the rules tab's union row, recomputes on paste
    kpis = [
        ('Cases in this workbook',
         f'=COUNTA({TABLE}[CASE_ID])', '#,##0'),
        ('Cases flagged by the selected rules',
         f"='{BLENDED_SHEET}'!G6", '#,##0'),
        ('Errors flagged by the selected rules',
         f"='{BLENDED_SHEET}'!H6", '#,##0'),
        ('Precision: share of flagged cases that are true errors',
         f"='{BLENDED_SHEET}'!D6", '0.0%'),
        ('Base error rate, all cases (compare precision against this)',
         f'=COUNTIF({TABLE}[over_threshold],1)/COUNTA({TABLE}[CASE_ID])', '0.0%'),
        # feedback 2026-08-21: the per-case average reads better than the
        # raw total (K6 = the union row's error dollars caught / cases flagged)
        ('Average error $ per case flagged (includes $0 error cases)',
         f"='{BLENDED_SHEET}'!K6", '$#,##0'),
        # overissuance share from the benefit pair (revision 2026-08-18): the
        # demo fills it automatically because the shipped input block carries
        # ORIGINAL/CORRECTED_BENEFIT_AMOUNT for every case
        ('Share of error cases that are overissuances',
         f'=IFERROR(SUMPRODUCT(({TABLE}[over_threshold]=1)'
         f'*({TABLE}[ORIGINAL_BENEFIT_AMOUNT]>{TABLE}[CORRECTED_BENEFIT_AMOUNT]))'
         f'/COUNTIF({TABLE}[over_threshold],1),"")', '0.0%'),
    ]
    for label, f, fmt in kpis:
        ws.cell(row=r, column=1, value=label).font = Font(size=13)
        c = ws.cell(row=r, column=2)
        c.value = f
        c.font = Font(bold=True, size=18)
        c.number_format = fmt
        c.fill = accent
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1

    para('What this is', bold=True)
    # wording agreed 2026-08-21; split into two paragraphs at "Paste in" and
    # the review-budget sentence added per the same day's feedback
    para('This workbook helps you test potential rules for flagging cases at risk of '
         'payment errors using your own internal data. It then allows you to either '
         'a) export those rules for use somewhere else or b) apply the rules to new '
         'cases here in Excel (i.e., this workbook will flag cases based on the rules '
         'you\'ve selected). It is critical to test rules before implementing them - '
         'you have access to more recent and more comprehensive data about errors '
         'than what was used to generate these rules. Some rules may no longer catch '
         'errors, others may flag too many cases without errors.')
    para('Paste in your own recent data (e.g., FY2025/FY2026 QC or QA cases), see '
         'how every rule performs on it, and keep the rules that perform well (e.g., '
         'at least 3-4 in 10 cases flagged have errors) and that make sense given '
         'what you know about your secondary review processes (e.g., can the types '
         'of errors flagged by the rule be found and fixed?). Add or remove rules to '
         'get to your "review budget" (the percentage of cases you can review '
         'annually). Each rule is a short, readable condition on case fields (for '
         'example: household size, income per person, deductions, benefit relative '
         f'to the maximum), selected from national and {state_name} rules mined on '
         'the public USDA SNAP Quality Control (QC) files for FY2022-2024.')
    para('Yellow highlighting means a cell is interactive: the yellow columns on the '
         f'"{DATA_SHEET}" tab are where you paste your data, and the yellow "Include?" '
         f'column on the "{BLENDED_SHEET}" tab (TRUE/FALSE) is where you keep or drop '
         'each rule. Everything else recomputes automatically.')

    para('How to use this workbook', bold=True)
    # step structure and wording agreed 2026-08-21: steps 1-3 are the core
    # path; steps 4 and 5 are ALTERNATIVE paths to using the selected rules,
    # indented under step 3; step 6 applies to any path
    for step in [
        f'1.  Map your data to the dictionary: the "{DICT_SHEET}" tab defines every '
        'input column and gives the QC technical-manual crosswalk.',
        f'2.  Import your testing data: paste your cases into the yellow columns of '
        f'the "{DATA_SHEET}" tab. Every figure in the workbook recomputes on your '
        'data. The gray columns to the right of the yellow ones are computed '
        'values (the variables the rules actually test, such as income per '
        'person): do not paste over them, and look there, not in your own '
        'columns, for what a rule is reading.',
        f'3.  Select rules on the "{BLENDED_SHEET}" tab: set the yellow Include? cell '
        'to FALSE for any rule you do not want. The combined results update in the '
        'orange rows at the top of that tab and in the figures above. At any point, '
        f'use the "{VIEWER_SHEET}" tab to pick a rule and see the actual cases it '
        'flags in your data, with the columns the rule uses highlighted in blue: an '
        'easy way to sanity-check a rule before keeping it.',
        'Steps 4 and 5 are alternative paths to using the rules to flag new '
        'high-risk cases: step 4 is for exporting them to use somewhere else, and '
        'step 5 is to use them here in Excel.',
        f'      4.  Export the selected rules (the "{EXPORT_SHEET}" tab): the rules '
        'still set to TRUE, with their exact logic. Filter your caseload in Excel, '
        'turn them into a query, or send them to your vendor.',
        f'      5.1  Screen new cases (the "{SCREEN_SHEET}" tab): paste cases with '
        'no review outcome into its yellow columns.',
        f'      5.2  See the flagged cases (the "{FLAGGED_SHEET}" tab): every new '
        'case a selected rule flags, with the rule that flagged it.',
        f'6.  Optional: share aggregate results back (the "{SHARE_SHEET}" tab): '
        'whichever path you took, this tab summarizes each rule\'s performance on '
        'your data. If you do not mind sending us back these aggregate performance '
        'numbers by rule, that will help us improve the models and the rules we '
        'generate for you as well as other states. Please copy/paste as values into '
        'a new workbook (do not send internal case-level data) in order to send '
        'these results back to us. If you have trouble, please reach out to '
        'eric.giannella@georgetown.edu.',
    ]:
        ws.merge_cells(f'A{r}:B{r}')
        c = ws.cell(row=r, column=1, value=step)
        c.font = Font(size=12)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        # ~100 wrapped chars per line across the A:B merge at 12pt, 15pt per line
        ws.row_dimensions[r].height = max(28, 15 * -(-len(step) // 100))
        r += 1
    r += 1

    para('Why to use internal data, not just the public file', bold=True)
    para('The public QC files are only available through FY24 and do not contain cases '
         'that were deemed ineligible, which in a few states represents a large share of '
         'errors. In addition, there are not that many observations per state per year '
         '(several hundred to a little over 1,000) so most state-derived rules are based '
         'on a small number of cases. Pasting the state\'s own internal data into the '
         f'"{DATA_SHEET}" tab makes every figure recompute so that you can select '
         'rules based on their performance with more comprehensive and recent data. '
         'You could try pasting QA data in to see what rules might work for QA '
         '(i.e., rules that address the subset of QC errors that can be found and '
         'fixed in QA).')
    para('How the rule lists were made', bold=True)
    # wording agreed 2026-08-21
    para('Candidate rules come from decision trees built by machine learning algorithms '
         '(xgboost and ranger) fitted to the QC data within household-size strata '
         '(1 / 2-3 / 4+), then filtered on a conservative lower confidence bound of '
         'training precision, filtered with a false-discovery-rate test against the '
         'stratum base rate, ranked using a stringent reliability measure of precision, '
         'and filled to a 10% review budget (i.e., enough rules to cover reviewing up '
         'to 10% of cases per year). Documentation can be found at the sources below:')
    link('The pipeline that builds these workbooks (GitHub)',
         f'{GITHUB}/tree/main/methods/excel_rules_for_states')
    link('The delivery rule lists, one CSV per state (GitHub)',
         f'{GITHUB}/tree/main/state_delivery_lists')
    link('The finished state workbooks (GitHub)',
         f'{GITHUB}/tree/main/methods/excel_rules_for_states/state_workbooks')
    r += 2
    para('Made by the Better Government Lab at Georgetown University and the '
         'University of Michigan', bold=True)
    ws.cell(row=r, column=1, value='Contributors:').font = Font(size=11)
    r += 1
    for name in ('Eric Giannella', 'Ziyu Shu', 'Ben Molin', 'Rachael Zuppke'):
        ws.cell(row=r, column=1, value=name).font = Font(size=11)
        r += 1
    ws.freeze_panes = 'A2'
    return ws


# ── data dictionary: one line per column on the Data tab ─────────────────────
# Input names follow features.R's state_col_map style; each description cross-
# references the SNAP QC technical documentation's variables (Origin R =
# reported on the QC Review Schedule, C = constructed during editing). A state
# supplies its own as-reported value for each concept; the public demo carries
# the research frame's RECONSTRUCTED (pre-QC-review) value, which puts the
# workbook's figures on the same scale the rules were mined on.
RAW_DESC = {
    'CASE_ID':   'case / review identifier — row identity only. QC manual: HHLDNO.',
    'REVIEW_FISCAL_YEAR': 'federal fiscal year of the review month (Oct-Sep).',
    'HOUSEHOLD_SIZE': 'certified SNAP unit size: the members ON THE CASE, as reported. '
                      'QC manual: CERTHHSZ (Origin R, reported); FSUSIZE is the post-QC '
                      'corrected household size. Do NOT supply the household\'s total '
                      'person count (the manual\'s RAWHSIZE) — people in the home who '
                      'are not unit members do not count. The demo carries the '
                      'reconstructed pre-QC-review unit size.',
    'NUM_CHILDREN': 'children in the unit. QC manual: FSNKID.',
    'NUM_ELDERLY': 'members aged 60+. QC manual: FSNELDER.',
    'NUM_DISABLED': 'disabled members. QC manual: FSNDIS.',
    'NUM_ABAWD': 'members with ABAWD status. QC manual: count of ABWDST1-18 coded 2..5.',
    'MARRIED_FLAG':   '1 if a spouse is present in the unit. QC manual: any REL1-16 = 2.',
    'EXPEDITED': '1 if the case received expedited service. QC manual: EXPEDSER in (1, 2).',
    'CATEGORICALLY_ELIGIBLE': '1 if the unit is categorically eligible. '
                              'QC manual: CAT_ELIG >= 1.',
    'HOMELESS_FLAG':  '1 if the unit is homeless. QC manual: HOMEDED present and not 1.',
    'MONTHS_SINCE_CERT': 'months since the last certification. QC manual: LASTCERT.',
    'EARNED_INCOME': 'total earned income, $/month, as reported. QC manual: FSEARN '
                     '(= FSWAGES + FSSLFEMP + FSOTHERN) is the QC-corrected total; the '
                     'demo carries the reconstructed pre-QC-review value.',
    'UNEARNED_INCOME': 'total unearned income, $/month, as reported. QC manual: FSUNEARN '
                       '(the sum of the ~30 unearned income-type fields, FY2024 '
                       'definition) is the QC-corrected total; the demo carries the '
                       'reconstructed pre-QC-review value.',
    'MEDICAL_DEDUCTION': 'medical expense deduction, $/month. QC manual: FSMEDDED; the '
                         'demo carries the reconstructed pre-QC-review value.',
    'DEPENDENT_CARE_DEDUCTION': 'dependent care deduction, $/month. QC manual: FSDEPDED.',
    'CHILD_SUPPORT_DEDUCTION': 'child support payment deduction, $/month. QC manual: FSCSDED.',
    'CHILD_SUPPORT_EXPENSES': 'child support payments, $/month, for states that treat '
                              'child support as an income EXCLUSION rather than a '
                              'deduction. QC manual: FSCSEXP. No rule reads this column: '
                              'the rules were mined on data standardized to the DEDUCTION '
                              'treatment (exclusion-state records get the expense amount '
                              'moved into the child support deduction and gross income '
                              'recomputed), so if amounts appear here that are NOT in your '
                              'CHILD_SUPPORT_DEDUCTION column, add them into '
                              'CHILD_SUPPORT_DEDUCTION to match the scale the rules were '
                              'mined on. The demo carries the QC file\'s expense field; '
                              'for exclusion-state rows those amounts are already '
                              'reflected in the demo\'s CHILD_SUPPORT_DEDUCTION.',
    'HOMELESS_DEDUCTION': 'homeless household shelter deduction, $/month. '
                          'QC manual: HOMELESS_DED.',
    'RENT':      'rent / mortgage, $/month. QC manual: RENT; the demo carries the '
                 'reconstructed pre-QC-review value.',
    'UTILITY_COSTS': 'utilities, $/month, including standard allowances. QC manual: '
                     'UTIL; the demo carries the reconstructed pre-QC-review value.',
    'ORIGINAL_BENEFIT_AMOUNT': 'the benefit amount the case was actually issued for the '
                               'review month, $ — before any review correction. QC '
                               'manual: RAWBEN. In a state\'s own data this is simply '
                               'the benefit paid.',
    'CORRECTED_BENEFIT_AMOUNT': 'the correct benefit amount as determined by the review, '
                                '$. QC manual: FSBEN. The workbook computes the error '
                                'amount as the difference between this and '
                                'ORIGINAL_BENEFIT_AMOUNT, and flags a payment error '
                                'when that difference exceeds the review year\'s federal '
                                'QC tolerance (hidden FederalTables sheet).',
    'STATUS': 'review disposition code: 2 = overissuance, 3 = underissuance, '
              '4 = ineligible household (the entire benefit is in error), 1 = correct. '
              'QC manual: STATUS. The public demo data carries no code-4 cases (the '
              'public QC files exclude ineligible households); include them when '
              'pasting internal data so ineligibility catches can be counted.',
}
FEAT_DESC = {
    'fiscal_year': 'REVIEW_FISCAL_YEAR, unchanged',
    'hh_size_raw': 'HOUSEHOLD_SIZE, unchanged',
    'hh_group':    'household-size stratum: 1 / 2-3 / 4+ (every rule applies within one stratum)',
    'HH_size_n':   'HOUSEHOLD_SIZE as a number, used inside rules',
    'bbce_state_i': 'state-year flag: 1 when at least half the year\'s cases are '
                    'CATEGORICALLY_ELIGIBLE (the state runs Broad-Based Categorical '
                    'Eligibility)',
    'children_i':  '1 if NUM_CHILDREN > 0',
    'elderly_disabled_i': '1 if NUM_ELDERLY + NUM_DISABLED > 0',
    'expedited_i': 'EXPEDITED, unchanged',
    'homeless':    'HOMELESS_FLAG, unchanged',
    'married':     'MARRIED_FLAG, unchanged',
    'medical_deductions': 'MEDICAL_DEDUCTION, unchanged',
    'months_since_cert_n': 'MONTHS_SINCE_CERT, unchanged',
    'percent_abawd': 'NUM_ABAWD / HOUSEHOLD_SIZE (the research frame divides by CERTHHSZ, '
                     'the certified unit size — the same concept HOUSEHOLD_SIZE carries)',
    'earned_by_hh_size':   'EARNED_INCOME / HOUSEHOLD_SIZE',
    'unearned_by_hh_size': 'UNEARNED_INCOME / HOUSEHOLD_SIZE',
    'gross_by_hh_size':    '(EARNED_INCOME + UNEARNED_INCOME) / HOUSEHOLD_SIZE — on '
                           'reported data this is the manual\'s RAWGROSS / CERTHHSZ',
    'rawben_rel_max':      'recomputed benefit / maximum allotment for the unit size '
                           '(via the hidden benefit-recomputation chain and FederalTables)',
    'unc_rawben_rel_max':  'recomputed benefit BEFORE the minimum/maximum caps / maximum allotment',
    'shelter_expenses_by_hh_size': '(RENT + UTILITY_COSTS) / HOUSEHOLD_SIZE',
    'total_deductions_by_hh_size': '(dependent care + child support + recomputed shelter + medical '
                                   '+ earned-income deductions) / HOUSEHOLD_SIZE',
    'utilities':   'UTILITY_COSTS, unchanged',
    'utilities_sua': 'standard utility allowance tier, computed from the pasted data: '
                     '0 = no utility amount; 1 = positive but more than $200 below '
                     'the most common positive UTILITY_COSTS value among cases in the '
                     'same fiscal year; 2 = within $200 of or above that value (the '
                     'high-SUA cluster, which in states with household-size or '
                     'regional SUA schedules covers all of the high variants). Rules '
                     'use this tier instead of utility dollars so they keep meaning '
                     'the same thing when SUA levels reset each October. Paste whole '
                     'fiscal years rather than a handful of cases: the anchor is '
                     'recomputed from what you paste.',
    'over_threshold':     '1 when total_error_amount exceeds the review year\'s federal '
                          'QC tolerance (FederalTables error_threshold) — what the rules '
                          'aim to catch',
    'total_error_amount': 'ABS(ORIGINAL_BENEFIT_AMOUNT - CORRECTED_BENEFIT_AMOUNT), '
                          'rounded to whole dollars (QC manual: AMTERR)',
}
HELPER_NOTE = ('_c_* columns (hidden): the benefit-recomputation chain — fiscal year, '
               'gross income, earned-income deduction, standard deduction and maximum '
               'allotment looked up per year x size from the hidden FederalTables sheet, '
               'net income before and after shelter, the shelter deduction with its '
               'elderly/disabled uncapping, and the recomputed benefit.')


# data type + example per input column (feedback 2026-08-22)
RAW_TYPE = {
    'CASE_ID': ('text or number', 'A123456'),
    'REVIEW_FISCAL_YEAR': ('whole number', '2025'),
    'HOUSEHOLD_SIZE': ('whole number', '3'),
    'NUM_CHILDREN': ('whole number', '2'), 'NUM_ELDERLY': ('whole number', '0'),
    'NUM_DISABLED': ('whole number', '1'), 'NUM_ABAWD': ('whole number', '0'),
    'MARRIED_FLAG': ('0 or 1', '1'), 'EXPEDITED': ('0 or 1', '0'),
    'CATEGORICALLY_ELIGIBLE': ('0 or 1', '1'), 'HOMELESS_FLAG': ('0 or 1', '0'),
    'MONTHS_SINCE_CERT': ('whole number', '7'),
    'EARNED_INCOME': ('dollars, monthly', '1450'),
    'UNEARNED_INCOME': ('dollars, monthly', '0'),
    'MEDICAL_DEDUCTION': ('dollars, monthly', '0'),
    'DEPENDENT_CARE_DEDUCTION': ('dollars, monthly', '250'),
    'CHILD_SUPPORT_DEDUCTION': ('dollars, monthly', '0'),
    'CHILD_SUPPORT_EXPENSES': ('dollars, monthly', '0'),
    'HOMELESS_DEDUCTION': ('dollars, monthly', '0'),
    'RENT': ('dollars, monthly', '950'), 'UTILITY_COSTS': ('dollars, monthly', '459'),
    'ORIGINAL_BENEFIT_AMOUNT': ('dollars, monthly', '512'),
    'CORRECTED_BENEFIT_AMOUNT': ('dollars, monthly', '468'),
    'STATUS': ('code 1 / 2 / 3 / 4', '2'),
}
FEAT_TYPE = {
    'fiscal_year': ('whole number', '2025'), 'hh_size_raw': ('whole number', '3'),
    'hh_group': ('text: 1, 2-3 or 4+', '2-3'), 'HH_size_n': ('whole number', '3'),
    'bbce_state_i': ('0 or 1', '1'), 'children_i': ('0 or 1', '1'),
    'elderly_disabled_i': ('0 or 1', '1'), 'expedited_i': ('0 or 1', '0'),
    'homeless': ('0 or 1', '0'), 'married': ('0 or 1', '1'),
    'medical_deductions': ('dollars', '0'), 'months_since_cert_n': ('whole number', '7'),
    'percent_abawd': ('share 0 to 1', '0.33'), 'earned_by_hh_size': ('dollars', '483.33'),
    'unearned_by_hh_size': ('dollars', '0'), 'gross_by_hh_size': ('dollars', '483.33'),
    'rawben_rel_max': ('ratio', '0.62'), 'unc_rawben_rel_max': ('ratio', '0.62'),
    'shelter_expenses_by_hh_size': ('dollars', '469.67'),
    'total_deductions_by_hh_size': ('dollars', '312.50'),
    'utilities': ('dollars', '459'), 'utilities_sua': ('tier 0 / 1 / 2', '2'),
    'over_threshold': ('0 or 1', '1'), 'total_error_amount': ('dollars', '44'),
}

# the Step 3 tab's columns, one dictionary row each (feedback 2026-08-22:
# a per-column dictionary instead of a long box at the top of the chart;
# the headers on Step 3 hyperlink to these rows). Characterization
# definitions follow state_delivery_lists/README.md.
STEP3_COLS = [
    ('Rule', 'text', 'Rule 42',
     'The rule\'s id: its rank in the delivery list it came from. Stable across '
     'workbook versions, so you can refer to a rule by number.'),
    ('HH size', 'text: 1, 2-3 or 4+', '2-3',
     'The household-size group the rule applies to. Every rule is mined and '
     'applied within one group; a case outside it is never flagged by the rule.'),
    ('What the rule says', 'text', 'no children present; earned income per person over $526',
     'The rule\'s conditions in plain English; all conditions must hold for a '
     'case to be flagged. The exact machine form is in the last column.'),
    ('Precision', 'text: counts', '12 errors of 30 cases flagged',
     'On the pasted data: of the cases this rule ALONE flags, how many are '
     'payment errors. Shown as counts so small numbers read as small numbers.'),
    ('Recall', 'percent', '6.4%',
     'On the pasted data: the share of ALL error cases that this rule ALONE '
     'catches. Rules overlap, so these do not add up to the combined rows.'),
    ('$ Recall', 'percent', '8.1%',
     'On the pasted data: the share of ALL error dollars that this rule ALONE '
     'catches.'),
    ('Flagged', 'whole number', '30',
     'Cases on the pasted data this rule flags.'),
    ('Errors', 'whole number', '12',
     'Of those flagged cases, how many are payment errors.'),
    ('Error $ caught', 'dollars', '$4,120',
     'The error dollars on the flagged cases.'),
    ('Workload %', 'percent', '2.5%',
     'Flagged cases as a share of the rule\'s own household-size group on the '
     'pasted data (in the orange combined rows: a share of ALL cases).'),
    ('Expected error $ by case', 'dollars', '$137',
     'Error $ caught divided by cases flagged: the average error dollars per '
     'case you would review under this rule.'),
    ('Include?', 'TRUE / FALSE', 'TRUE',
     'Set to FALSE to drop the rule from the combined results and from the '
     'export. The yellow cell is the only thing to edit on this tab.'),
    ('Rule source pool', 'text', 'national',
     'Where the rule was mined: "national" (all 49 states\' public QC data) or '
     '"state" (this state\'s own public QC data).'),
    ('Train precision (natl.)', 'percent', '77.6%',
     'Precision when the rule was mined, on the national training data '
     '(FY2022-2024). Fixed context; does not recompute from pasted data.'),
    ('Train precision lower bound', 'percent', '61.4%',
     'A conservative (99% lower confidence bound) version of train precision; '
     'the statistic the rules were ranked on. Fixed context.'),
    ('Train error $ per flagged case', 'dollars', '$210',
     'Average error dollars per flagged case on the national training data. '
     'Fixed context.'),
    ('Natl. error cases behind rule', 'whole number', '38',
     'How many error cases nationwide (FY2022-2024) the rule matched; the '
     'sample behind the four share columns that follow. NOT this state\'s count.'),
    ('Error elements caught (to 75%)', 'text: element and share', 'shelter deduction 0.49; medical deduction 0.19',
     'Of the rule\'s national error cases, which QC error ELEMENTS (the part of '
     'the case that was wrong) they involved, with each element\'s share, '
     'listed until 75% of the cases are covered. "shelter deduction 0.49" '
     'means 49% of the rule\'s error cases involved the shelter deduction.'),
    ('Error natures caught (to 75%)', 'text: nature and share', 'wrong amount, known item 0.42',
     'Same, for the QC error NATURE (how the error happened): for example '
     'the item was known but the amount was wrong, or an include/exclude '
     'decision was wrong. Shares listed until 75% of the cases are covered.'),
    ('Share overissuance', 'percent', '100%',
     'Of the rule\'s national error cases, the share that were overissuances '
     '(the household was paid more than the correct amount).'),
    ('Share agency-caused', 'percent', '60%',
     'Of the rule\'s national error cases, the share QC coded as agency-caused '
     '(client-caused is roughly the rest).'),
    ('Share discovered in case file', 'percent', '34%',
     'Of the rule\'s national error cases, the share the QC reviewer found '
     'from the case record itself rather than from client contact or a fresh '
     'data match: a rough guide to which errors a desk review can find.'),
    ('Share at certification', 'percent', '79%',
     'Of the rule\'s national error cases, the share that arose at the '
     'agency\'s certification or recertification action rather than later.'),
    ('Exact expression', 'text', 'children_i <= 0 & earned_by_hh_size > 526',
     'The rule\'s conditions as machine logic, using the constructed-variable '
     'names defined above. This is what the Export tab carries.'),
]


def data_dictionary(wb, hdr):
    """A visible tab documenting every Data column (the input fields a state
    supplies and the constructed model variables), then every column of the
    Step 3 rules tab. Returns {Step 3 header text: dictionary row} so the
    rules tab's headers can hyperlink to their definitions."""
    ws = wb.create_sheet(DICT_SHEET, wb.sheetnames.index(DATA_SHEET) + 1)
    ws.sheet_view.showGridLines = False
    blue = PatternFill('solid', fgColor='2F5496')
    gray = PatternFill('solid', fgColor='F2F2F2')
    wrap = Alignment(wrap_text=True, vertical='top')
    for cl, w in (('A', 30), ('B', 20), ('C', 26), ('D', 100)):
        ws.column_dimensions[cl].width = w
    ws.merge_cells('A1:D1')
    c = ws['A1']
    c.value = (f'Step 1: map your data to this dictionary — every column on the '
               f'"{DATA_SHEET}" tab, then every column on the "{BLENDED_SHEET}" tab')
    c.fill = blue; c.font = Font(bold=True, size=14, color='FFFFFF')
    ws.row_dimensions[1].height = 28

    def header(r, txt):
        ws.merge_cells(f'A{r}:D{r}')
        c = ws.cell(row=r, column=1, value=txt)
        c.fill = blue; c.font = Font(bold=True, color='FFFFFF'); c.alignment = wrap
        ws.row_dimensions[r].height = max(18, 15 * -(-len(txt) // 150))
        return r + 1

    def cols(r):
        for ci, t in enumerate(('column', 'data type', 'example', 'definition'), 1):
            c = ws.cell(row=r, column=ci, value=t)
            c.fill = gray; c.font = Font(bold=True, size=10)
        return r + 1

    def entry(r, name, typ, example, definition):
        ws.cell(row=r, column=1, value=name).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=typ).font = Font(size=10)
        ws.cell(row=r, column=3, value=example).font = Font(size=10)
        c = ws.cell(row=r, column=4, value=definition)
        c.font = Font(size=10); c.alignment = wrap
        ws.row_dimensions[r].height = max(15, 13 * -(-len(definition) // 115))
        return r + 1

    r = header(3, 'Input fields (yellow block) — what a state supplies: its own '
                  'as-reported values, mapped onto these columns. Each '
                  'definition cross-references the SNAP QC technical '
                  'documentation; this public-data copy carries the research '
                  'frame\'s reconstructed pre-QC-review values. A blank cell '
                  'means MISSING, not zero: a case missing an input a rule '
                  'needs is not flagged by that rule, and a column you do not '
                  'collect can be left entirely empty (rules that use it stop '
                  'flagging). Enter zeros as zeros.')
    r = cols(r)
    for name in RAW_COLS:
        typ, ex = RAW_TYPE.get(name, ('value', ''))
        r = entry(r, name, typ, ex, RAW_DESC.get(name, ''))
    r = header(r + 1, 'Constructed variables (gray block) — computed by formula from the '
                      'input fields; the rules reference these. Do not paste over them.')
    r = cols(r)
    for name in hdr:
        if name.startswith('_') or name not in FEAT_DESC:
            continue
        typ, ex = FEAT_TYPE.get(name, ('formula', ''))
        r = entry(r, name, typ, ex, FEAT_DESC[name])
    r += 1
    ws.merge_cells(f'A{r}:D{r}')
    c = ws.cell(row=r, column=1, value=HELPER_NOTE)
    c.font = Font(size=9, color='808080')
    r += 2
    r = header(r, f'Columns on the "{BLENDED_SHEET}" tab — one row per column. The '
                  'first twelve recompute from the pasted data; the columns from '
                  '"Rule source pool" onward describe each rule as mined on NATIONAL '
                  'data and do not change when you paste. Click a column header on '
                  f'the "{BLENDED_SHEET}" tab to jump to its row here.')
    r = cols(r)
    step3_rows = {}
    for name, typ, ex, definition in STEP3_COLS:
        step3_rows[name] = r
        r = entry(r, name, typ, ex, definition)
    ws.freeze_panes = 'A2'
    return step3_rows


# ── Step 5.1 / 5.2: screen new cases, list the flagged case x rule pairs ─────
SCREEN_TABLE = 'ScreenData'
# same paste contract minus the outcome columns: new cases have no review
# outcome yet, and the corrected benefit / status only exist after a review
SCREEN_COLS = [c for c in RAW_COLS
               if c not in ('CORRECTED_BENEFIT_AMOUNT', 'STATUS')]
FLAG_CAP = 5000               # case x rule pairs displayed on Step 5.2


def screening_tabs(wb, R, dat_hdr):
    """Step 5.1: a second input table with the Step 2 contract minus the
    outcome columns, the same feature formulas, and one hidden 0/1 column per
    rule (weighted by that rule's Include? flag). Step 5.2: one row per
    flagged case x rule pair, enumerated with a binary-search MATCH on the
    running pair count plus AGGREGATE for the j-th matching rule (pattern
    COM-verified 2026-08-18). Both tables auto-extend on paste."""
    from openpyxl.utils import get_column_letter
    bq = qref(BLENDED_SHEET)
    rules = read_delivery_tab(wb[BLENDED_SHEET])
    sel = selection_refs(wb, BLENDED_SHEET)
    nr = len(rules)
    assert nr and len(sel) == nr, 'rules tab / RuleFlags selection mismatch'

    helpers, feats = feature_formulas(R, SCREEN_TABLE)
    feat_cols = [h for h in dat_hdr
                 if h in feats and h not in ('over_threshold', 'total_error_amount')]
    hdr = (SCREEN_COLS + feat_cols + [h for h, _ in helpers]
           + [f'_r{j + 1}' for j in range(nr)] + ['_hits', '_cum'])
    ws = wb.create_sheet(SCREEN_SHEET)
    yellow = PatternFill('solid', fgColor='FFFF99')
    gray = PatternFill('solid', fgColor='D9D9D9')
    for ci, name in enumerate(hdr, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = Font(bold=True)
        if name in SCREEN_COLS:
            c.fill = yellow
        elif not name.startswith('_'):
            c.fill = gray
    hlet = lambda name: get_column_letter(hdr.index(name) + 1)
    # the single seeded row: its formulas define the table's calculated
    # columns, which Excel fills down when a state pastes cases
    for name, f in list(helpers) + [(n, feats[n]) for n in feat_cols]:
        ws.cell(row=2, column=hdr.index(name) + 1, value=f)
    for j, (conds, hh) in enumerate(rules):
        f = (f'={sel[11 + j]}*{rule_term(conds, hh, SCREEN_TABLE)}'
             if conds else '=0')
        ws.cell(row=2, column=hdr.index(f'_r{j + 1}') + 1, value=f)
    ws.cell(row=2, column=hdr.index('_hits') + 1,
            value=f'=SUM({SCREEN_TABLE}[[#This Row],[_r1]:[_r{nr}]])')
    ws.cell(row=2, column=hdr.index('_cum') + 1,
            value=f'=${hlet("_hits")}2+N(${hlet("_cum")}1)')
    for ci, name in enumerate(hdr, 1):
        if name in feats:
            ws.cell(row=2, column=ci).fill = gray
        if name.startswith('_'):
            ws.column_dimensions[get_column_letter(ci)].hidden = True
    make_table(ws, SCREEN_TABLE, f'A1:{get_column_letter(len(hdr))}2')
    ws.freeze_panes = 'B2'
    ws.cell(row=1, column=1).comment = Comment(
        'Optional: paste NEW cases to screen against the selected rules. Same '
        f'yellow input columns as the "{DATA_SHEET}" tab minus the outcome '
        'columns (new cases have no review outcome yet); definitions on the '
        f'"{DICT_SHEET}" tab. The gray columns are formulas — do not paste '
        f'over them. Flagged cases appear on the "{FLAGGED_SHEET}" tab. NB '
        'bbce_state_i recomputes from the pasted rows themselves, so paste '
        'full years rather than a handful of cases.', 'snap_dashboard')

    # ── Step 5.2 ─────────────────────────────────────────────────────────────
    blue = PatternFill('solid', fgColor='2F5496')
    grayF = PatternFill('solid', fgColor='F2F2F2')
    wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    wsf = wb.create_sheet(FLAGGED_SHEET)
    wsf.sheet_view.showGridLines = False
    for cl, w in {'A': 16, 'B': 12, 'C': 14, 'D': 9, 'E': 70}.items():
        wsf.column_dimensions[cl].width = w
    for cl in ('G', 'H', 'I'):
        wsf.column_dimensions[cl].hidden = True
    wsf.merge_cells('A1:E1')
    c = wsf['A1']; c.value = 'Flagged New Cases'
    c.fill = blue; c.font = Font(bold=True, size=16, color='FFFFFF')
    wsf.row_dimensions[1].height = 30
    wsf.merge_cells('A2:E2')
    c = wsf['A2']
    c.value = (f'One row per case x rule: every case pasted into the "{SCREEN_SHEET}" '
               'tab that a rule currently set to TRUE flags, with the rule that '
               f'flagged it. A case flagged by several rules appears once per rule. '
               f'Showing the first {FLAG_CAP:,} pairs; cell B3 holds the total.')
    c.fill = grayF; c.font = Font(size=12); c.alignment = wrap
    wsf.row_dimensions[2].height = 46
    wsf['A3'] = 'Flagged case x rule pairs:'
    wsf['A3'].font = Font(bold=True)
    wsf['B3'] = f'=SUM({SCREEN_TABLE}[_hits])'
    wsf['B3'].font = Font(bold=True)
    wsf['B3'].number_format = '#,##0'
    for ci, txt in enumerate(['Case ID', 'Household size', 'Benefit amount',
                              'Rule', 'What the rule says'], 1):
        c = wsf.cell(row=4, column=ci, value=txt)
        c.fill = grayF; c.font = Font(bold=True, size=10)
    RBLK = f'{SCREEN_TABLE}[[_r1]:[_r{nr}]]'
    for k in range(1, FLAG_CAP + 1):
        r = 4 + k
        g = f'ROW()-4'
        wsf.cell(row=r, column=7, value=(
            f'=IF({g}>$B$3,"",IFERROR(MATCH({g}-1,{SCREEN_TABLE}[_cum],1),0)+1)'))
        wsf.cell(row=r, column=8, value=(
            f'=IF($G{r}="","",{g}-(INDEX({SCREEN_TABLE}[_cum],$G{r})'
            f'-INDEX({SCREEN_TABLE}[_hits],$G{r})))'))
        wsf.cell(row=r, column=9, value=(
            f'=IF($G{r}="","",_xlfn.AGGREGATE(15,6,'
            f'(COLUMN({RBLK})-COLUMN({SCREEN_TABLE}[_r1])+1)'
            f'/(INDEX({RBLK},$G{r},0)>0),$H{r}))'))
        wsf.cell(row=r, column=1, value=(
            f'=IF($I{r}="","",INDEX({SCREEN_TABLE}[CASE_ID],$G{r}))'))
        wsf.cell(row=r, column=2, value=(
            f'=IF($I{r}="","",INDEX({SCREEN_TABLE}[HOUSEHOLD_SIZE],$G{r}))'))
        c = wsf.cell(row=r, column=3, value=(
            f'=IF($I{r}="","",INDEX({SCREEN_TABLE}[ORIGINAL_BENEFIT_AMOUNT],$G{r}))'))
        c.number_format = '$#,##0'
        wsf.cell(row=r, column=4, value=(
            f'=IF($I{r}="","",INDEX({bq}!$A$11:$A${10 + nr},$I{r}))'))
        c = wsf.cell(row=r, column=5, value=(
            f'=IF($I{r}="","",INDEX({bq}!$C$11:$C${10 + nr},$I{r}))'))
        c.font = Font(size=10)
    wsf.freeze_panes = 'A5'
    return nr


def share_tab(wb, state_name):
    """Step 6: aggregate per-rule performance on the pasted Step 2 data, in a
    block a state can copy back to us — every rule alone (independent of
    Include?), with the denominators that make results poolable across
    states, plus the ineligible-household catches the public files can never
    show (STATUS = 4)."""
    bq = qref(BLENDED_SHEET)
    rules_ws = wb[BLENDED_SHEET]
    rules = read_delivery_tab(rules_ws)
    nr = len(rules)
    blue = PatternFill('solid', fgColor='2F5496')
    grayF = PatternFill('solid', fgColor='F2F2F2')
    yellow = PatternFill('solid', fgColor='FFFF99')
    orange = PatternFill('solid', fgColor='F8CBAD')
    wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws = wb.create_sheet(SHARE_SHEET)
    ws.sheet_view.showGridLines = False
    for cl, w in {'A': 30, 'B': 9, 'C': 60, 'D': 11, 'E': 12, 'F': 11,
                  'G': 11, 'H': 18}.items():
        ws.column_dimensions[cl].width = w
    ws.merge_cells('A1:H1')
    c = ws['A1']; c.value = f'Share aggregate results back — {state_name}'
    c.fill = blue; c.font = Font(bold=True, size=16, color='FFFFFF')
    ws.row_dimensions[1].height = 30
    ws.merge_cells('A2:H2')
    c = ws['A2']
    # wording agreed 2026-08-19
    c.value = ('Optional: this tab displays aggregate performance of every rule on the '
               f'data pasted into the "{DATA_SHEET}" tab. It would help us improve the '
               'models (for your state and others) if you can share back the aggregate '
               'performance by rule - that will tell us what kinds of errors can be '
               'reliably found out of sample or what kinds of rules might be more '
               'robust out of sample. Fill in the yellow cells so we know what the '
               'numbers cover. Ineligible cases are not included in the public QC '
               'sample, which is why we want a separate count of them and which rule '
               'might capture them. To send the results back to us, please copy / '
               'paste the whole sheet into a new excel workbook (paste as values) and '
               'send it back to us.')
    c.fill = grayF; c.font = Font(size=12); c.alignment = wrap
    ws.row_dimensions[2].height = 80
    meta = [('State', state_name, False),
            ('Fiscal years pasted into Step 2 (e.g., 2025-2026)', '', True),
            ('Data pasted (QC / QA / pre-auth / other)', '', True)]
    for i, (label, val, editable) in enumerate(meta):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=val)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        if editable:
            c.fill = yellow
    dv = DataValidation(type='list',
                        formula1='"QC review data,QA review data,pre-auth data,other"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add('B6')
    dens = [('Total cases', f'=COUNTA({TABLE}[CASE_ID])', '#,##0'),
            ('Error cases', f'=COUNTIF({TABLE}[over_threshold],1)', '#,##0'),
            ('Error $', f'=SUMIFS({TABLE}[total_error_amount],'
                        f'{TABLE}[over_threshold],1)', '$#,##0'),
            ('Ineligible-household cases (STATUS = 4)',
             f'=COUNTIF({TABLE}[STATUS],4)', '#,##0')]
    for i, (label, f, fmt) in enumerate(dens):
        r = 8 + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=f)
        c.fill = orange; c.number_format = fmt
    hdrs = ['Rule', 'HH size', 'What the rule says', 'Flagged', 'Errors caught',
            'Precision', '$ Recall', 'Ineligible households flagged (STATUS=4)']
    HR = 13
    for ci, txt in enumerate(hdrs, 1):
        c = ws.cell(row=HR, column=ci, value=txt)
        c.fill = grayF; c.font = Font(bold=True, size=10)
    tot_ed = f'SUMIFS({TABLE}[total_error_amount],{TABLE}[over_threshold],1)'
    for j, (conds, hh) in enumerate(rules):
        r = HR + 1 + j
        src = 11 + j
        # rule id / stratum / plain text are static on the rules tab — copy
        # the values so this block pastes cleanly as text
        ws.cell(row=r, column=1, value=rules_ws.cell(row=src, column=1).value)
        ws.cell(row=r, column=2, value=rules_ws.cell(row=src, column=2).value)
        c = ws.cell(row=r, column=3, value=rules_ws.cell(row=src, column=3).value)
        c.font = Font(size=10); c.alignment = wrap
        if conds is None:
            continue
        n = countifs(conds, hh)
        e = countifs(conds, hh, extra=f'{TABLE}[over_threshold],1')
        d = countifs(conds, hh, extra=f'{TABLE}[over_threshold],1',
                     col='total_error_amount', fn='SUMIFS')
        i4 = countifs(conds, hh, extra=f'{TABLE}[STATUS],4')
        for ci, (f, fmt) in enumerate([
                (f'={n}', '#,##0'), (f'={e}', '#,##0'),
                (f'=IFERROR($E{r}/$D{r},0)', '0.0%'),
                (f'=IFERROR({d}/{tot_ed},0)', '0.0%'),
                (f'={i4}', '#,##0')], 4):
            c = ws.cell(row=r, column=ci, value=f)
            c.number_format = fmt
    ws.freeze_panes = f'A{HR + 1}'
    return nr


# ── feature inputs: which raw columns each feature needs ─────────────────────
# Blank means MISSING, never zero: a feature only computes for a case when
# every input it needs holds a number on that row; otherwise the feature goes
# blank and no rule condition can match it (mirroring the R pipeline, where a
# missing value never matches). Deleting a whole input column therefore
# disables every rule that needs it. The demo block carries the research
# frame's reconstructed values with its zero-imputation already applied, so
# the shipped figures are unchanged. The QC outcome features are guarded on
# the benefit pair (2026-08-18): a case missing either benefit amount gets a
# blank error amount and a blank over-threshold flag, so it counts as neither
# an error nor a clean case.
BEN_INPUTS = ['REVIEW_FISCAL_YEAR', 'HOUSEHOLD_SIZE', 'EARNED_INCOME',
              'UNEARNED_INCOME', 'MEDICAL_DEDUCTION', 'DEPENDENT_CARE_DEDUCTION',
              'CHILD_SUPPORT_DEDUCTION', 'HOMELESS_DEDUCTION', 'RENT',
              'UTILITY_COSTS', 'NUM_ELDERLY', 'NUM_DISABLED']
FEATURE_INPUTS = {
    'fiscal_year': ['REVIEW_FISCAL_YEAR'],
    'hh_size_raw': ['HOUSEHOLD_SIZE'],
    'hh_group': ['HOUSEHOLD_SIZE'],
    'HH_size_n': ['HOUSEHOLD_SIZE'],
    'bbce_state_i': ['REVIEW_FISCAL_YEAR', 'CATEGORICALLY_ELIGIBLE'],
    'children_i': ['NUM_CHILDREN'],
    'elderly_disabled_i': ['NUM_ELDERLY', 'NUM_DISABLED'],
    'expedited_i': ['EXPEDITED'],
    'homeless': ['HOMELESS_FLAG'],
    'married': ['MARRIED_FLAG'],
    'medical_deductions': ['MEDICAL_DEDUCTION'],
    'months_since_cert_n': ['MONTHS_SINCE_CERT'],
    'percent_abawd': ['NUM_ABAWD', 'HOUSEHOLD_SIZE'],
    'earned_by_hh_size': ['EARNED_INCOME', 'HOUSEHOLD_SIZE'],
    'unearned_by_hh_size': ['UNEARNED_INCOME', 'HOUSEHOLD_SIZE'],
    'gross_by_hh_size': ['EARNED_INCOME', 'UNEARNED_INCOME', 'HOUSEHOLD_SIZE'],
    'rawben_rel_max': BEN_INPUTS,
    'unc_rawben_rel_max': BEN_INPUTS,
    'shelter_expenses_by_hh_size': ['RENT', 'UTILITY_COSTS', 'HOUSEHOLD_SIZE'],
    'total_deductions_by_hh_size': BEN_INPUTS,
    'utilities': ['UTILITY_COSTS'],
    'utilities_sua': ['UTILITY_COSTS', 'REVIEW_FISCAL_YEAR'],
    'over_threshold': ['ORIGINAL_BENEFIT_AMOUNT', 'CORRECTED_BENEFIT_AMOUNT',
                       'REVIEW_FISCAL_YEAR'],
    'total_error_amount': ['ORIGINAL_BENEFIT_AMOUNT', 'CORRECTED_BENEFIT_AMOUNT'],
}


def T(col, table=TABLE):
    return f'{table}[[#This Row],[{col}]]'


def feature_formulas(R, table=TABLE):
    """name -> formula for every feature the current delivery vocabulary uses,
    mirroring 1_data_munging_..._for_using_public_qc_data.R. Helpers are
    prefixed '_c_'; order matters (left to right). Each feature carries a
    per-row guard (FEATURE_INPUTS): a case missing any input the feature
    needs gets a blank, never a zero, so no rule condition can match it.
    `table` parameterizes the structured references so the Step 5.1
    screening table reuses the exact same formulas (2026-08-18)."""
    T = lambda col: f'{table}[[#This Row],[{col}]]'   # noqa: E731 — shadow on purpose
    sz = f'MIN(MAX({T("HOUSEHOLD_SIZE")},1),20)'
    hh = f'MAX({T("HOUSEHOLD_SIZE")},1)'
    helpers = [
        ('_c_fy',      f'={T("REVIEW_FISCAL_YEAR")}'),
        ('_c_eld',     f'=IF({T("NUM_ELDERLY")}+{T("NUM_DISABLED")}>0,1,0)'),
        ('_c_gross',   f'={T("EARNED_INCOME")}+{T("UNEARNED_INCOME")}'),
        # _xlfn. prefix: post-2007 functions written straight into the XML
        # need it, or Excel renders #NAME?
        ('_c_ernded',  f'=_xlfn.FLOOR.MATH({T("EARNED_INCOME")}*0.2)'),
        ('_c_stdded',  f'=INDEX({R["SDBLK"]},MATCH({T("_c_fy")},{R["SDYRS"]},1),{sz})'),
        ('_c_benmax',  f'=INDEX({R["MABLK"]},MATCH({T("_c_fy")},{R["MAYRS"]},1),{sz})'),
        ('_c_netbs',   f'={T("_c_gross")}-({T("_c_ernded")}+{T("DEPENDENT_CARE_DEDUCTION")}'
                       f'+{T("MEDICAL_DEDUCTION")}+{T("CHILD_SUPPORT_DEDUCTION")}+{T("_c_stdded")})'),
        ('_c_maxsh',   f'=IF({T("_c_eld")}=1,1000000000,'
                       f'INDEX({R["MAXSH"]},MATCH({T("_c_fy")},{R["YEARS"]},1)))'),
        # NB: the munging script does NOT floor the shelter deduction; only the
        # net incomes and the benefit are floored (calculate_raw_benefits)
        ('_c_sltded',  f'=MIN(MAX({T("RENT")}+{T("UTILITY_COSTS")}'
                       f'-MAX({T("_c_netbs")}*0.5,0),0),{T("_c_maxsh")})'),
        ('_c_netan',   f'=_xlfn.FLOOR.MATH({T("_c_netbs")}-({T("_c_sltded")}'
                       f'+{T("HOMELESS_DEDUCTION")}))'),
        ('_c_benunc',  f'=_xlfn.FLOOR.MATH({T("_c_benmax")}-0.3*{T("_c_netan")})'),
        ('_c_benrec',  f'=MIN(MAX({T("_c_benunc")},IF({T("HOUSEHOLD_SIZE")}<3,'
                       f'INDEX({R["MINAL"]},MATCH({T("_c_fy")},{R["YEARS"]},1)),0)),'
                       f'{T("_c_benmax")})'),
    ]
    # a row with blank inputs must never surface an ERROR cell: the year
    # lookups return #N/A on a blank fiscal year and cascade down the chain
    # (seen on Step 5.1's empty seeded row, 2026-08-18). IFERROR blanks the
    # helper instead; with complete inputs it is a no-op (the validation
    # gate re-verifies), and the guarded features ignore blank helpers.
    helpers = [(n, f'=IFERROR({f[1:]},"")') for n, f in helpers]
    feats = {
        'fiscal_year': f'={T("REVIEW_FISCAL_YEAR")}',
        'hh_size_raw': f'={T("HOUSEHOLD_SIZE")}',
        'hh_group':    f'=IF({T("HOUSEHOLD_SIZE")}>=4,"4+",'
                       f'IF({T("HOUSEHOLD_SIZE")}>=2,"2-3","1"))',
        'HH_size_n':   f'={T("HOUSEHOLD_SIZE")}',
        'bbce_state_i':
            # state-year regime flag: share of the year's cases categorically
            # eligible reaching 0.5 (munging script, 2026-08-13 decision)
            f'=IF(COUNTIFS({table}[REVIEW_FISCAL_YEAR],{T("REVIEW_FISCAL_YEAR")},'
            f'{table}[CATEGORICALLY_ELIGIBLE],1)'
            f'/COUNTIFS({table}[REVIEW_FISCAL_YEAR],{T("REVIEW_FISCAL_YEAR")})>=0.5,1,0)',
        'children_i':  f'=IF({T("NUM_CHILDREN")}>0,1,0)',
        'elderly_disabled_i': f'={T("_c_eld")}',
        'expedited_i': f'={T("EXPEDITED")}',
        'homeless':    f'={T("HOMELESS_FLAG")}',
        'married':     f'={T("MARRIED_FLAG")}',
        'medical_deductions':  f'={T("MEDICAL_DEDUCTION")}',
        'months_since_cert_n': f'={T("MONTHS_SINCE_CERT")}',
        'percent_abawd': f'={T("NUM_ABAWD")}/{hh}',
        'earned_by_hh_size':   f'={T("EARNED_INCOME")}/{hh}',
        'unearned_by_hh_size': f'={T("UNEARNED_INCOME")}/{hh}',
        'gross_by_hh_size':    f'=({T("EARNED_INCOME")}+{T("UNEARNED_INCOME")})/{hh}',
        'rawben_rel_max':      f'={T("_c_benrec")}/{T("_c_benmax")}',
        'shelter_expenses_by_hh_size': f'=({T("RENT")}+{T("UTILITY_COSTS")})/{hh}',
        'total_deductions_by_hh_size':
            f'=({T("DEPENDENT_CARE_DEDUCTION")}+{T("CHILD_SUPPORT_DEDUCTION")}'
            f'+{T("_c_sltded")}+{T("MEDICAL_DEDUCTION")}+{T("_c_ernded")})/{hh}',
        'unc_rawben_rel_max': f'={T("_c_benunc")}/{T("_c_benmax")}',
        'utilities':   f'={T("UTILITY_COSTS")}',
        # SUA tier (vocabulary variant staged 2026-08-22; design + result in
        # methods/v250_benchmark_2024_utilrel/): 0 = no utility amount,
        # 1 = positive below (state-year mode - 200), 2 = at/above that
        # (the HIGH-SUA cluster). The anchor is the MODE of positive
        # UTILITY_COSTS within the row's fiscal year, computed from the
        # pasted data itself. The mode is NOT computed inside this row
        # formula: an IF(...) array inside MODE.SNGL in a non-array table
        # cell collapses to a single cell under implicit intersection and
        # IFERROR hides the failure (every row read tier 0 in the first
        # build, 2026-08-22, while the pandas mirror passed). It lives in
        # the hidden per-year mode block on FederalTables (R["SUAMODE"] /
        # R["SUAYRS"], array-entered there), and each row does a plain
        # INDEX/MATCH; a year absent from the block yields mode 0 -> tier 2
        # for every positive row, handled by the IFERROR->0 below only for
        # real lookup errors. Tie rule: MODE.SNGL returns the first-
        # occurring tied value (the build frame's mode_pos the smallest);
        # the frame has no tied state-year cells, so the demo matches.
        'utilities_sua':
            f'=IF({T("UTILITY_COSTS")}<=0,0,IFERROR(IF({T("UTILITY_COSTS")}'
            f'<INDEX({R["SUAMODE"]},MATCH({T("REVIEW_FISCAL_YEAR")},{R["SUAYRS"]},0))'
            f'-200,1,2),0))',
        # the QC outcome, recomputed from the benefit pair exactly as the
        # munging script defines it: error amount = |RAWBEN - FSBEN| rounded,
        # error flag = amount STRICTLY OVER the year's federal QC tolerance
        # (FederalTables error_threshold; munging script lines 165/198)
        'over_threshold':
            f'=IF(ROUND(ABS({T("ORIGINAL_BENEFIT_AMOUNT")}'
            f'-{T("CORRECTED_BENEFIT_AMOUNT")}),0)'
            f'>INDEX({R["ERRTHR"]},MATCH({T("REVIEW_FISCAL_YEAR")},{R["YEARS"]},1)),1,0)',
        'total_error_amount':
            f'=ROUND(ABS({T("ORIGINAL_BENEFIT_AMOUNT")}'
            f'-{T("CORRECTED_BENEFIT_AMOUNT")}),0)',
    }
    for name, f in list(feats.items()):
        cols = list(dict.fromkeys(FEATURE_INPUTS.get(name, [])))
        if cols:
            # IFERROR inside the guard (2026-08-18 review): inputs that are
            # numeric but unusable (e.g. a "25" fiscal-year typo) pass the
            # COUNT guard yet blank the helper chain, and an error-valued
            # feature cell would poison every per-case rule term it touches.
            # Degrade to blank instead — same meaning as a missing input.
            # A real formula bug still cannot hide: the validation gate
            # compares values against the frame on every build.
            guard = f'COUNT({",".join(T(c) for c in cols)})={len(cols)}'
            feats[name] = f'=IF({guard},IFERROR({f[1:]},""),"")'
    return helpers, feats


# ── validation: mirror every formula in pandas, score against the frame ──────
def _excel_floor(x, sig=1.0):
    return np.floor(np.asarray(x, float) / sig) * sig


def mirror_features(raw, ftabs):
    """Compute what the Excel formulas will produce, from the raw block."""
    g = lambda c: raw[c].fillna(0).astype(float).values     # Excel blank -> 0
    yd, sd, ma = ftabs
    fy = g('REVIEW_FISCAL_YEAR').astype(int)
    sz20 = np.clip(g('HOUSEHOLD_SIZE'), 1, 20).astype(int)
    hh = np.maximum(g('HOUSEHOLD_SIZE'), 1)
    lk = lambda tbl, col: np.array([tbl.loc[tbl.year <= y, col].iloc[-1] for y in fy])
    stdded = np.array([sd.loc[sd.year <= y, str(s)].iloc[-1] for y, s in zip(fy, sz20)])
    benmax = np.array([ma.loc[ma.year <= y, str(s)].iloc[-1] for y, s in zip(fy, sz20)])
    eld = ((g('NUM_ELDERLY') + g('NUM_DISABLED')) > 0).astype(int)
    ernded = _excel_floor(g('EARNED_INCOME') * 0.2)
    netbs = (g('EARNED_INCOME') + g('UNEARNED_INCOME')
             - (ernded + g('DEPENDENT_CARE_DEDUCTION') + g('MEDICAL_DEDUCTION')
                + g('CHILD_SUPPORT_DEDUCTION') + stdded))
    maxsh = np.where(eld == 1, 1e9, lk(yd, 'max_shelter_deduction'))
    sltded = np.minimum(np.maximum(g('RENT') + g('UTILITY_COSTS')
                                   - np.maximum(netbs * 0.5, 0), 0), maxsh)
    netan = _excel_floor(netbs - (sltded + g('HOMELESS_DEDUCTION')))
    benunc = _excel_floor(benmax - 0.3 * netan)
    minal = lk(yd, 'min_allotment')
    benrec = np.minimum(np.maximum(
        benunc, np.where(g('HOUSEHOLD_SIZE') < 3, minal, 0)), benmax)
    fyshare = pd.Series(g('CATEGORICALLY_ELIGIBLE')).groupby(fy).transform('mean')
    out = {
        'fiscal_year': fy, 'hh_size_raw': g('HOUSEHOLD_SIZE'),
        'HH_size_n': g('HOUSEHOLD_SIZE'),
        'hh_group': np.where(g('HOUSEHOLD_SIZE') >= 4, '4+',
                             np.where(g('HOUSEHOLD_SIZE') >= 2, '2-3', '1')),
        'bbce_state_i': (fyshare.values >= 0.5).astype(int),
        'children_i': (g('NUM_CHILDREN') > 0).astype(int),
        'elderly_disabled_i': eld,
        'expedited_i': g('EXPEDITED'),
        'homeless': g('HOMELESS_FLAG'),
        'married': g('MARRIED_FLAG'),
        'medical_deductions': g('MEDICAL_DEDUCTION'),
        'months_since_cert_n': g('MONTHS_SINCE_CERT'),
        'percent_abawd': g('NUM_ABAWD') / hh,
        'earned_by_hh_size': g('EARNED_INCOME') / hh,
        'unearned_by_hh_size': g('UNEARNED_INCOME') / hh,
        'gross_by_hh_size': (g('EARNED_INCOME') + g('UNEARNED_INCOME')) / hh,
        'rawben_rel_max': benrec / benmax,
        'shelter_expenses_by_hh_size': (g('RENT') + g('UTILITY_COSTS')) / hh,
        'total_deductions_by_hh_size': (g('DEPENDENT_CARE_DEDUCTION')
                                        + g('CHILD_SUPPORT_DEDUCTION') + sltded
                                        + g('MEDICAL_DEDUCTION') + ernded) / hh,
        'unc_rawben_rel_max': benunc / benmax,
        'utilities': g('UTILITY_COSTS'),
    }
    # SUA tier mirror of the EXCEL formula (mode of positive rounded
    # utilities within fiscal year, smallest tied value, $200 band). The
    # validation gate compares this against the frame's canonical
    # utilities_sua column (features.R add_sua_tier, promoted 2026-08-22),
    # so a drift between the workbook formula and the canonical definition
    # fails the build. Keep this in step with features.R.
    util = g('UTILITY_COSTS')
    tier = np.zeros(len(util), dtype=int)
    for y in np.unique(fy):
        sel = fy == y
        pos = np.round(util[sel & (util > 0)])
        if len(pos):
            vals, cnt = np.unique(pos, return_counts=True)
            mode = vals[cnt == cnt.max()].min()
            t = np.where(util[sel] <= 0, 0, np.where(util[sel] < mode - 200, 1, 2))
            tier[sel] = t
    out['utilities_sua'] = tier
    # QC outcome recomputed from the benefit pair, mirroring the Excel
    # formulas (and the munging script's own definition)
    errdiff = np.round(np.abs(g('ORIGINAL_BENEFIT_AMOUNT')
                              - g('CORRECTED_BENEFIT_AMOUNT')))
    out['total_error_amount'] = errdiff
    out['over_threshold'] = (errdiff > lk(yd, 'error_threshold')).astype(int)
    return pd.DataFrame(out)


def validate(raw, frame, elem_free, feat_names):
    ad = os.path.join(REPO, 'additional_data')
    yd = pd.read_csv(os.path.join(ad, 'year_data.csv')).sort_values('year')
    sd = pd.read_csv(os.path.join(ad, 'standard_deductions.csv')).sort_values('year')
    ma = pd.read_csv(os.path.join(ad, 'max_allotments.csv')).sort_values('year')
    for t in (sd, ma):
        t.columns = [str(c).strip() for c in t.columns]
    mir = mirror_features(raw, (yd, sd, ma))
    print(f'\nformula validation vs the munged frame '
          f'({len(frame)} rows, reconstructed pre-QC-review inputs):')
    print(f'  {"feature":32s} {"all rows":>9s}')
    worst = []
    for c in feat_names:
        if c not in mir.columns or c not in frame.columns:
            continue
        if c == 'hh_group':
            ok = mir[c].astype(str).values == frame[c].astype(str).values
        else:
            a = mir[c].astype(float).values
            # over_threshold / total_error_amount live in the workbook in
            # build_workbook_v2's transformed form; compare against that
            if c == 'over_threshold':
                b = frame['is_error'].astype(float).values
            elif c == 'total_error_amount':
                b = np.round(np.abs(pd.to_numeric(
                    frame[c], errors='coerce').fillna(0).values))
            else:
                b = pd.to_numeric(frame[c], errors='coerce').fillna(0).values
            ok = np.isclose(a, b, atol=1e-6, rtol=1e-9)
        r_all = ok.mean()
        print(f'  {c:32s} {r_all:9.1%}')
        worst.append((c, r_all))
    bad = [c for c, r in worst if r < 0.995]
    if bad:
        print(f'  WARNING: match below 99.5% for: {", ".join(bad)}')
    return not bad


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('live_workbook')
    ap.add_argument('-o', '--out', required=True)
    ap.add_argument('--state', default=os.environ.get('SNAP_STATE', 'WA'))
    a = ap.parse_args()
    cfg = STATE_REGISTRY.get(a.state.upper())
    frame_csv = os.path.join(PKG, '.frames', f'{cfg["abbr"].lower()}_frame.csv')

    raw, elem_free, frame = raw_frame(cfg, frame_csv)
    print(f'raw extract: {len(raw)} rows x {len(RAW_COLS)} input fields')

    shutil.copy(a.live_workbook, a.out)
    wb = openpyxl.load_workbook(a.out)
    # the Dashboard (per-rule threshold tuner) stays hidden: engine plumbing;
    # unhide it in Excel to tune thresholds interactively
    dat = wb[DATA_SHEET]
    NROW = dat.max_row
    assert NROW - 1 == len(raw), f'workbook has {NROW-1} cases, extract {len(raw)}'

    hdr = [c.value for c in next(dat.iter_rows(min_row=1, max_row=1))]
    ncol0 = len(hdr)
    R = federal_tables(wb, cfg['name'])
    helpers, feats = feature_formulas(R)

    missing = [h for h in hdr
               if not h.startswith('_') and h not in feats and h not in RAW_COLS]
    assert not missing, f'no formula for Data columns: {missing}'
    # Excel table column names are case-insensitively unique; a collision makes
    # Excel reject the whole file as damaged, so fail here instead
    # (RAW_COLS are already in hdr: the build stage writes them on the left)
    names = hdr + [h for h, _ in helpers]
    low = [n.lower() for n in names]
    dups = sorted({n for n in low if low.count(n) > 1})
    assert not dups, f'case-insensitive duplicate Data columns: {dups}'
    ok = validate(raw, frame, elem_free, [h for h in hdr if h in feats])
    if not ok:
        # HARD GATE (2026-08-16): the demo must sit on the reconstructed
        # pre-QC-review scale the rules were mined on — a state's internal
        # data is effectively that scale, and scoring rules against
        # QC-corrected values misleads (v1 post-mortem: 21 of 114 WA rules
        # never fired). With reconstructed inputs the match is 100%; any
        # shortfall means corrected values crept back into the input block.
        if os.environ.get('SNAP_ALLOW_VALIDATION_MISMATCH') == '1':
            print('  SNAP_ALLOW_VALIDATION_MISMATCH=1: continuing despite '
                  'validation failures (debug only)')
        else:
            raise SystemExit(
                'validation failed: formula features do not reproduce the '
                'research frame. The Data-tab demo must carry RECONSTRUCTED '
                '(pre-QC-review) values, never QC-corrected ones — see the '
                'columns flagged above. Set SNAP_ALLOW_VALIDATION_MISMATCH=1 '
                'only to debug.')

    # 1. the raw contract values go into the LEFT block (headers written by
    #    build_workbook_v2 from the shared layout); the helper columns are
    #    appended AFTER the feature + hit columns so every positional
    #    reference stays valid
    for j, name in enumerate(RAW_COLS, 1):
        assert dat.cell(row=1, column=j).value == name, (
            f'Data column {j} is {dat.cell(row=1, column=j).value!r}, expected '
            f'{name!r}: layout drift against workbook_layout.py')
        vals = raw[name].tolist()
        for i, v in enumerate(vals):
            dat.cell(row=2 + i, column=j,
                     value=None if (isinstance(v, float) and np.isnan(v)) else v)
    c = ncol0
    helper_idx = []
    for name, formula in helpers:
        c += 1
        helper_idx.append(c)
        dat.cell(row=1, column=c, value=name)
        for rr in range(2, NROW + 1):
            dat.cell(row=rr, column=c, value=formula)
    LAST = CL(c)

    # 2. the feature columns become formulas (values overwritten in place)
    n_swapped = 0
    for ci, name in enumerate(hdr, 1):
        if name in feats:
            for rr in range(2, NROW + 1):
                dat.cell(row=rr, column=ci, value=feats[name])
            n_swapped += 1
    print(f'features -> formulas: {n_swapped} columns | helpers: {len(helpers)} | '
          f'raw: {len(RAW_COLS)} -> Data!A1:{LAST}{NROW}')

    # 3. rebind the table over the widened range (with calculated-column
    #    declarations so pasted rows fill the hidden formula columns)
    for t in list(dat.tables):
        del dat.tables[t]
    make_table(dat, TABLE, f'A1:{LAST}{NROW}')
    n_modes = write_sua_mode_block(wb, R, NROW)
    print(f'SUA mode block: {n_modes} array-entered per-year cells on FederalTables')

    # 4. presentation: inputs (left block) light yellow — the workbook-wide
    #    "yellow means interactive" convention (revision 2026-08-18); computed
    #    features gray; helpers hidden
    feat_fill = PatternFill('solid', fgColor='D9D9D9')
    raw_fill = PatternFill('solid', fgColor='FFFF99')
    for i in range(1, len(RAW_COLS) + 1):
        h = dat.cell(row=1, column=i)
        h.fill = raw_fill
        h.font = Font(bold=True)
    for ci, name in enumerate(hdr, 1):
        if name in feats:
            h = dat.cell(row=1, column=ci)
            h.fill = feat_fill
            h.font = Font(bold=True)
            for rr in range(2, NROW + 1):
                dat.cell(row=rr, column=ci).fill = feat_fill
    for i in helper_idx:
        dat.column_dimensions[CL(i)].hidden = True
    # collapse button sits left of the group (summary-left): the input block
    # and the computed block each collapse independently
    dat.sheet_properties.outlinePr.summaryRight = False
    for i in range(3, len(RAW_COLS) + 1):          # keep CASE_ID + year visible
        dat.column_dimensions[CL(i)].outline_level = 1
    for ci, name in enumerate(hdr, 1):
        if name in feats and name != 'fiscal_year':
            dat.column_dimensions[CL(ci)].outline_level = 1
    note = ('Yellow input columns (this left block) are what a state supplies, '
            'mapping its own reported case fields onto them. The gray columns '
            'to the right are FORMULAS computed from the inputs; do not paste '
            'values over them. A blank cell means MISSING, not zero: cases '
            'missing an input a rule needs are not flagged by that rule, and '
            'a column you do not collect can be left entirely empty. Enter '
            'zeros as zeros. See the Data Dictionary tab for every definition '
            'and the QC technical-manual crosswalk.')
    dat.cell(row=1, column=1).comment = Comment(note, 'snap_dashboard')
    dat.freeze_panes = 'B2'

    step3_rows = data_dictionary(wb, hdr)
    # Step 3 column headers hyperlink to their dictionary rows (feedback
    # 2026-08-22); the header row is row 4 on the rules tab
    rules_ws = wb[BLENDED_SHEET]
    linked = 0
    for ci in range(1, rules_ws.max_column + 1):
        h = rules_ws.cell(row=4, column=ci)
        if h.value in step3_rows:
            h.hyperlink = f"#'{DICT_SHEET}'!A{step3_rows[h.value]}"
            h.font = Font(name='Calibri', bold=True, size=10, color='0563C1',
                          underline='single')
            linked += 1
    print(f'Step 3 headers linked to the dictionary: {linked}')
    background_tab(wb, cfg['name'])
    n51 = screening_tabs(wb, R, hdr)
    n6 = share_tab(wb, cfg['name'])
    print(f'screening tabs: {n51} rule columns | share-back rows: {n6}')

    # (the viewer's CASE_ID and grid formulas are written by the build stage
    # since 2026-08-18: they read the Data sheet's _view_cum column, which
    # lives inside the CaseData table and auto-extends over pasted rows)

    # tab colors (Eric's WA scheme, 2026-08-19): orange = the core path
    # (Steps 1-3), pale orange = the optional post-Step-3 tabs, dark blue =
    # Start Here and the viewer (set where those sheets are created)
    for name in (DICT_SHEET, DATA_SHEET, BLENDED_SHEET):
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = 'C55A11'
    for name in (EXPORT_SHEET, SCREEN_SHEET, FLAGGED_SHEET, SHARE_SHEET):
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = 'FDE9D9'

    # sheet order: Start Here, then the step tabs in order, then everything else
    head_names = ['Start Here', DICT_SHEET, DATA_SHEET, BLENDED_SHEET,
                  EXPORT_SHEET, SCREEN_SHEET, FLAGGED_SHEET, SHARE_SHEET]
    head = [wb[n] for n in head_names if n in wb.sheetnames]
    wb._sheets = head + [s for s in wb._sheets if s.title not in head_names]
    wb.active = 0

    wb.save(a.out)
    print('saved', a.out)


if __name__ == '__main__':
    main()
