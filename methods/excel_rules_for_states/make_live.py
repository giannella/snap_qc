"""
Turn a built workbook into a LIVE one: every figure recomputes from the Data tab.

What changes
  Data          becomes an Excel Table (sheet keeps its name), so pasting more
                rows extends every reference automatically, and columns are
                bound by NAME rather than by position.
  Rules tab     per-rule metrics and the combined rows become COUNTIFS/SUMIFS
                over that table; the hard-coded denominators go live too.
                Per-rule Recall and $ Recall divide by ALL pasted errors /
                error dollars (grand totals, decision 2026-08-18), matching
                build_workbook_v2.score_list; the orange combined rows keep
                their per-stratum denominators.
  Dashboard     its fixed Data!$X$2:$X$3000 ranges become table references.

Usage:  python make_live.py <built_workbook.xlsx> [-o out.xlsx]

Run postprocess_workbook.py afterwards to drop the stale calc chain.
"""
import argparse
import os
import re
import shutil

import openpyxl
from openpyxl.utils import get_column_letter as CL
from openpyxl.utils import column_index_from_string

from workbook_layout import DATA_SHEET, BLENDED_SHEET, qref
# the rule-to-formula translation is shared with make_input_workbook's
# Step 5/6 tabs (extracted 2026-08-18)
from live_formulas import (countifs, make_table, pack_chunks,
                           read_delivery_tab, rule_terms, selection_refs)

TABLE = 'CaseData'            # table object name; the sheet keeps its name


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('workbook')
    ap.add_argument('-o', '--out')
    a = ap.parse_args()
    out = a.out or a.workbook.replace('.xlsx', '_LIVE.xlsx')
    shutil.copy(a.workbook, out)

    wb = openpyxl.load_workbook(out)
    dat = wb[DATA_SHEET]
    nat = wb[BLENDED_SHEET] if BLENDED_SHEET in wb.sheetnames else None
    hdr = [c.value for c in next(dat.iter_rows(min_row=1, max_row=1))]
    NROW, NCOL = dat.max_row, len(hdr)
    print(f'Data: {NROW-1} cases x {NCOL} columns')

    # ── 1. read the rules straight out of the Exact expression column ────────
    N0 = 11
    nt_rules = read_delivery_tab(nat, first_row=N0)
    print(f'rules: {len(nt_rules)}')

    # ── 2. hidden per-case hit columns, inside the table so they auto-fill ───
    sel_by_row = selection_refs(wb, BLENDED_SHEET)
    print('selection columns found: %d' % len(sel_by_row))
    sel_n = lambda i: sel_by_row[N0 + i]
    blocks = [('nat', nt_rules, sel_n)] if nat else []
    newcols = {}
    c = NCOL
    for tag, rules, sel in blocks:
        terms = rule_terms(rules, sel)
        parts = []
        for k, chunk in enumerate(pack_chunks(terms)):
            c += 1
            name = f'_{tag}{k+1}'
            newcols[name] = c
            dat.cell(row=1, column=c, value=name)
            f = '=' + ('+'.join(chunk) if chunk else '0')
            assert len(f) < 8000, f'{name}: formula {len(f)} chars, too long'
            for rr in range(2, NROW + 1):
                dat.cell(row=rr, column=c, value=f)
            parts.append(name)
        c += 1
        hit = f'_hit_{tag}'
        newcols[hit] = c
        dat.cell(row=1, column=c, value=hit)
        expr = '+'.join(f'{TABLE}[[#This Row],[{p}]]' for p in parts)
        for rr in range(2, NROW + 1):
            dat.cell(row=rr, column=c, value=f'=IF({expr}>0,1,0)')
    LAST = CL(c)
    print(f'added {len(newcols)} computed columns -> Data!A1:{LAST}{NROW}')

    make_table(dat, TABLE, f'A1:{LAST}{NROW}')
    for name, idx in newcols.items():
        dat.column_dimensions[CL(idx)].hidden = True

    # ── 3. live denominators ─────────────────────────────────────────────────
    def denom(hh, what):
        if what == 'cases':
            return f'COUNTIF({TABLE}[hh_group],"{hh}")' if hh != 'all' \
                else f'COUNTA({TABLE}[hh_group])'
        if what == 'errors':
            return (f'COUNTIFS({TABLE}[hh_group],"{hh}",{TABLE}[over_threshold],1)'
                    if hh != 'all' else f'COUNTIF({TABLE}[over_threshold],1)')
        return (f'SUMIFS({TABLE}[total_error_amount],{TABLE}[hh_group],"{hh}",'
                f'{TABLE}[over_threshold],1)' if hh != 'all'
                else f'SUMIFS({TABLE}[total_error_amount],{TABLE}[over_threshold],1)')

    HITS = {'nat': f'{TABLE}[_hit_nat]'}
    OV, AMT = f'{TABLE}[over_threshold]', f'{TABLE}[total_error_amount]'

    def combined(ws, row, hh, tag, cols):
        """cols maps logical name -> column index on that sheet."""
        hitp = HITS[tag]
        st = '' if hh == 'all' else f'*({TABLE}[hh_group]="{hh}")'
        F, E, D = (CL(cols['flagged']), CL(cols['errors']), CL(cols['dollars']))
        put = lambda k, f: ws.cell(row=row, column=cols[k], value='=' + f)
        put('flagged', f'SUMPRODUCT(({hitp}){st})')
        put('errors',  f'SUMPRODUCT(({hitp}){st}*({OV}))')
        put('dollars', f'SUMPRODUCT(({hitp}){st}*({OV})*({AMT}))')
        put('prec',    f'IF(${F}{row}=0,0,${E}{row}/${F}{row})')
        put('rec',     f'IFERROR(${E}{row}/{denom(hh,"errors")},0)')
        put('drec',    f'IFERROR(${D}{row}/{denom(hh,"dollars")},0)')
        # summary-row workload: share of ALL cases, so the household-size
        # rows are the percentage-point portions that sum to the 'all' row
        put('work',    f'IFERROR(${F}{row}/{denom("all","cases")},0)')
        put('exp',     f'IFERROR(IF(${F}{row}=0,"",${D}{row}/${F}{row}),"")')

    NT = dict(prec=4, rec=5, drec=6, flagged=7, errors=8, dollars=9, work=10, exp=11)
    for i, hh in enumerate(['all', '1', '2-3', '4+']):
        if nat:
            combined(nat, 6 + i, hh, 'nat', NT)

    # ── 4. per-rule rows ─────────────────────────────────────────────────────
    def rule_row(ws, row, conds, hh, cols):
        if conds is None:
            return
        F, E, D = (CL(cols['flagged']), CL(cols['errors']), CL(cols['dollars']))
        n = countifs(conds, hh)
        e = countifs(conds, hh, extra=f'{TABLE}[over_threshold],1')
        d = countifs(conds, hh, extra=f'{TABLE}[over_threshold],1',
                     col='total_error_amount', fn='SUMIFS')
        put = lambda k, f: ws.cell(row=row, column=cols[k], value='=' + f)
        put('flagged', n); put('errors', e); put('dollars', d)
        # below 10 flagged cases a percentage misleads; show the counts
        put('prec', f'IF(${F}{row}<10,${E}{row}&" errors of "&${F}{row}'
                    f'&" cases flagged",IFERROR(${E}{row}/${F}{row},0))')
        # per-rule Recall / $ Recall: this rule ALONE over ALL pasted errors /
        # error dollars (grand totals; must match build_workbook_v2.score_list)
        put('rec',  f'IFERROR(${E}{row}/{denom("all","errors")},0)')
        put('drec', f'IFERROR(${D}{row}/{denom("all","dollars")},0)')
        put('work', f'IFERROR(${F}{row}/{denom(hh,"cases")},0)')
        put('exp',  f'IFERROR(IF(${F}{row}=0,"",${D}{row}/${F}{row}),"")')

    for i, (conds, hh) in enumerate(nt_rules):
        rule_row(nat, N0 + i, conds, hh, NT)

    # ── 5. Dashboard: fixed ranges -> table references ───────────────────────
    swapped = 0
    dq = qref(DATA_SHEET)
    rng_pat = re.compile(re.escape(dq) + r'!\$([A-Z]{1,3})\$\d+:\$[A-Z]{1,3}\$\d+')
    for row in wb['Dashboard'].iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and f'{dq}!$' in v:
                def sub(m):
                    return f'{TABLE}[{hdr[column_index_from_string(m.group(1)) - 1]}]'
                nv = rng_pat.sub(sub, v)
                if nv != v:
                    cell.value = nv; swapped += 1
    print(f'Dashboard: {swapped} formulas rebound to the table')

    wb.save(out)
    print('saved', out)
    print('\nnow run:  python postprocess_workbook.py "%s"' % out)


if __name__ == '__main__':
    main()
