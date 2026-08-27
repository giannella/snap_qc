"""
Build a state's SNAP QC rules workbook end to end.

    python make_state.py                 # default state (WA)
    python make_state.py VA              # any state with a delivery list
    python make_state.py all             # every state with a delivery list
    python make_state.py WA --refresh    # re-export the frame from the rds first

The deliverable is ONE file per state, state_workbooks/<ABBR>/
SNAP_flagging_rules_<ABBR>.xlsx: raw input fields as values, every model
feature an in-workbook formula, and the rules tab (the blended delivery
list after the rule_selection.py transform). No tuning of any kind runs in
the workbook. The intermediate stages land in .build/out_<ABBR>/ and are
kept for crosscheck_rules.py, which reads the plain build's static values
against .build/effective_rules_<ABBR>.csv.

(--v2, --live and --recon are accepted as no-ops for old command lines: the
full chain is always run, and the v1 builder is retired to
custom_one_off/legacy_dashboard.)

Stages:
  1. build_workbook_v2.py   every sheet, formula and value  (.build)
  2. make_live.py           Data becomes an Excel table; the rules tabs
                            recompute from it, so pasted rows flow through
  3. make_input_workbook.py  the Data tab gains the input contract (values)
                            and every model feature becomes an in-workbook
                            formula, for states that can run neither Python
                            nor R; adds Start Here, Data Dictionary and
                            FederalTables, and runs the validation gate
  4. postprocess_workbook.py   calc-chain removal (checkboxes retired 2026-08-18)
  5. verify_workbook.py / verify_workbook_win.ps1   open-in-Excel probe

Close the workbook in Excel first: writing underneath an open session produces
OneDrive conflict copies and can clobber the new file.
"""
import json
import os
import shutil
import subprocess
import sys

PKG = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PKG)
from workbook_layout import (BLENDED_SHEET, DATA_SHEET, EXPORT_SHEET,  # noqa: E402
                             FLAGGED_SHEET, SHARE_SHEET, VIEWER_SHEET)

BUILD_DIR = os.path.join(PKG, '.build')
PY = sys.executable
PROBES = [f"{BLENDED_SHEET}!C5", f"{BLENDED_SHEET}!G5", f"{BLENDED_SHEET}!H5",
          f"{VIEWER_SHEET}!A1", f"{VIEWER_SHEET}!A4",
          f"{EXPORT_SHEET}!A4"]
# the delivered (stage-3) workbook additionally carries Start Here and the
# Step 5/6 tabs; probe the KPIs (base error rate row 7, overissuance share
# row 9 — the KPI block starts at row 3), the pair counter, and a Step 6
# per-rule metric cell (first rule row, Flagged column)
PROBES_FINAL = PROBES + ["Start Here!B7", "Start Here!B9",
                         f"{FLAGGED_SHEET}!B3", f"{SHARE_SHEET}!D14"]


def run(script, *args, env=None):
    print(f'\n$ {os.path.basename(script)} {" ".join(args)}')
    p = subprocess.run([PY, os.path.join(PKG, script), *args], cwd=PKG,
                       env={**os.environ, **(env or {})})
    if p.returncode != 0:
        raise SystemExit(f'stage failed: {script}')


def in_use(path):
    if sys.platform == 'win32':
        # no lsof on Windows: an exclusive-open probe catches an Excel lock
        try:
            with open(path, 'r+b'):
                return False
        except FileNotFoundError:
            return False
        except OSError:
            return True
    return subprocess.run(['lsof', path], capture_output=True).returncode == 0


def verify(wbk, probes=PROBES):
    if sys.platform == 'darwin':
        run('verify_workbook.py', wbk, *probes)
    elif sys.platform == 'win32':
        print(f'\n$ verify_workbook_win.ps1 {os.path.basename(wbk)}')
        p = subprocess.run(
            ['powershell', '-NoProfile', '-ExecutionPolicy', 'Bypass',
             '-File', os.path.join(PKG, 'verify_workbook_win.ps1'),
             wbk] + probes, cwd=PKG)
        if p.returncode != 0:
            raise SystemExit(f'verify failed: {wbk}')
    else:
        print('verify stage skipped (no Excel automation on this platform); '
              'open the workbook once in Excel before shipping it')


def build_one(state, refresh, want_verify):
    env = {'SNAP_STATE': state,
           'SNAP_OUT_DIR': os.environ.get('SNAP_OUT_DIR')
           or os.path.join(BUILD_DIR, f'out_{state}')}
    if refresh:
        env['SNAP_REFRESH_FRAME'] = '1'
    run('build_workbook_v2.py', env=env)

    out = json.load(open(os.path.join(BUILD_DIR, 'target.json')))['out']
    live = out.replace('.xlsx', '_LIVE.xlsx')
    recon = out.replace('.xlsx', '_RECON.xlsx')
    run('make_live.py', out, '-o', live)
    run('make_input_workbook.py', live, '-o', recon, '--state', state)

    ck = os.path.join(BUILD_DIR, 'checkbox_cells.json')
    for wbk in (out, live, recon):
        run('postprocess_workbook.py', wbk, ck)

    deliver_dir = os.path.join(PKG, 'state_workbooks', state)
    os.makedirs(deliver_dir, exist_ok=True)
    # delivery filename convention agreed 2026-08-21
    deliver = os.path.join(deliver_dir, f'SNAP_flagging_rules_{state}.xlsx')
    if in_use(deliver):
        raise SystemExit(f'{deliver} is open in Excel — close it and re-run.')
    shutil.copy(recon, deliver)
    # drop leftovers from the era when all three variants were delivered
    for stale in (f'snap_qc_dashboard_{state}_LIVE.xlsx',
                  f'snap_qc_dashboard_{state}_RECON.xlsx',
                  f'snap_qc_dashboard_{state}.xlsx',   # pre-2026-08-21 name
                  f'{state.lower()}_cases.csv'):
        p = os.path.join(deliver_dir, stale)
        if os.path.isfile(p) and not in_use(p):
            os.remove(p)

    if want_verify:
        for wbk in (out, live):
            verify(wbk)
        # the delivered file's LIVE union figures must equal the plain
        # build's STATIC ones: a formula that silently collapses
        # (2026-08-22: an in-row array evaluated to 0 everywhere, with no
        # error cell) shows up here as a union count that drifts. The static
        # union is recomputed from the plain build's RuleFlags 0/1 hit
        # matrix (every rule selected) and the Data tab's static
        # over_threshold column, the same quantities crosscheck_rules.py
        # verifies against R, so no Excel evaluation is involved.
        g5, h5 = static_union(out, state)
        expect = [f"{BLENDED_SHEET}!G5={g5}", f"{BLENDED_SHEET}!H5={h5}"]
        probes = [p for p in PROBES_FINAL
                  if p not in (f"{BLENDED_SHEET}!G5", f"{BLENDED_SHEET}!H5")] + expect
        verify(deliver, probes)
    print(f'\nDone: {deliver}')


def static_union(plain_xlsx, state=None):
    """(cases flagged, errors caught) by the union of the SHIPPED rules, from
    the plain build's static RuleFlags hit matrix (rows FLAG0.., one 0/1
    column per rule starting at column B) and the Data tab's over_threshold.
    Measurement-tier rules (ship = FALSE in the effective list) are excluded,
    matching the live union's Include? defaults."""
    import openpyxl
    wb = openpyxl.load_workbook(plain_xlsx, read_only=True)
    rf = wb['RuleFlags']
    nr = sum(1 for c in next(rf.iter_rows(min_row=1, max_row=1))[1:]
             if isinstance(c.value, int))
    # RuleFlags column j (1-based after the lead column) is effective-list
    # rule j; the ship column says which columns count toward the deployed
    # union. No ship column (or no csv) means every rule ships.
    keep = [True] * nr
    if state:
        eff = os.path.join(BUILD_DIR, f'effective_rules_{state}.csv')
        if os.path.isfile(eff):
            import csv as _csv
            with open(eff, newline='', encoding='utf-8') as fh:
                rows = list(_csv.DictReader(fh))
            if rows and 'ship' in rows[0]:
                keep = [str(r.get('ship', 'True')).strip().lower()
                        not in ('false', '0') for r in rows[:nr]]
    dat = wb[DATA_SHEET]
    hdr = [c.value for c in next(dat.iter_rows(min_row=1, max_row=1))]
    ov = hdr.index('over_threshold')
    err = [row[ov] for row in dat.iter_rows(min_row=2, values_only=True)]
    flagged = errors = 0
    for i, row in enumerate(rf.iter_rows(min_row=5, max_col=1 + nr, values_only=True)):
        if i >= len(err):
            break
        if any(v == 1 for v, k in zip(row[1:], keep) if k):
            flagged += 1
            if err[i] == 1:
                errors += 1
    return flagged, errors


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('-')]
    target = (args[0] if args else 'WA').upper()
    refresh = '--refresh' in sys.argv
    want_verify = '--no-verify' not in sys.argv

    if target == 'ALL':
        import states as STATE_REGISTRY
        todo = STATE_REGISTRY.all_abbrs()
        print(f'building {len(todo)} states: {" ".join(todo)}')
        failed = []
        for st in todo:
            try:
                build_one(st, refresh, want_verify)
            except SystemExit as e:
                print(f'FAILED {st}: {e}')
                failed.append(st)
        if failed:
            raise SystemExit(f'{len(failed)} state(s) failed: {" ".join(failed)}')
    else:
        build_one(target, refresh, want_verify)


if __name__ == '__main__':
    main()
