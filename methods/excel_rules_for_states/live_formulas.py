"""Shared Excel-formula builders for the live stages (extracted from
make_live.py 2026-08-18 so make_input_workbook's Step 5/6 tabs reuse the SAME
rule-to-formula translation instead of re-implementing it).

Everything here is string assembly over an Excel table's structured
references; nothing touches a workbook.
"""
import re

from openpyxl.utils import get_column_letter as CL

CHUNK_LIMIT = 7500            # max chars per hit-column formula (Excel cap 8192)
COND = re.compile(r'([A-Za-z_][A-Za-z0-9_]*)\s*(>=|<=|>|<|=)\s*(-?[0-9.]+)')


def parse(text):
    """'a > 1 & b <= 2' -> [('a','>','1'), ('b','<=','2')]  (None if unusable)"""
    if (not text or 'no combination' in str(text) or 'not selected' in str(text)
            or 'not deployed' in str(text)):
        return None
    out = COND.findall(str(text))
    return out or None


def countifs(conds, hh, extra='', col=None, fn='COUNTIFS', table='CaseData'):
    """A COUNTIFS/SUMIFS over the table for one rule."""
    head = f'{table}[{col}],' if fn == 'SUMIFS' else ''
    parts = [f'{table}[hh_group],"{hh}"']
    for v, op, t in conds:
        parts.append(f'{table}[{v}],"{op if op != "=" else ""}{t}"')
    if extra:
        parts.append(extra)
    return f'{fn}({head}{",".join(parts)})'


def rule_term(conds, hh, table='CaseData'):
    """One rule's per-case test (product of guards), no selection weight.

    Each numeric condition carries an ISNUMBER guard: when an input column is
    deleted the features built from it go blank (make_input_workbook gates
    them), and a blank must never satisfy a condition — Excel otherwise ranks
    any text above any number, so ``"" >= x`` would flag every case."""
    t = [f'({table}[[#This Row],[hh_group]]="{hh}")']
    for v, op, val in conds:
        ref = f'{table}[[#This Row],[{v}]]'
        t.append(f'ISNUMBER({ref})*({ref}{"=" if op == "=" else op}{val})')
    return '*'.join(t)


def rule_terms(rows, sel_ref, table='CaseData'):
    """Per-case test terms, one per rule, weighted by that rule's Include? flag."""
    terms = []
    for i, (conds, hh) in enumerate(rows):
        if conds is None:
            continue
        terms.append(f'{sel_ref(i)}*' + rule_term(conds, hh, table))
    return terms


def pack_chunks(terms, limit=CHUNK_LIMIT):
    """Split the per-rule terms into as many helper-column formulas as the
    formula length cap requires."""
    chunks, cur, size = [], [], 1
    for t in terms:
        if cur and size + len(t) + 1 > limit:
            chunks.append(cur)
            cur, size = [], 1
        cur.append(t)
        size += len(t) + 1
    if cur:
        chunks.append(cur)
    return chunks or [[]]


def read_delivery_tab(ws, first_row=11):
    """(conds, hh) per rule row, read from a rules tab's Exact-expression
    column (found by header rather than assumed position)."""
    if ws is None:
        return []
    expr_col = next((c for c in range(1, ws.max_column + 1)
                     if ws.cell(row=4, column=c).value == 'Exact expression'),
                    13)
    out, r = [], first_row
    while ws.cell(row=r, column=1).value:
        out.append((parse(ws.cell(row=r, column=expr_col).value),
                    ws.cell(row=r, column=2).value))
        r += 1
    return out


def make_table(ws, name, ref, formula_row=2):
    """Create an Excel Table whose formula columns carry explicit
    <calculatedColumnFormula> declarations (read from `formula_row`, the top
    data row), so Excel fills them into rows added by a paste. Without the
    declaration Excel's fill-on-expand heuristic skips HIDDEN columns — seen
    2026-08-18: a paste into Step 5.1 extended the table and filled the
    visible feature columns but left the hidden _r*/_hits/_cum columns empty
    past the seeded row, so no case was ever flagged."""
    from openpyxl.utils.cell import range_boundaries
    from openpyxl.worksheet.table import (Table, TableColumn, TableFormula,
                                          TableStyleInfo)
    min_col, min_row, max_col, _ = range_boundaries(ref)
    cols = []
    for i, c in enumerate(range(min_col, max_col + 1), 1):
        tc = TableColumn(id=i, name=str(ws.cell(row=min_row, column=c).value))
        v = ws.cell(row=formula_row, column=c).value
        if isinstance(v, str) and v.startswith('='):
            tc.calculatedColumnFormula = TableFormula(attr_text=v[1:])
        cols.append(tc)
    tbl = Table(displayName=name, ref=ref, tableColumns=cols)
    tbl.tableStyleInfo = TableStyleInfo(name='TableStyleLight1',
                                        showRowStripes=False)
    ws.add_table(tbl)
    return tbl


def selection_refs(wb, sheet_name):
    """{rules-tab row -> 'RuleFlags!$X$2'}: which RuleFlags selection cell
    watches each rule row. The selection vector's column order is NOT the
    order rules appear on the sheet, so read it out of RuleFlags row 2."""
    rf = wb['RuleFlags']
    out = {}
    for cc in range(2, rf.max_column + 1):
        v = rf.cell(row=2, column=cc).value
        if not isinstance(v, str):
            continue
        m = re.search(r"(?:'([^']+)'|([A-Za-z_]\w*))!\$[A-Z]+\$(\d+)", v)
        if not m:
            continue
        if (m.group(1) or m.group(2)) == sheet_name:
            out[int(m.group(3))] = f'RuleFlags!${CL(cc)}$2'
    return out
